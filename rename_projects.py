"""Standardize project names in Invoice details sheet."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))
import openpyxl

DATA_FILE = Path(__file__).parent / "data" / "CaddyCheckProjectsInfo.xlsx"

# {old_name_lower: canonical_name}  — case-insensitive match, exact canonical output
RENAMES = {
    # Proxy Denderleew (spelling fix; Delhaize prefix removed by general rule below)
    "proxy delhaize denderleeuw":               "Proxy Denderleew",
    "proxy delhaize denderleew":                "Proxy Denderleew",

    # Proxy Denderhoutem
    "proxy delhaize denderhoutem (back)":       "Proxy Denderhoutem",
    "proxy delhaize denderhoutem (top)":        "Proxy Denderhoutem",
    "proxy delhaize denderhoutem":              "Proxy Denderhoutem",

    # Proxy Berchem (explicit for canonical casing)
    "proxy delhaize berchem":                   "Proxy Berchem",
    "proxy berchem":                            "Proxy Berchem",

    # Proxy Ooigem (fix ALL CAPS)
    "proxy delhaize ooigem":                    "Proxy Ooigem",
    "proxy ooigem":                             "Proxy Ooigem",

    # Spar Dadizele
    "spar dadizele":                            "Spar Dadizele",
    "spar dadizele (2nd)":                      "Spar Dadizele",

    # Spar Denderleew (vim3 variant)
    "spar denderleew (vim3)":                   "Spar Denderleew",

    # Plus Simpelveld Schouteten (multiple spelling variants)
    "plus simpelveld schouteten":               "Plus Simpelveld Schouteten",
    "plus simpelveld schouteden":               "Plus Simpelveld Schouteten",
    "plus simpleveld schouteden":               "Plus Simpelveld Schouteten",

    # Plus Klazienaveen Fischer
    "plus klazienaveen":                        "Plus Klazienaveen Fischer",
    "plus klazienaveen fischer":                "Plus Klazienaveen Fischer",

    # Plus Kamerik Romijn (case fix)
    "plus kamerik romijn":                      "Plus Kamerik Romijn",

    # Plus Landgraaf Arts (case fix)
    "plus landgraaf arts":                      "Plus Landgraaf Arts",

    # AH Merksem MAATJES (strip _x000D_ / \r)
    "ah merksem maatjes":                       "AH Merksem MAATJES",

    # AH Dendermonde (typo fix)
    "ah dendemonde":                            "AH Dendermonde",

    # AD Geraardsbergen (_x000D_ suffix)
    "ad geraardsbergen":                        "AD Geraardsbergen",

    # AD Aartselar
    "ad aartselar (1 top)":                     "AD Aartselar",
    "ad aartselar (4 back)":                    "AD Aartselar",
    "ad aartselar (4 top)":                     "AD Aartselar",
    "ad delhaize aartselaar":                   "AD Aartselar",

    # AD Arendonk (ALL CAPS fix)
    "ad arendonk":                              "AD Arendonk",

    # AD Soumagne (spelling variants)
    "ad soumagane":                             "AD Soumagne",
    "ad soumagune":                             "AD Soumagne",

    # AD Denderleeuw (spelling fix)
    "ad denderleew":                            "AD Denderleeuw",

    # Alvo Kontich (typo fix)
    "alvo konitch":                             "Alvo Kontich",

    # CM Alma Mol (case fix)
    "cm alma mol":                              "CM Alma Mol",

    # CM Waregem (typo fix)
    "cm waregram":                              "CM Waregem",

    # Coop Kockengen (spelling fix)
    "coop cockengen":                           "Coop Kockengen",

    # Edeka Dortmund (ALL CAPS fix)
    "edeka dortmund":                           "Edeka Dortmund",

    # Delhaize Belval Luxemburg (odd capitalisation)
    "delhaize belval luxemburg":                "Delhaize Belval Luxemburg",

    # Edeka Vogel-Lewandowski Dueren
    "edeka vogel dueren":                       "Edeka Vogel-Lewandowski Dueren",

    # Jumbo Groningen Pestman (case fix)
    "jumbo groningen pestman":                  "Jumbo Groningen Pestman",

    # Jumbo Olm Krimpen aan den Ijssel (ALL CAPS fix)
    "jumbo olm krimpen aan den ijssel":         "Jumbo Olm Krimpen aan den Ijssel",

    # Jumbo Gravendeel Van Der Hoek
    "jumbo 's gravendeel van der hoek":         "Jumbo Gravendeel Van Der Hoek",
    "jumbo gravendeel":                         "Jumbo Gravendeel Van Der Hoek",

    # Coop Bert Stuut
    "coop bert stuut (back)":                   "Coop Bert Stuut",
    "coop bert stuut (top)":                    "Coop Bert Stuut",

    # Jumbo Eindhoven Boschdijk
    "jumbo eindhoven bosdijk":                  "Jumbo Eindhoven Boschdijk",
    "jumbo eindhoven bosdijk (from side to td)": "Jumbo Eindhoven Boschdijk",
    "jumbo eindhoven boschdijk":                "Jumbo Eindhoven Boschdijk",

    # Eurospar (fix EuroSpar capitalisation; merge (2nd) variant)
    "eurospar dadizele":                        "Eurospar Dadizele",
    "eurospar dadizele (2nd)":                  "Eurospar Dadizele",
    "eurospar denderleew":                      "Eurospar Denderleew",

    # Delhaize Sint-Lievens-Houtem → Proxy Sint-Lievens-Houtem
    "delhaize sint-lievens-houtem":             "Proxy Sint-Lievens-Houtem",

    # Plus langraaf Aarts extra checkouts → Plus Landgraaf Arts
    "plus langraaf aarts extra checkouts":      "Plus Landgraaf Arts",

    # Alma Geel → CM Alma Geel
    "alma geel":                                "CM Alma Geel",

    # Plus Delft
    "plus delft (1 top)":                       "Plus Delft",
    "plus delft (4 back)":                      "Plus Delft",

    # Plus Maastricht Caberg
    "plus maastricht caberg":                   "Plus Maastricht Caberg",
    "plus maastricht caberg  (from back to td)": "Plus Maastricht Caberg",

    # Plus Meerkerk
    "plus meerkerk (back)":                     "Plus Meerkerk",
    "plus meerkerk (top)":                      "Plus Meerkerk",

    # Jumbo Uden De Laak
    "jumbo uden de laak (back)":               "Jumbo Uden De Laak",
    "jumbo uden de laak (from back to td)":    "Jumbo Uden De Laak",
}

wb = openpyxl.load_workbook(DATA_FILE)

total = 0
for sheet_name in ["Invoice details", "Projects overview"]:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    rows_list = list(ws.iter_rows(values_only=False))
    header = [str(c.value).strip().lower() if c.value else "" for c in rows_list[0]]
    proj_col = next((i+1 for i, h in enumerate(header) if "project" in h and "name" in h), None)
    if proj_col is None:
        proj_col = next((i+1 for i, h in enumerate(header) if "project" in h), None)
    if proj_col is None:
        print(f"  Skipping {sheet_name}: no project column found")
        continue
    updated = 0
    for row in rows_list[1:]:
        cell = row[proj_col - 1]
        if not cell.value:
            continue
        # Strip carriage returns and both cases of _x000D_
        raw = str(cell.value).strip().replace("\r", "").replace("_x000D_", "").replace("_x000d_", "")
        key = raw.lower()
        if key in RENAMES:
            cell.value = RENAMES[key]
            updated += 1
        elif raw.lower().startswith("ad delhaize "):
            # "AD Delhaize X" → "AD X"
            new_name = "AD " + raw[len("AD Delhaize "):]
            cell.value = new_name
            updated += 1
        elif raw.lower().startswith("proxy delhaize "):
            # "Proxy Delhaize X" → "Proxy X"
            new_name = "Proxy " + raw[len("Proxy Delhaize "):]
            cell.value = new_name
            updated += 1
        elif raw.lower().startswith("edeka ") and " - " in raw:
            # "Edeka Vogel - Dueren" → "Edeka Vogel Dueren"
            new_name = raw.replace(" - ", " ")
            cell.value = new_name
            updated += 1
    print(f"{sheet_name}: {updated} cells renamed.")
    total += updated

wb.save(DATA_FILE)
print(f"\nTotal: {total} cells renamed. Saved.")
