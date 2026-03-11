"""Service for exporting dashboard summaries to Excel."""
import logging
import datetime
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import OUTPUT_DIR
from models.project import Project
from models.invoice import Invoice, DebtSummary

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill(start_color="2D6A9F", end_color="2D6A9F", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL = PatternFill(start_color="EBF3FA", end_color="EBF3FA", fill_type="solid")
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _write_header(ws, row: int, headers: list):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _BORDER


def _write_row(ws, row: int, values: list, alt: bool = False):
    fill = _ALT_FILL if alt else None
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        if fill:
            cell.fill = fill
        cell.border = _BORDER
        cell.alignment = Alignment(wrap_text=True)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value or ""))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def export_dashboard_excel(
    projects: List[Project],
    invoices: List[Invoice],
    debt_summaries: List[DebtSummary],
    yearly_summary: dict,
    output_dir: Path = OUTPUT_DIR,
) -> Path:
    """Export a dashboard summary workbook."""
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"CaddyCheck_Dashboard_{timestamp}.xlsx"

    wb = openpyxl.Workbook()

    # ── Sheet 1: Projects ──────────────────────────────────────────────────────
    ws_proj = wb.active
    ws_proj.title = "Projects"
    headers = ["Project Name", "Country", "# Cams", "Payment Month",
               "Installation Year", "Status", "Activation Date", "License EOP"]
    _write_header(ws_proj, 1, headers)
    for i, p in enumerate(projects, 2):
        _write_row(ws_proj, i, [
            p.project_name, p.country, p.num_cams, p.payment_month,
            p.installation_year, p.status,
            p.activation_date.strftime("%Y-%m-%d") if p.activation_date else "",
            p.license_eop.strftime("%Y-%m-%d") if p.license_eop else "",
        ], alt=(i % 2 == 0))
    _auto_width(ws_proj)

    # ── Sheet 2: Invoice Details ───────────────────────────────────────────────
    ws_inv = wb.create_sheet("Invoice Details")
    headers = ["Invoice #", "Project Name", "Maint. Year", "Amount",
               "Cameras", "Payment Date", "Paid", "Year"]
    _write_header(ws_inv, 1, headers)
    for i, inv in enumerate(invoices, 2):
        _write_row(ws_inv, i, [
            inv.invoice_number, inv.project_name, inv.maintenance_year,
            inv.payment_amount, inv.cameras_number,
            inv.payment_date.strftime("%Y-%m-%d") if inv.payment_date else "",
            inv.paid, inv.year,
        ], alt=(i % 2 == 0))
    _auto_width(ws_inv)

    # ── Sheet 3: Debt Summary ──────────────────────────────────────────────────
    ws_debt = wb.create_sheet("Debt Summary")
    headers = ["Project Name", "Country", "# Cams", "Payment Month",
               "Install Year", "Status", "Expected (€)", "Paid (€)",
               "Cancelled (€)", "Unpaid (€)"]
    _write_header(ws_debt, 1, headers)
    for i, ds in enumerate(debt_summaries, 2):
        _write_row(ws_debt, i, [
            ds.project_name, ds.country, ds.num_cams, ds.payment_month,
            ds.installation_year, ds.status,
            ds.total_expected, ds.total_paid, ds.total_cancelled, ds.total_unpaid,
        ], alt=(i % 2 == 0))
    _auto_width(ws_debt)

    # ── Sheet 4: Yearly Summary ────────────────────────────────────────────────
    ws_year = wb.create_sheet("Yearly Summary")
    _write_header(ws_year, 1, ["Year", "Total Paid (€)"])
    for i, (yr, total) in enumerate(sorted(yearly_summary.items()), 2):
        _write_row(ws_year, i, [yr, total], alt=(i % 2 == 0))
    _auto_width(ws_year)

    wb.save(output_path)
    logger.info("Dashboard exported to %s", output_path)
    return output_path
