"""Service for reading and parsing Excel data files."""
import logging
import datetime
from pathlib import Path
from typing import List, Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl.cell.cell import MergedCell

from config.settings import (
    PROJECTS_FILE,
    SHEET_PROJECTS_OVERVIEW,
    SHEET_INVOICE_DETAILS,
    SHEET_PROJECT_PAYMENT_SUMMARY,
    SHEET_YEARLY_PAYMENT_SUMMARY,
    normalize_month,
    get_project_overrides,
    get_data_paths,
)
from models.project import Project
from models.invoice import Invoice, DebtSummary

logger = logging.getLogger(__name__)


def _safe_int(val) -> Optional[int]:
    try:
        if pd.isna(val):
            return None
        return int(val)
    except Exception:
        return None


def _safe_float(val) -> Optional[float]:
    try:
        if pd.isna(val):
            return None
        return float(val)
    except Exception:
        return None


def _safe_str(val) -> str:
    try:
        if pd.isna(val):
            return ""
        return str(val).strip()
    except Exception:
        return ""


def _safe_datetime(val) -> Optional[datetime.datetime]:
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass
    if isinstance(val, datetime.datetime):
        return val
    if isinstance(val, datetime.date):
        return datetime.datetime.combine(val, datetime.time())
    try:
        return pd.to_datetime(val).to_pydatetime()
    except Exception:
        return None


def load_projects(filepath: Path = None) -> List[Project]:
    """Load and parse all projects from 'Projects overview' sheet."""
    if filepath is None:
        filepath = get_data_paths()["projects_file"]
    logger.info("Loading projects from %s", filepath)
    try:
        df = pd.read_excel(filepath, sheet_name=SHEET_PROJECTS_OVERVIEW, header=0)
    except Exception as e:
        logger.error("Failed to load projects: %s", e)
        return []

    projects = []
    for _, row in df.iterrows():
        name = _safe_str(row.get("Project Name"))
        if not name:
            continue

        # Parse M-1Y .. M-9Y columns (may contain invoice numbers)
        inv_numbers = {}
        for i in range(1, 10):
            col = f"M-{i}Y"
            v = _safe_float(row.get(col))
            if v is not None:
                inv_numbers[i] = v

        p = Project(
            project_name=name,
            country=_safe_str(row.get("Country")),
            num_cams=_safe_int(row.get("# Cams")) or 0,
            payment_month=normalize_month(_safe_str(row.get("payment month"))),
            installation_year=_safe_int(row.get("Installation Year")),
            project_approval=_safe_str(row.get("Project Approval")),
            activation_date=_safe_datetime(row.get("Activation date")),
            detection_type=_safe_str(row.get("Detection Type")),
            cart_type=_safe_str(row.get("Cart Type")),
            vim_version=_safe_str(row.get("VIM Version")),
            status=_safe_str(row.get("Status")),
            license_eop=_safe_datetime(row.get("Licsense EOP ")),
            caddy_back=_safe_str(row.get(" CaddyBack")),
            maintenance_invoice_numbers=inv_numbers,
        )
        projects.append(p)

    # Apply saved per-project rate overrides
    overrides = get_project_overrides()
    for proj in projects:
        key = proj.project_name.lower().strip()
        if key in overrides:
            proj.rate_y1_override = overrides[key].get("y1_rate")
            proj.rate_y2_override = overrides[key].get("y2_rate")

    logger.info("Loaded %d projects", len(projects))
    return projects


def load_invoices(filepath: Path = None) -> List[Invoice]:
    """Load and parse all invoices from 'Invoice details' sheet."""
    if filepath is None:
        filepath = get_data_paths()["projects_file"]
    logger.info("Loading invoices from %s", filepath)
    try:
        df = pd.read_excel(filepath, sheet_name=SHEET_INVOICE_DETAILS, header=0)
    except Exception as e:
        logger.error("Failed to load invoices: %s", e)
        return []

    invoices = []
    for _, row in df.iterrows():
        project_name = _safe_str(row.get("Project name"))
        if not project_name:
            continue

        year_val = row.get("Year")
        year_int: Optional[int] = None
        try:
            if not pd.isna(year_val):
                year_int = int(year_val)
        except Exception:
            pass

        inv = Invoice(
            invoice_number=_safe_float(row.get("Invoice Number")),
            project_name=project_name,
            maintenance_year=_safe_str(row.get("Maintenance Year")),
            payment_amount=_safe_float(row.get("Payment amount")) or 0.0,
            cameras_number=_safe_float(row.get("Cameras number")),
            payment_date=_safe_datetime(row.get("Payment Date")),
            paid=_safe_str(row.get("Paid")),
            year=year_int,
        )
        invoices.append(inv)

    logger.info("Loaded %d invoices", len(invoices))
    return invoices


def compute_debt_summaries(
    projects: List[Project],
    invoices: List[Invoice],
    current_year: Optional[int] = None,
) -> List[DebtSummary]:
    """
    For each project, compute expected vs. paid vs. unpaid.

    Expected = sum of get_expected_amount() for each year from installation_year to current_year.
    Paid = sum of paid invoice amounts.
    """
    if current_year is None:
        current_year = datetime.datetime.now().year

    # Index invoices by project name (case-insensitive)
    from collections import defaultdict
    invoice_map: dict = defaultdict(list)
    for inv in invoices:
        invoice_map[inv.project_name.lower().strip()].append(inv)

    summaries = []
    for proj in projects:
        key = proj.project_name.lower().strip()
        proj_invoices = invoice_map.get(key, [])

        # Expected: sum across all years from installation_year to current_year
        total_expected = 0.0
        if proj.installation_year and proj.installation_year <= current_year:
            for yr in range(proj.installation_year, current_year + 1):
                total_expected += proj.get_expected_amount(yr)

        total_paid = sum(inv.payment_amount for inv in proj_invoices if inv.is_paid())
        total_cancelled = sum(inv.payment_amount for inv in proj_invoices if inv.is_cancelled())

        ds = DebtSummary(
            project_name=proj.project_name,
            country=proj.country,
            num_cams=proj.num_cams,
            payment_month=proj.payment_month,
            installation_year=proj.installation_year,
            status=proj.status,
            total_expected=total_expected,
            total_paid=total_paid,
            total_cancelled=total_cancelled,
        )
        summaries.append(ds)

    return summaries


def get_yearly_summary(invoices: List[Invoice]) -> dict:
    """Return {year: total_paid_amount} from paid invoices."""
    result: dict = {}
    for inv in invoices:
        if inv.is_paid() and inv.year:
            result[inv.year] = result.get(inv.year, 0.0) + inv.payment_amount
    return result


def get_monthly_summary(invoices: List[Invoice], year: Optional[int] = None) -> dict:
    """Return {month_num: total_paid_amount} for a given year (or all years)."""
    result: dict = {}
    for inv in invoices:
        if not inv.is_paid():
            continue
        if year and inv.year != year:
            continue
        if inv.payment_date:
            m = inv.payment_date.month
            result[m] = result.get(m, 0.0) + inv.payment_amount
    return result


def get_projects_for_month(
    projects: List[Project],
    month_name: str,
) -> List[Project]:
    """Return active projects whose payment_month matches the given month name."""
    normalized = normalize_month(month_name)
    return [
        p for p in projects
        if p.payment_month == normalized and p.is_active()
    ]


# ── Column indices (1-based, openpyxl) for Projects overview ──────────────────
_PROJ_COL = {
    "Project Name":    1,
    "Country":         2,
    "# Cams":          3,
    "payment month":   4,
    "Installation Year": 5,
    "Project Approval": 6,
    "Activation date": 7,
    "Detection Type":  8,
    "Cart Type":       9,
    "VIM Version":     10,
    "Status":          20,
    "Licsense EOP ":   21,
    " CaddyBack":      22,
}

# ── Column indices (1-based, openpyxl) for Invoice details ────────────────────
_INV_COL = {
    "Invoice Number":  1,
    "Project name":    2,
    "Maintenance Year": 3,
    "Payment amount":  4,
    "Cameras number":  5,
    "Payment Date":    6,
    "Paid":            7,
    "Year":            8,
}


def _safe_write(ws, row: int, col: int, value):
    """Write to a cell, skipping read-only merged cells."""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def save_projects_to_excel(
    projects: List[Project],
    filepath: Path = None,
) -> None:
    """
    Write modified project data back to the 'Projects overview' sheet.
    Matches rows by Project Name (column A). Only updates editable fields;
    all other cells and formatting are preserved.
    """
    if filepath is None:
        filepath = get_data_paths()["projects_file"]
    logger.info("Saving %d projects to %s", len(projects), filepath)
    wb = openpyxl.load_workbook(filepath)
    ws = wb[SHEET_PROJECTS_OVERVIEW]

    # Build lookup: project_name -> Project
    proj_map = {p.project_name.strip(): p for p in projects}

    existing_names = set()
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        if isinstance(cell, MergedCell) or not cell.value:
            continue
        name = str(cell.value).strip()
        existing_names.add(name)
        proj = proj_map.get(name)
        if proj is None:
            continue

        _safe_write(ws, row_idx, _PROJ_COL["Country"],           proj.country or None)
        _safe_write(ws, row_idx, _PROJ_COL["# Cams"],            proj.num_cams or None)
        _safe_write(ws, row_idx, _PROJ_COL["payment month"],
                    proj.payment_month[:3] if proj.payment_month else None)
        _safe_write(ws, row_idx, _PROJ_COL["Installation Year"],  proj.installation_year)
        _safe_write(ws, row_idx, _PROJ_COL["Activation date"],    proj.activation_date)
        _safe_write(ws, row_idx, _PROJ_COL["Status"],             proj.status or None)
        _safe_write(ws, row_idx, _PROJ_COL["Licsense EOP "],      proj.license_eop)

    # Append rows for brand-new projects not yet in the sheet
    for proj in projects:
        if proj.project_name.strip() in existing_names:
            continue
        row_data = [None] * 22
        row_data[_PROJ_COL["Project Name"] - 1]       = proj.project_name
        row_data[_PROJ_COL["Country"] - 1]            = proj.country or None
        row_data[_PROJ_COL["# Cams"] - 1]             = proj.num_cams or None
        row_data[_PROJ_COL["payment month"] - 1]      = proj.payment_month[:3] if proj.payment_month else None
        row_data[_PROJ_COL["Installation Year"] - 1]  = proj.installation_year
        row_data[_PROJ_COL["Activation date"] - 1]    = proj.activation_date
        row_data[_PROJ_COL["Status"] - 1]             = proj.status or None
        row_data[_PROJ_COL["Licsense EOP "] - 1]      = proj.license_eop
        ws.append(row_data)
        logger.info("Appended new project row: %s", proj.project_name)

    wb.save(filepath)
    logger.info("Projects saved successfully.")


def save_invoices_to_excel(
    invoices: List[Invoice],
    filepath: Path = None,
) -> None:
    """
    Write modified invoice data back to the 'Invoice details' sheet.
    Matches rows by Invoice Number (column A) or row position.
    Only updates editable fields; all other cells and formatting are preserved.
    """
    if filepath is None:
        filepath = get_data_paths()["projects_file"]
    logger.info("Saving %d invoices to %s", len(invoices), filepath)
    wb = openpyxl.load_workbook(filepath)
    ws = wb[SHEET_INVOICE_DETAILS]

    def _invoice_identity(invoice_number, project_name, maintenance_year, year):
        try:
            inv_no = float(invoice_number) if invoice_number not in (None, "") else None
        except (ValueError, TypeError):
            inv_no = None
        return (
            inv_no,
            str(project_name or "").strip().lower(),
            str(maintenance_year or "").strip(),
            int(year) if year not in (None, "") else None,
        )

    # Build lookup: invoice_number -> Invoice and (project, maint year, year) -> Invoice
    inv_by_number: dict = {}
    inv_by_name_year: dict = {}
    for inv in invoices:
        if inv.invoice_number:
            inv_by_number[inv.invoice_number] = inv
        key = (inv.project_name.lower().strip(), inv.maintenance_year, inv.year)
        inv_by_name_year[key] = inv

    existing_keys = set()

    for row_idx in range(2, ws.max_row + 1):
        inv_no_cell = ws.cell(row=row_idx, column=_INV_COL["Invoice Number"])
        if isinstance(inv_no_cell, MergedCell):
            continue

        existing_keys.add(
            _invoice_identity(
                ws.cell(row=row_idx, column=_INV_COL["Invoice Number"]).value,
                ws.cell(row=row_idx, column=_INV_COL["Project name"]).value,
                ws.cell(row=row_idx, column=_INV_COL["Maintenance Year"]).value,
                ws.cell(row=row_idx, column=_INV_COL["Year"]).value,
            )
        )

        inv: Optional[Invoice] = None
        if inv_no_cell.value:
            try:
                inv = inv_by_number.get(float(inv_no_cell.value))
            except (ValueError, TypeError):
                pass

        if inv is None:
            # Fallback: match by name + maint year + year
            name_cell = ws.cell(row=row_idx, column=_INV_COL["Project name"])
            my_cell = ws.cell(row=row_idx, column=_INV_COL["Maintenance Year"])
            yr_cell = ws.cell(row=row_idx, column=_INV_COL["Year"])
            if name_cell.value:
                try:
                    yr_val = int(yr_cell.value) if yr_cell.value else None
                except (ValueError, TypeError):
                    yr_val = None
                key = (
                    str(name_cell.value).lower().strip(),
                    str(my_cell.value or "").strip(),
                    yr_val,
                )
                inv = inv_by_name_year.get(key)

        if inv is None:
            continue

        _safe_write(ws, row_idx, _INV_COL["Invoice Number"],   inv.invoice_number if inv.invoice_number else None)
        _safe_write(ws, row_idx, _INV_COL["Project name"],     inv.project_name or None)
        _safe_write(ws, row_idx, _INV_COL["Maintenance Year"], inv.maintenance_year or None)
        _safe_write(ws, row_idx, _INV_COL["Payment amount"],   inv.payment_amount if inv.payment_amount else None)
        _safe_write(ws, row_idx, _INV_COL["Cameras number"],   int(inv.cameras_number) if inv.cameras_number else None)
        _safe_write(ws, row_idx, _INV_COL["Payment Date"],     inv.payment_date)
        _safe_write(ws, row_idx, _INV_COL["Paid"],             inv.paid or None)
        _safe_write(ws, row_idx, _INV_COL["Year"],             inv.year)

    # Append brand-new invoices that do not already exist in the sheet.
    for inv in invoices:
        identity = _invoice_identity(inv.invoice_number, inv.project_name, inv.maintenance_year, inv.year)
        if identity in existing_keys:
            continue

        row_data = [None] * 8
        row_data[_INV_COL["Invoice Number"] - 1] = inv.invoice_number if inv.invoice_number else None
        row_data[_INV_COL["Project name"] - 1] = inv.project_name or None
        row_data[_INV_COL["Maintenance Year"] - 1] = inv.maintenance_year or None
        row_data[_INV_COL["Payment amount"] - 1] = inv.payment_amount if inv.payment_amount else None
        row_data[_INV_COL["Cameras number"] - 1] = int(inv.cameras_number) if inv.cameras_number else None
        row_data[_INV_COL["Payment Date"] - 1] = inv.payment_date
        row_data[_INV_COL["Paid"] - 1] = inv.paid or None
        row_data[_INV_COL["Year"] - 1] = inv.year
        ws.append(row_data)
        existing_keys.add(identity)
        logger.info("Appended new invoice row: %s / %s", inv.invoice_number, inv.project_name)

    wb.save(filepath)
    logger.info("Invoices saved successfully.")


def get_next_invoice_number(invoices: List[Invoice]) -> int:
    """Return the next free monthly invoice number (max existing + 1)."""
    max_no = 0
    for inv in invoices:
        if inv.invoice_number:
            try:
                n = int(inv.invoice_number)
                if n > max_no:
                    max_no = n
            except (ValueError, TypeError):
                pass
    return max_no + 1


def append_monthly_invoice_rows(
    invoice_number: int,
    projects: List[Project],
    year: int,
    filepath: Path = None,
) -> int:
    """
    Append new rows to 'Invoice details' for each project in a monthly invoice.
    Skips any project already present with this invoice_number (safe to re-run).
    Returns number of rows appended.
    """
    if filepath is None:
        filepath = get_data_paths()["projects_file"]
    wb = openpyxl.load_workbook(filepath)
    ws = wb[SHEET_INVOICE_DETAILS]

    # Collect existing (invoice_number, project_name_lower) to avoid duplicates
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        inv_no = row[_INV_COL["Invoice Number"] - 1]
        proj_name = row[_INV_COL["Project name"] - 1]
        if inv_no and proj_name:
            try:
                existing.add((float(inv_no), str(proj_name).strip().lower()))
            except (ValueError, TypeError):
                pass

    appended = 0
    for proj in sorted(projects, key=lambda p: p.project_name):
        if proj.num_cams <= 0:
            continue
        key = (float(invoice_number), proj.project_name.strip().lower())
        if key in existing:
            continue

        maint_label = proj.get_maintenance_year_label(year)
        amount = proj.get_expected_amount(year)

        row_data = [None] * 8
        row_data[_INV_COL["Invoice Number"] - 1]    = invoice_number
        row_data[_INV_COL["Project name"] - 1]      = proj.project_name
        row_data[_INV_COL["Maintenance Year"] - 1]  = maint_label
        row_data[_INV_COL["Payment amount"] - 1]    = amount
        row_data[_INV_COL["Cameras number"] - 1]    = proj.num_cams
        row_data[_INV_COL["Paid"] - 1]              = "No"
        row_data[_INV_COL["Year"] - 1]              = year

        ws.append(row_data)
        existing.add(key)
        appended += 1

    if appended > 0:
        wb.save(filepath)
        logger.info("Appended %d invoice rows for invoice #%d", appended, invoice_number)

    return appended
