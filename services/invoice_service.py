"""Service for generating monthly invoice Excel files from the template."""
import logging
import shutil
import datetime
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers

from config.settings import (
    INVOICE_TEMPLATE,
    OUTPUT_DIR,
    INVOICE_BILL_TO_NAME,
    INVOICE_BILL_TO_ADDRESS_1,
    INVOICE_BILL_TO_ADDRESS_2,
    INVOICE_BILL_TO_VAT,
    INVOICE_BANK_DETAILS,
    INVOICE_COMPANY_NAME,
    INVOICE_COMPANY_REG,
    INVOICE_PAYMENT_DAYS,
    RATE_Y1_PER_CAM,
    RATE_Y2_PLUS_PER_CAM,
    normalize_month,
)
from models.project import Project

logger = logging.getLogger(__name__)

# ── Column indices in the invoice template (1-based) ──────────────────────────
COL_SUPERMARKET = 1   # A  – project/store name
COL_UNITS = 7         # G  – number of cameras/units
COL_YEAR = 8          # H  – maintenance year label (Y1, Y2 …)
COL_PRICE = 9         # I  – price per unit
COL_LINE_TOTAL = 10   # J  – line total

# ── Fixed row numbers in the template header/footer ───────────────────────────
ROW_COMPANY_NAME = 3
ROW_DATE_LABEL = 4
ROW_INVOICE_NO = 5
ROW_PAYMENT_DUE = 8
ROW_COMPANY_REG = 9
ROW_BILL_TO_LABEL = 11
ROW_BILL_TO_NAME = 12
ROW_BILL_TO_ADDR1 = 13
ROW_BILL_TO_ADDR2 = 14
ROW_BILL_TO_VAT = 15
ROW_PROJECT_NAME_LABEL = 11   # column E
ROW_PROJECT_TITLE = 12        # column E – "Iretailcheck - Maintenance – Dec 2025"
ROW_HEADER = 19               # Supermarket | … | Units | Year | Price | Line Total
ROW_FIRST_DATA = 20           # First project data row in template

# Offset from last data row for footer sections
FOOTER_SUBTOTAL_OFFSET = 1
FOOTER_DISCOUNT_OFFSET = 2
FOOTER_LICENSE_OFFSET = 3
FOOTER_TAX_RATE_OFFSET = 3
FOOTER_TAX_OFFSET = 4
FOOTER_TOTAL_OFFSET = 5
FOOTER_BLANK_1 = 6
FOOTER_BLANK_2 = 7
FOOTER_BLANK_3 = 8
FOOTER_BLANK_4 = 9
FOOTER_BLANK_5 = 10
FOOTER_BANK_OFFSET = 11


def _determine_maintenance_year(
    project: Project,
    invoice_year: int,
) -> tuple:
    """
    Return (maintenance_year_int, rate, label) for a project in the given invoice year.

    Isolated helper – easy to adjust business rules here.
    """
    my = project.get_maintenance_year(invoice_year)
    label = f"Y{my}"
    rate = RATE_Y1_PER_CAM if my == 1 else RATE_Y2_PLUS_PER_CAM
    return my, rate, label


def generate_monthly_invoice(
    projects: List[Project],
    month_name: str,
    year: int,
    invoice_number: Optional[int] = None,
    output_dir: Path = None,
    template_path: Path = None,
) -> Path:
    """
    Generate a monthly invoice Excel file for the given month/year.

    Parameters
    ----------
    projects : list of Project
        Projects whose payment_month matches the selected month.
    month_name : str
        Full month name, e.g. 'December'.
    year : int
        Invoice year, e.g. 2025.
    invoice_number : int, optional
        Invoice number to use; auto-incremented from last template if None.
    output_dir : Path
        Directory to write the generated file.
    template_path : Path
        Path to the template .xlsx file.

    Returns
    -------
    Path
        Path to the generated Excel file.
    """
    from config.settings import get_data_paths
    paths = get_data_paths()
    if output_dir is None:
        output_dir = paths["output_dir"]
    if template_path is None:
        template_path = paths["invoice_template"]
    output_dir.mkdir(parents=True, exist_ok=True)
    month_abbr = month_name[:3]
    filename = f"CC_M_inv_{invoice_number or 'auto'}_{month_abbr}_{year}.xlsx"
    output_path = output_dir / filename

    # Copy template
    shutil.copy2(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    # Use first sheet (template has one sheet named "July")
    ws = wb.worksheets[0]

    # ── Invoice date and due date ──────────────────────────────────────────────
    invoice_date = datetime.datetime.now().date()
    due_date = invoice_date + datetime.timedelta(days=INVOICE_PAYMENT_DAYS)

    # ── Header fields ──────────────────────────────────────────────────────────
    _set_cell(ws, ROW_DATE_LABEL, COL_LINE_TOTAL, invoice_date)
    if invoice_number:
        _set_cell(ws, ROW_INVOICE_NO, COL_LINE_TOTAL, invoice_number)
    _set_cell(ws, ROW_PAYMENT_DUE, COL_LINE_TOTAL, due_date)

    # Bill-to section
    _set_cell(ws, ROW_BILL_TO_NAME, 1, INVOICE_BILL_TO_NAME)
    _set_cell(ws, ROW_BILL_TO_ADDR1, 1, INVOICE_BILL_TO_ADDRESS_1)
    _set_cell(ws, ROW_BILL_TO_ADDR2, 1, INVOICE_BILL_TO_ADDRESS_2)
    _set_cell(ws, ROW_BILL_TO_VAT, 1, INVOICE_BILL_TO_VAT)

    # Invoice title (E12)
    title = f"Iretailcheck - Maintenance - {month_name[:3]} {year}"
    _set_cell(ws, ROW_PROJECT_TITLE, 5, title)

    # ── Clear existing data rows (rows 20 to max) ─────────────────────────────
    # We'll rebuild all project rows from row 20 onwards
    from openpyxl.cell.cell import MergedCell
    max_template_row = ws.max_row
    for r in range(ROW_FIRST_DATA, max_template_row + 1):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    # ── Write project rows ────────────────────────────────────────────────────
    row = ROW_FIRST_DATA
    subtotal = 0.0

    # Sort projects by name
    sorted_projects = sorted(projects, key=lambda p: p.project_name)

    for proj in sorted_projects:
        if proj.num_cams <= 0:
            logger.warning("Skipping %s – 0 cameras", proj.project_name)
            continue

        _, rate, label = _determine_maintenance_year(proj, year)
        line_total = proj.num_cams * rate
        subtotal += line_total

        _set_cell(ws, row, COL_SUPERMARKET, proj.project_name)
        _set_cell(ws, row, COL_UNITS, proj.num_cams)
        _set_cell(ws, row, COL_YEAR, label)
        _set_cell(ws, row, COL_PRICE, rate)
        _set_cell(ws, row, COL_LINE_TOTAL, line_total)

        row += 1

    # ── Footer rows ───────────────────────────────────────────────────────────
    footer_start = row  # first row after last project

    # Subtotal
    _set_cell(ws, footer_start + 0, COL_PRICE, "Subtotal")
    _set_cell(ws, footer_start + 0, COL_LINE_TOTAL, subtotal)

    # Discount
    _set_cell(ws, footer_start + 1, COL_PRICE, "Discount")
    _set_cell(ws, footer_start + 1, COL_LINE_TOTAL, None)

    # License Period / Tax Rate
    license_date = datetime.datetime(year, 12, 1)
    _set_cell(ws, footer_start + 2, COL_SUPERMARKET, "License Period")
    _set_cell(ws, footer_start + 3, COL_SUPERMARKET, license_date)
    _set_cell(ws, footer_start + 2, COL_PRICE, "Tax/VAT Rate")
    _set_cell(ws, footer_start + 2, COL_LINE_TOTAL, 0)

    # Tax/VAT
    _set_cell(ws, footer_start + 3, COL_PRICE, "Tax/VAT ")
    _set_cell(ws, footer_start + 3, COL_LINE_TOTAL, 0)

    # Total
    _set_cell(ws, footer_start + 4, COL_PRICE, "Total")
    _set_cell(ws, footer_start + 4, COL_LINE_TOTAL, subtotal)

    # ── Bank details ──────────────────────────────────────────────────────────
    bank_row = footer_start + 11
    bd = INVOICE_BANK_DETAILS
    _set_cell(ws, bank_row, 1, "Bank details:")
    _set_cell(ws, bank_row, 2, f"Account name: {bd['account_name']}")
    _set_cell(ws, bank_row, COL_PRICE, "Signature")

    _set_cell(ws, bank_row + 1, 2, f"Account No:{bd['account_no']}")
    _set_cell(ws, bank_row + 1, COL_PRICE, bd["signature_name"])

    _set_cell(ws, bank_row + 2, 1, bd["website"])
    _set_cell(ws, bank_row + 2, 2, f"Bank name: {bd['bank_name']}")
    _set_cell(ws, bank_row + 2, COL_PRICE, invoice_date)

    _set_cell(ws, bank_row + 3, 2, f"SWIFT Code: {bd['swift']}")
    _set_cell(ws, bank_row + 4, 2, f"Branch: {bd['branch']}")
    _set_cell(ws, bank_row + 5, 2, f"IBAN Number: {bd['iban']}")

    _set_cell(ws, bank_row + 7, 1, "If you have any questions concerning this Invoice contact us")
    _set_cell(ws, bank_row + 9, 1, "Thank you for your business!")
    _set_cell(ws, bank_row + 10, 1, f",     {bd['website']}")
    _set_cell(ws, bank_row + 11, 1, bd["address"])
    _set_cell(ws, bank_row + 12, 1, bd["contact"])

    wb.save(output_path)
    logger.info("Invoice saved to %s", output_path)
    return output_path


def _set_cell(ws, row: int, col: int, value):
    """Set a cell value, skipping read-only merged cells."""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def generate_monthly_invoice_pdf(
    projects: List[Project],
    month_name: str,
    year: int,
    invoice_number: Optional[int] = None,
    output_dir: Path = None,
) -> Path:
    """
    Generate a monthly invoice as a PDF file using reportlab.

    Returns the path to the generated PDF.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_RIGHT
    import datetime as dt

    from config.settings import get_data_paths
    paths = get_data_paths()
    if output_dir is None:
        output_dir = paths["output_dir"]
    output_dir.mkdir(parents=True, exist_ok=True)

    month_abbr = month_name[:3]
    filename = f"CC_M-inv_{invoice_number or 'auto'}_{month_abbr}_{year}.pdf"
    output_path = output_dir / filename

    invoice_date = dt.datetime.now().date()
    due_date = invoice_date + dt.timedelta(days=INVOICE_PAYMENT_DAYS)
    bd = INVOICE_BANK_DETAILS

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    bold_style = ParagraphStyle("bold", parent=normal, fontName="Helvetica-Bold")
    small = ParagraphStyle("small", parent=normal, fontSize=8)
    right_style = ParagraphStyle("right", parent=normal, alignment=TA_RIGHT)

    story = []

    # Header table: company left, invoice details right
    header_data = [
        [
            Paragraph(f"<b>{INVOICE_COMPANY_NAME}</b>", bold_style),
            "",
            Paragraph(f"<b>Date:</b> {invoice_date.strftime('%d/%m/%Y')}", right_style),
        ],
        [
            Paragraph(INVOICE_COMPANY_REG, small),
            "",
            Paragraph(f"<b>Invoice No:</b> {invoice_number or ''}", right_style),
        ],
        [
            "",
            "",
            Paragraph(f"<b>Payment Due:</b> {due_date.strftime('%d/%m/%Y')}", right_style),
        ],
    ]
    header_table = Table(header_data, colWidths=[9 * cm, 3 * cm, 6 * cm])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.4 * cm))

    # Bill-to / Invoice title
    billed_data = [
        [
            Paragraph("<b>Bill To:</b>", bold_style),
            "",
            Paragraph(f"<b>Iretailcheck - Maintenance - {month_abbr} {year}</b>", bold_style),
        ],
        [Paragraph(INVOICE_BILL_TO_NAME, normal), "", ""],
        [Paragraph(INVOICE_BILL_TO_ADDRESS_1, normal), "", ""],
        [Paragraph(INVOICE_BILL_TO_ADDRESS_2, normal), "", ""],
        [Paragraph(INVOICE_BILL_TO_VAT, normal), "", ""],
    ]
    billed_table = Table(billed_data, colWidths=[7 * cm, 2 * cm, 9 * cm])
    billed_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(billed_table)
    story.append(Spacer(1, 0.5 * cm))

    # Project rows table
    col_headers = ["Supermarket / Store", "Units", "Year", "Price (€)", "Line Total (€)"]
    table_data = [col_headers]
    subtotal = 0.0

    for proj in sorted(projects, key=lambda p: p.project_name):
        if proj.num_cams <= 0:
            continue
        _, rate, label = _determine_maintenance_year(proj, year)
        line_total = proj.num_cams * rate
        subtotal += line_total
        table_data.append([
            proj.project_name,
            str(proj.num_cams),
            label,
            f"{rate:,.0f}",
            f"{line_total:,.0f}",
        ])

    # Footer rows
    table_data.append(["", "", "", "Subtotal", f"{subtotal:,.0f}"])
    table_data.append(["", "", "", "Discount", ""])
    table_data.append(["", "", "", "Tax/VAT Rate", "0%"])
    table_data.append(["", "", "", "Tax/VAT", "0"])
    table_data.append(["", "", "", Paragraph("<b>Total</b>", bold_style),
                        Paragraph(f"<b>{subtotal:,.0f}</b>", bold_style)])

    col_widths = [8.5 * cm, 1.8 * cm, 1.5 * cm, 3 * cm, 3.2 * cm]
    inv_table = Table(table_data, colWidths=col_widths, repeatRows=1)

    n_data = len(table_data)
    n_footer = 5
    last_data_row = n_data - n_footer - 1

    inv_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c5282")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (1, 0), (-1, 0), "RIGHT"),
        ("FONTSIZE", (0, 1), (-1, last_data_row), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, last_data_row), [colors.white, colors.HexColor("#f7fafc")]),
        ("ALIGN", (1, 1), (-1, last_data_row), "RIGHT"),
        ("ALIGN", (3, last_data_row + 1), (-1, -1), "RIGHT"),
        ("FONTSIZE", (0, last_data_row + 1), (-1, -1), 9),
        ("BACKGROUND", (3, -1), (-1, -1), colors.HexColor("#ebf4ff")),
        ("GRID", (0, 0), (-1, last_data_row), 0.3, colors.HexColor("#cbd5e0")),
        ("LINEABOVE", (3, last_data_row + 1), (-1, last_data_row + 1), 0.5, colors.HexColor("#2c5282")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(inv_table)
    story.append(Spacer(1, 0.8 * cm))

    # Bank details
    bank_data = [
        [Paragraph("<b>Bank Details:</b>", bold_style), Paragraph("<b>Signature:</b>", bold_style)],
        [f"Account name: {bd['account_name']}", bd["signature_name"]],
        [f"Account No: {bd['account_no']}", invoice_date.strftime('%d/%m/%Y')],
        [f"Bank name: {bd['bank_name']}", ""],
        [f"SWIFT Code: {bd['swift']}", ""],
        [f"Branch: {bd['branch']}", ""],
        [f"IBAN: {bd['iban']}", ""],
        [bd["website"], ""],
    ]
    bank_table = Table(bank_data, colWidths=[12 * cm, 6 * cm])
    bank_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("BOX", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e0")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f4f8")),
    ]))
    story.append(bank_table)
    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph("If you have any questions concerning this Invoice, contact us.", small))
    story.append(Paragraph("<b>Thank you for your business!</b>", bold_style))
    story.append(Paragraph(bd["address"], small))
    story.append(Paragraph(bd["contact"], small))

    doc.build(story)
    logger.info("PDF invoice saved to %s", output_path)
    return output_path


def archive_sent_invoice_pdf(
    pdf_path: Path,
    archive_name: Optional[str] = None,
    archive_dir: Path = None,
) -> Path:
    """Copy a sent PDF into a persistent archive folder and return the archived path."""
    from config.settings import get_data_paths

    paths = get_data_paths()
    if archive_dir is None:
        archive_dir = paths["output_dir"] / "sent_invoices"
    archive_dir.mkdir(parents=True, exist_ok=True)

    target_path = archive_dir / (archive_name or pdf_path.name)
    shutil.copy2(pdf_path, target_path)
    logger.info("Archived sent PDF to %s", target_path)
    return target_path


def get_invoice_preview_data(
    projects: List[Project],
    month_name: str,
    year: int,
) -> list:
    """
    Return a list of dicts representing the invoice line items for preview.

    Each dict has: project_name, num_cams, maintenance_year, rate, line_total.
    """
    rows = []
    subtotal = 0.0
    for proj in sorted(projects, key=lambda p: p.project_name):
        if proj.num_cams <= 0:
            continue
        _, rate, label = _determine_maintenance_year(proj, year)
        line_total = proj.num_cams * rate
        subtotal += line_total
        rows.append({
            "project_name": proj.project_name,
            "num_cams": proj.num_cams,
            "maintenance_year": label,
            "rate": rate,
            "line_total": line_total,
        })
    rows.append({
        "project_name": "",
        "num_cams": "",
        "maintenance_year": "",
        "rate": "TOTAL",
        "line_total": subtotal,
    })
    return rows
