"""CaddyCheck CRM — Streamlit web app (role-based access)."""
import datetime
import calendar
import io
import logging
import re
import sys
import zipfile
from pathlib import Path
from typing import Optional

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import openpyxl

# Make sure project root is on the path
sys.path.insert(0, str(Path(__file__).parent))

import math

logger = logging.getLogger(__name__)

def _safe_int(v, default=0):
    """Convert v to int safely, returning default for None/NaN/empty."""
    try:
        if v is None:
            return default
        if isinstance(v, float) and math.isnan(v):
            return default
        return int(v)
    except Exception:
        return default

def _safe_float(v, default=0.0):
    """Convert v to float safely, returning default for None/NaN/empty."""
    try:
        if v is None:
            return default
        f = float(v)
        return default if math.isnan(f) else f
    except Exception:
        return default

def _safe_str(v):
    """Convert v to str, returning '' for None/NaN."""
    if v is None:
        return ""
    try:
        if isinstance(v, float) and math.isnan(v):
            return ""
    except Exception:
        pass
    return str(v)


def _normalize_project_name_key(value: object) -> str:
    text = _safe_str(value).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def _project_name_matches(candidate: object, existing_keys: set[str]) -> bool:
    candidate_key = _normalize_project_name_key(candidate)
    if not candidate_key:
        return False
    if candidate_key in existing_keys:
        return True
    for existing_key in existing_keys:
        if candidate_key in existing_key or existing_key in candidate_key:
            return True

    from difflib import SequenceMatcher

    return any(SequenceMatcher(None, candidate_key, existing_key).ratio() >= 0.88 for existing_key in existing_keys)


def _suggest_best_project_match(candidate: object, project_names: list[str]) -> tuple[str, float]:
    candidate_key = _normalize_project_name_key(candidate)
    if not candidate_key or not project_names:
        return "", 0.0

    from difflib import SequenceMatcher

    best_name = ""
    best_score = 0.0
    for project_name in project_names:
        project_key = _normalize_project_name_key(project_name)
        if not project_key:
            continue
        if candidate_key == project_key:
            return project_name, 1.0
        if candidate_key in project_key or project_key in candidate_key:
            score = 0.95
        else:
            score = SequenceMatcher(None, candidate_key, project_key).ratio()
        if score > best_score:
            best_score = score
            best_name = project_name
    return best_name, best_score


def _suggest_project_matches(candidate: object, project_names: list[str], limit: int = 5) -> list[str]:
    candidate_key = _normalize_project_name_key(candidate)
    if not candidate_key or not project_names:
        return []

    from difflib import SequenceMatcher

    scored_matches: list[tuple[float, str]] = []
    for project_name in project_names:
        project_key = _normalize_project_name_key(project_name)
        if not project_key:
            continue
        if candidate_key == project_key:
            score = 1.0
        elif candidate_key in project_key or project_key in candidate_key:
            score = 0.95
        else:
            score = SequenceMatcher(None, candidate_key, project_key).ratio()
        scored_matches.append((score, project_name))

    scored_matches.sort(key=lambda item: (-item[0], item[1].lower()))
    return [name for score, name in scored_matches[:limit] if score >= 0.55]


def _get_exact_existing_project_match(candidate: object, project_names: list[str]) -> str:
    candidate_key = _normalize_project_name_key(candidate)
    if not candidate_key:
        return ""
    for project_name in project_names:
        if _normalize_project_name_key(project_name) == candidate_key:
            return project_name
    return ""


def _invoice_category_label(invoice) -> str:
    return str(getattr(invoice, "maintenance_year", "")).strip().lower()


def _is_paid_trial_category(invoice) -> bool:
    checker = getattr(invoice, "is_paid_trial_category", None)
    if callable(checker):
        try:
            return bool(checker())
        except Exception:
            pass
    return "paid trial" in _invoice_category_label(invoice)


def _is_new_installation_category(invoice) -> bool:
    checker = getattr(invoice, "is_new_installation_category", None)
    if callable(checker):
        try:
            return bool(checker())
        except Exception:
            pass
    return _invoice_category_label(invoice) == "y1"


def _is_maintenance_category(invoice) -> bool:
    checker = getattr(invoice, "is_maintenance_category", None)
    if callable(checker):
        try:
            return bool(checker())
        except Exception:
            pass
    return not _is_new_installation_category(invoice) and not _is_paid_trial_category(invoice)

from config.settings import (
    MONTH_ORDER,
    canonical_project_name,
    get_email_config,
    save_email_config,
    load_sent_invoices_log,
    append_sent_invoice_log,
    save_sent_invoices_log,
    load_license_change_log,
    append_license_change_log,
    load_orders_records,
    save_orders_records,
)

try:
    from config.settings import load_project_change_log, append_project_change_log
except Exception:
    # Backward-compatible fallback: keep app startup alive if Cloud deploy
    # temporarily runs mixed code versions during rollout.
    def load_project_change_log() -> list:
        return []

    def append_project_change_log(entry: dict) -> None:
        return None
from services.supabase_service import (
    load_projects,
    load_invoices,
    upsert_projects,
    upsert_invoices,
    delete_projects as delete_projects_supabase,
    rename_invoice_project_names as rename_invoice_project_names_supabase,
    replace_invoice_rows,
    get_next_invoice_number as _supa_next_inv_no,
    append_invoice_rows as _supa_append_invoice,
    load_orders as load_orders_supabase,
    create_orders as create_orders_supabase,
    update_order as update_order_supabase,
    delete_order as delete_order_supabase,
    download_sent_invoice_pdf as download_sent_invoice_pdf_supabase,
    upload_sent_invoice_pdf as upload_sent_invoice_pdf_supabase,
    upload_order_pdf as upload_order_pdf_supabase,
    download_order_pdf as download_order_pdf_supabase,
    create_order_pdf_signed_url as create_order_pdf_signed_url_supabase,
    create_sent_invoice_pdf_signed_url as create_sent_invoice_pdf_signed_url_supabase,
)

ORDER_STATUS_OPTIONS = [
    "New",
    "Ordered",
    "In Progress",
    "Installed",
    "Active",
    "Cancelled",
]

TICKET_SUBCATEGORY_OPTIONS = ["PushOut", "TopDown", "BackTray", "License"]
TICKET_TITLE_OPTIONS = [
    "Detection improvement",
    "License Problem",
    "Milestone connection",
    "DB corruption",
    "Camera mounting",
    "ROI of the detector not optimal",
]
TICKET_CAMERA_OPTIONS = [""] + [f"K{i}B" for i in range(1, 11)] + [f"K{i}TD" for i in range(1, 11)]

SUPPORTED_ORDER_FILE_SUFFIXES = {".pdf", ".xlsx", ".xlsm", ".xls", ".csv"}
from services.excel_service import (
    compute_debt_summaries,
    get_yearly_summary,
    get_monthly_invoice_projects,
    load_projects as load_projects_excel,
    load_invoices as load_invoices_excel,
    save_projects_to_excel,
    delete_projects_from_excel,
    rename_invoice_project_names_in_excel,
    save_invoices_to_excel,
    get_next_invoice_number as _excel_next_inv_no,
    append_monthly_invoice_rows as _excel_append_invoice,
)
from services.invoice_service import (
    archive_sent_invoice_pdf,
    generate_monthly_invoice,
    generate_monthly_invoice_pdf,
    get_invoice_preview_data,
)
from models.invoice import group_monthly_invoices

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CaddyCheck CRM",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Role-based login ──────────────────────────────────────────────────────────
# Passwords are stored in Streamlit secrets (secrets.toml or Streamlit Cloud secrets).
# Format in .streamlit/secrets.toml:
#   [passwords]
#   admin  = "your_admin_password"
#   viewer = "your_viewer_password"
#
# Roles: "admin" → can edit data   |   "viewer" → read-only

ROLES = {
    "admin":  {"label": "Admin",  "can_edit": True},
    "viewer": {"label": "Viewer", "can_edit": False},
}

APP_BUILD = "viewer-open-2026-04-23-1"

def _check_login(username: str, password: str):
    """Return role string if credentials match, else None."""
    normalized_username = str(username or "").strip().lower()
    normalized_password = str(password or "").strip()

    try:
        passwords = st.secrets.get("passwords", {})
    except Exception:
        passwords = {}

    accepted_passwords = {
        role: str(secret).strip()
        for role, secret in passwords.items()
        if str(secret).strip()
    }
    if normalized_username in accepted_passwords and normalized_password == accepted_passwords[normalized_username]:
        return normalized_username  # role == username key
    return None

def _login_form():
    st.markdown("## 🔐 CaddyCheck CRM Login")
    st.caption(f"Build: {APP_BUILD}")
    try:
        password_roles = [role for role, value in st.secrets.get("passwords", {}).items() if str(value).strip()]
    except Exception:
        password_roles = []
    if not password_roles:
        st.warning("Login is not configured. Add [passwords] secrets for admin and viewer in Streamlit Cloud.")
    with st.form("login_form"):
        username = st.selectbox("Role", list(ROLES.keys()), format_func=lambda k: ROLES[k]["label"])
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Login")
    if submitted:
        role = _check_login(username, password)
        if role:
            st.session_state["role"] = role
            st.rerun()
        else:
            st.error("Incorrect password. Try again.")


def _consume_flash_success(current_page: str) -> str:
    message = st.session_state.get("_flash_success", "")
    if not message:
        return ""
    flash_page = st.session_state.get("_flash_success_page")
    if flash_page not in (None, current_page):
        return ""
    st.session_state.pop("_flash_success", None)
    st.session_state.pop("_flash_success_page", None)
    return message

# ── Public renewal token handler (no login required) ──────────────────────────
_renew_token = st.query_params.get("token", "")
if _renew_token:
    from services.supabase_service import process_renewal_token as _process_token
    st.markdown("## 🔑 Subscription Renewal")
    with st.spinner("Validating renewal link…"):
        try:
            if not re.fullmatch(r"[A-Za-z0-9_-]{20,128}", _renew_token):
                _result = {"success": False, "message": "Invalid renewal link."}
            else:
                _result = _process_token(_renew_token)
        except Exception as _e:
            logger.exception("Failed to process renewal token")
            _result = {"success": False, "message": "Unable to validate this renewal link right now."}
    if _result["success"]:
        st.success(f"✅ {_result['message']}")
        st.markdown(
            f"**Project:** {_result['project_name']}  \n"
            f"**Valid until:** {_result['valid_until'].strftime('%B %d, %Y')}  \n"
            f"**Cameras licensed:** {_result['cameras_allowed']}"
        )
        st.info("Your subscription has been updated. You can close this page.")
    else:
        st.error(f"❌ {_result['message']}")
        st.caption("Contact your account manager if you believe this is an error.")
    st.stop()

# Gate access
if "role" not in st.session_state:
    _login_form()
    st.stop()

ROLE       = st.session_state["role"]
CAN_EDIT   = ROLES[ROLE]["can_edit"]

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .metric-card {
        background: white;
        border-radius: 10px;
        padding: 16px 20px;
        border-left: 5px solid #2D6A9F;
        box-shadow: 0 2px 6px rgba(0,0,0,0.08);
        margin-bottom: 4px;
    }
    .metric-title { font-size: 13px; font-weight: 600; color: #666; margin-bottom: 4px; }
    .metric-value { font-size: 26px; font-weight: 800; color: #2C3E50; }
    .card-income  { border-left-color: #27AE60; }
    .card-paid    { border-left-color: #2980B9; }
    .card-debt    { border-left-color: #E74C3C; }
    .card-monthly { border-left-color: #16A085; }
    .card-yearly  { border-left-color: #2C3E50; }
    .card-projects{ border-left-color: #8E44AD; }
    .card-cameras { border-left-color: #F39C12; }
    .stDataFrame thead tr th { background: #2D6A9F !important; color: white !important; }
</style>
""", unsafe_allow_html=True)


# ── Data loading (cached) ─────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def load_data():
    source_name = "Supabase"
    try:
        projects = load_projects()
        invoices = load_invoices()
    except RuntimeError as exc:
        if "Supabase credentials not configured" not in str(exc):
            raise
        projects = load_projects_excel()
        invoices = load_invoices_excel()
        source_name = "Excel (local fallback)"
    debt_summaries = compute_debt_summaries(projects, invoices)
    yearly_summary = get_yearly_summary(invoices)
    return projects, invoices, debt_summaries, yearly_summary, source_name


def _is_excel_source(source_name: str) -> bool:
    return source_name.startswith("Excel")


def _save_projects(projects, source_name: str) -> None:
    if _is_excel_source(source_name):
        save_projects_to_excel(projects)
        return
    upsert_projects(projects)


def _delete_projects(project_names, source_name: str) -> int:
    if _is_excel_source(source_name):
        return delete_projects_from_excel(project_names)
    return delete_projects_supabase(project_names)


def _rename_invoice_project_names(rename_map, source_name: str) -> int:
    if _is_excel_source(source_name):
        return rename_invoice_project_names_in_excel(rename_map)
    return rename_invoice_project_names_supabase(rename_map)


def _save_invoices(invoices, source_name: str) -> None:
    if _is_excel_source(source_name):
        save_invoices_to_excel(invoices)
        return
    upsert_invoices(invoices)


def _get_next_invoice_number(invoices, source_name: str) -> int:
    if _is_excel_source(source_name):
        return _excel_next_inv_no(invoices)
    return _supa_next_inv_no()


def _suggest_month_invoice_number(month_projects, invoices, year: int, source_name: str) -> int:
    """Prefer an existing invoice number for the selected month/year when possible."""
    month_project_keys = {
        _safe_str(project.project_name).strip().lower()
        for project in month_projects
        if _safe_str(project.project_name).strip()
    }
    candidates: dict[int, int] = {}
    for inv in invoices:
        if inv.year != year:
            continue
        project_key = _safe_str(inv.project_name).strip().lower()
        if project_key not in month_project_keys:
            continue
        invoice_number = _safe_int(inv.invoice_number, default=0)
        if invoice_number <= 0:
            continue
        candidates[invoice_number] = candidates.get(invoice_number, 0) + 1

    if candidates:
        return max(candidates.items(), key=lambda item: (item[1], item[0]))[0]
    return _get_next_invoice_number(invoices, source_name)


def _normalize_query_text(value: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9+ ]", " ", _safe_str(value).lower())).strip()


def _extract_question_year(question: str) -> Optional[int]:
    match = re.search(r"\b(20\d{2})\b", question)
    return int(match.group(1)) if match else None


def _extract_question_invoice_number(question: str) -> Optional[int]:
    match = re.search(r"(?:invoice\s*#?|inv\s*#?)\s*(\d{3,6})", question, re.IGNORECASE)
    if match:
        return int(match.group(1))
    standalone = re.search(r"\b(8\d{3,5})\b", question)
    return int(standalone.group(1)) if standalone else None


def _project_license_date(project) -> Optional[datetime.date]:
    if not project.license_eop:
        return None
    if isinstance(project.license_eop, datetime.datetime):
        return project.license_eop.date()
    if isinstance(project.license_eop, datetime.date):
        return project.license_eop
    return None


def _add_months(base_date: datetime.date, months: int) -> datetime.date:
    month_index = base_date.month - 1 + months
    year = base_date.year + month_index // 12
    month = month_index % 12 + 1
    day = min(base_date.day, calendar.monthrange(year, month)[1])
    return datetime.date(year, month, day)


def _license_status(project, today: Optional[datetime.date] = None) -> str:
    today = today or datetime.date.today()
    if str(getattr(project, "status", "")).strip().lower() == "offline":
        return "Offline"
    license_date = _project_license_date(project)
    if license_date is None:
        return "Missing"
    if license_date < today:
        return "Expired"
    next_month = today.month % 12 + 1
    next_month_year = today.year if today.month < 12 else today.year + 1
    if license_date.month == next_month and license_date.year == next_month_year:
        return "Update Next Month"
    return "Active"


def _extract_question_month(question: str) -> str:
    q = _normalize_query_text(question)
    for month in MONTH_ORDER:
        full = month.lower()
        short = month[:3].lower()
        if re.search(rf"\b{re.escape(full)}\b", q) or re.search(rf"\b{re.escape(short)}\b", q):
            return month
    return ""


def _match_project_from_question(question: str, projects, invoices) -> str:
    q = _normalize_query_text(question)
    candidates = sorted({p.project_name for p in projects} | {i.project_name for i in invoices}, key=len, reverse=True)
    for name in candidates:
        normalized = _normalize_query_text(name)
        if normalized and normalized in q:
            return name
    return ""


def _match_country_from_question(question: str, projects) -> str:
    q = _normalize_query_text(question)
    countries = sorted({p.country for p in projects if p.country}, key=len, reverse=True)
    for country in countries:
        normalized = _normalize_query_text(country)
        if normalized and normalized in q:
            return country
    return ""


def _build_invoice_answer_df(invoice_rows, projects) -> pd.DataFrame:
    country_map = {p.project_name.lower().strip(): p.country for p in projects}
    return pd.DataFrame([
        {
            "Invoice #": str(int(inv.invoice_number)) if inv.invoice_number else "—",
            "Project": inv.project_name,
            "Country": country_map.get(inv.project_name.lower().strip(), ""),
            "Maint. Year": inv.maintenance_year,
            "Amount (€)": float(inv.payment_amount),
            "Paid": inv.paid,
            "Year": inv.year or "",
        }
        for inv in invoice_rows
    ])


def _build_project_answer_df(project_rows) -> pd.DataFrame:
    return pd.DataFrame([
        {
            "Project": p.project_name,
            "Country": p.country,
            "# Cams": p.num_cams,
            "Payment Month": p.payment_month,
            "Status": p.status,
        }
        for p in project_rows
    ])


def _answer_data_question(question: str, projects, invoices, debt_summaries) -> tuple[str, Optional[pd.DataFrame]]:
    q = _normalize_query_text(question)
    if not q:
        return "Ask a question about projects, invoices, debt, or sent PDFs.", None

    project_name = _match_project_from_question(question, projects, invoices)
    country_name = _match_country_from_question(question, projects)
    year = _extract_question_year(question)
    invoice_number = _extract_question_invoice_number(question)
    month_name = _extract_question_month(question)

    if "next invoice" in q:
        next_inv = _get_next_invoice_number(invoices, _data_path)
        return f"The next invoice number is {next_inv}.", None

    if invoice_number is not None and ("invoice" in q or "inv" in q):
        invoice_rows = [inv for inv in invoices if _safe_int(inv.invoice_number, default=0) == invoice_number]
        if year is not None:
            invoice_rows = [inv for inv in invoice_rows if inv.year == year]

        if ("sent" in q or "emailed" in q or "email" in q):
            sent_rows = [row for row in load_sent_invoices_log() if _safe_int(row.get("invoice_number"), default=0) == invoice_number]
            if not sent_rows:
                return f"Invoice {invoice_number} has no sent PDF log entry.", None
            df = pd.DataFrame([
                {
                    "Sent At": _safe_str(row.get("sent_at", "")).replace("T", " ")[:19],
                    "Invoice #": _safe_int(row.get("invoice_number"), default=0),
                    "Month": _safe_str(row.get("month", "")),
                    "Year": _safe_int(row.get("year"), default=0),
                    "PDF": _safe_str(row.get("pdf_filename", "")),
                    "To": ", ".join(row.get("recipients", [])),
                    "Saved To Ledger": "Yes" if row.get("saved_to_ledger") else "No",
                }
                for row in reversed(sent_rows)
            ])
            return f"Found {len(df)} sent PDF log record(s) for invoice {invoice_number}.", df

        if not invoice_rows:
            return f"Invoice {invoice_number} is not in the invoice ledger.", None

        total_amount = sum(float(inv.payment_amount) for inv in invoice_rows)
        df = _build_invoice_answer_df(invoice_rows, projects)
        return f"Invoice {invoice_number} has {len(df)} row(s) totaling €{total_amount:,.0f}.", df

    if ("sent" in q or "emailed" in q or "email" in q) and ("pdf" in q or "invoice" in q):
        rows = load_sent_invoices_log()
        if invoice_number is not None:
            rows = [row for row in rows if _safe_int(row.get("invoice_number"), default=0) == invoice_number]
        if year is not None:
            rows = [row for row in rows if _safe_int(row.get("year"), default=0) == year]
        if project_name:
            rows = [row for row in rows if project_name.lower() in _safe_str(row.get("subject", "")).lower()]
        if not rows:
            return "No sent PDF invoices match that question.", None
        df = pd.DataFrame([
            {
                "Sent At": _safe_str(row.get("sent_at", "")).replace("T", " ")[:19],
                "Invoice #": _safe_int(row.get("invoice_number"), default=0),
                "Month": _safe_str(row.get("month", "")),
                "Year": _safe_int(row.get("year"), default=0),
                "PDF": _safe_str(row.get("pdf_filename", "")),
                "To": ", ".join(row.get("recipients", [])),
            }
            for row in reversed(rows)
        ])
        return f"Found {len(df)} sent PDF invoice record(s).", df

    if "active" in q and "project" in q and ("how many" in q or "count" in q or "number" in q):
        active_projects = [p for p in projects if p.is_active()]
        if country_name:
            active_projects = [p for p in active_projects if p.country == country_name]
        return f"There are {len(active_projects)} active project(s){' in ' + country_name if country_name else ''}.", None

    if ("project" in q or "bill" in q) and month_name:
        month_projects = get_projects_for_month(projects, month_name)
        if year is not None:
            filtered_projects = month_projects
        else:
            filtered_projects = month_projects
        if country_name:
            filtered_projects = [p for p in filtered_projects if p.country == country_name]
        if not filtered_projects:
            return f"No projects found for {month_name}{' ' + str(year) if year else ''}.", None
        df = _build_project_answer_df(filtered_projects)
        return f"Found {len(df)} project(s) billed in {month_name}{' ' + str(year) if year else ''}.", df

    if project_name and any(token in q for token in ["project", "status", "country", "payment month", "camera", "cams", "details"]):
        project = next((p for p in projects if p.project_name == project_name), None)
        if project is None:
            return f"I could not find project details for {project_name}.", None
        df = pd.DataFrame([{
            "Project": project.project_name,
            "Country": project.country,
            "# Cams": project.num_cams,
            "Payment Month": project.payment_month,
            "Install Year": project.installation_year or "",
            "Status": project.status,
        }])
        return f"Here are the project details for {project_name}.", df

    if "debt" in q or "unpaid" in q:
        debt_rows = [inv for inv in invoices if inv.is_unpaid()]
        if invoice_number is not None:
            debt_rows = [inv for inv in debt_rows if _safe_int(inv.invoice_number, default=0) == invoice_number]
        if year is not None:
            debt_rows = [inv for inv in debt_rows if inv.year == year]
        if country_name:
            project_names_in_country = {p.project_name for p in projects if p.country == country_name}
            debt_rows = [inv for inv in debt_rows if inv.project_name in project_names_in_country]
        if project_name:
            debt_rows = [inv for inv in debt_rows if inv.project_name == project_name]
        if "trial" in q:
            debt_rows = [inv for inv in debt_rows if _is_paid_trial_category(inv)]
        elif "y1" in q or "first year" in q or "new installation" in q:
            debt_rows = [inv for inv in debt_rows if _is_new_installation_category(inv)]
        elif "y2+" in q or "maintenance" in q:
            debt_rows = [inv for inv in debt_rows if _is_maintenance_category(inv)]
        if not debt_rows:
            return "No unpaid invoice rows match that question.", None
        total_amount = sum(float(inv.payment_amount) for inv in debt_rows)
        df = _build_invoice_answer_df(debt_rows, projects)
        return f"I found {len(df)} unpaid invoice row(s) totaling €{total_amount:,.0f}.", df

    if "top" in q and "debt" in q:
        debt_by_project = {}
        for inv in invoices:
            if not inv.is_unpaid():
                continue
            if year is not None and inv.year != year:
                continue
            debt_by_project.setdefault(inv.project_name, 0.0)
            debt_by_project[inv.project_name] += float(inv.payment_amount)
        if not debt_by_project:
            return "No unpaid debt rows match that question.", None
        df = pd.DataFrame([
            {"Project": name, "Debt (€)": amount}
            for name, amount in sorted(debt_by_project.items(), key=lambda item: item[1], reverse=True)[:10]
        ])
        return "Here are the top debt projects.", df

    if "invoice" in q and project_name:
        project_invoices = [inv for inv in invoices if inv.project_name == project_name]
        if year is not None:
            project_invoices = [inv for inv in project_invoices if inv.year == year]
        if not project_invoices:
            return f"No invoice rows found for {project_name}{' in ' + str(year) if year else ''}.", None
        df = _build_invoice_answer_df(project_invoices, projects)
        return f"I found {len(df)} invoice row(s) for {project_name}.", df

    if "country" in q and ("debt" in q or "unpaid" in q) and not country_name:
        return "I could not match the country name in that question.", None

    return (
        "I can answer questions like: 'What is the Y1 debt for 2026?', 'Show unpaid invoices for AD Denderleeuw', "
        "'How many active projects are in Belgium?', 'What invoices exist for Rewe Schorn - Bergheim?', "
        "'Show invoice 8676', 'Was invoice 8676 sent?', 'Which projects are billed in April?', or 'Show sent PDF invoices for 2026'.",
        None,
    )


def _find_invoice_header_row(worksheet) -> Optional[tuple[int, dict[str, int]]]:
    for row_idx in range(1, min(worksheet.max_row, 80) + 1):
        labels = {}
        for col_idx in range(1, min(worksheet.max_column, 20) + 1):
            cell_value = _safe_str(worksheet.cell(row=row_idx, column=col_idx).value).strip().lower()
            if not cell_value:
                continue
            normalized = re.sub(r"\s+", " ", cell_value)
            if "supermarket" in normalized or normalized == "project" or "project name" in normalized:
                labels["project_name"] = col_idx
            elif normalized in ("units", "cameras", "camera", "qty", "quantity"):
                labels["cameras_number"] = col_idx
            elif normalized in ("year", "maint. year", "maintenance year"):
                labels["maintenance_year"] = col_idx
            elif "line total" in normalized or normalized in ("amount", "amount (€)", "total"):
                labels["payment_amount"] = col_idx
        if {"project_name", "cameras_number", "payment_amount"}.issubset(labels):
            return row_idx, labels
    return None


def _infer_maintenance_year_label(title: str) -> str:
    normalized_title = _safe_str(title).strip().lower()
    if "trial" in normalized_title:
        return "Paid Trial-0.5Y"
    return ""


def _extract_invoice_number(worksheet) -> Optional[int]:
    preferred_cells = [(5, 10), (5, 9), (4, 10)]
    for row_idx, col_idx in preferred_cells:
        try:
            return int(worksheet.cell(row=row_idx, column=col_idx).value)
        except (TypeError, ValueError):
            pass

    for row_idx in range(1, min(worksheet.max_row, 20) + 1):
        for col_idx in range(1, min(worksheet.max_column, 12) + 1):
            label = _safe_str(worksheet.cell(row=row_idx, column=col_idx).value).strip().lower()
            if "invoice" in label and "#" in label or label == "invoice #":
                for neighbor_col in range(col_idx + 1, min(col_idx + 4, worksheet.max_column) + 1):
                    try:
                        return int(worksheet.cell(row=row_idx, column=neighbor_col).value)
                    except (TypeError, ValueError):
                        continue
    return None


def _extract_invoice_title_and_year(worksheet) -> tuple[str, Optional[int]]:
    candidate_text = []
    for row_idx in range(1, min(worksheet.max_row, 20) + 1):
        for col_idx in range(1, min(worksheet.max_column, 12) + 1):
            value = _safe_str(worksheet.cell(row=row_idx, column=col_idx).value).strip()
            if value:
                candidate_text.append(value)

    title = ""
    for value in candidate_text:
        if "maintenance" in value.lower() or "trial" in value.lower():
            title = value
            break
    if not title and candidate_text:
        title = candidate_text[0]

    year_match = re.search(r"(20\d{2})", title)
    if not year_match:
        for value in candidate_text:
            year_match = re.search(r"(20\d{2})", value)
            if year_match:
                break
    return title, int(year_match.group(1)) if year_match else None


_PROJECT_ORDER_HEADER_ALIASES = {
    "project_name": {
        "project", "project name", "store", "store name", "site", "site name", "customer", "supermarket",
    },
    "country": {"country", "market"},
    "num_cams": {"cams", "cameras", "camera", "units", "qty", "quantity", "number of cameras", "# cams", "# cameras"},
    "payment_amount": {"amount", "payment amount", "total amount", "order total", "total", "value"},
    "payment_month": {"payment month", "billing month", "invoice month", "charge month"},
    "installation_year": {"install year", "installation year", "installed year", "year installed"},
    "activation_date": {"activation date", "activation", "go live", "go live date", "start date"},
    "status": {"status", "project status"},
    "license_eop": {"license eop", "license expiry", "license end", "valid until", "valid to"},
}


def _normalize_upload_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _safe_str(value).strip().lower()).strip()


def _find_project_order_header_row(df: pd.DataFrame) -> Optional[tuple[int, dict[str, int]]]:
    max_rows = min(len(df), 40)
    max_cols = min(len(df.columns), 30)
    for row_idx in range(max_rows):
        labels: dict[str, int] = {}
        for col_idx in range(max_cols):
            normalized = _normalize_upload_header(df.iat[row_idx, col_idx])
            if not normalized:
                continue
            for field, aliases in _PROJECT_ORDER_HEADER_ALIASES.items():
                if normalized in aliases:
                    labels.setdefault(field, col_idx)
        if "project_name" in labels:
            return row_idx, labels
    return None


def _load_project_order_pdf_as_dataframe(file_bytes: bytes) -> pd.DataFrame:
    try:
        import pdfplumber
    except Exception as exc:
        raise RuntimeError(f"PDF import requires pdfplumber: {exc}")

    extracted_rows: list[list[str]] = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for table in tables:
                for row in table or []:
                    normalized_row = [_safe_str(cell) for cell in (row or [])]
                    if any(cell.strip() for cell in normalized_row):
                        extracted_rows.append(normalized_row)

    if not extracted_rows:
        raise ValueError("Could not find any table rows in the uploaded PDF order")

    max_cols = max(len(row) for row in extracted_rows)
    padded_rows = [row + [""] * (max_cols - len(row)) for row in extracted_rows]
    return pd.DataFrame(padded_rows)


def _extract_project_order_pdf_text(file_bytes: bytes) -> str:
    try:
        import pdfplumber
    except Exception as exc:
        raise RuntimeError(f"PDF import requires pdfplumber: {exc}")

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        pages = [page.extract_text() or "" for page in pdf.pages]
    return "\n".join(pages)


def _parse_order_amount_token(value: str) -> Optional[float]:
    token = _safe_str(value).replace("EUR", "").replace("€", "").strip()
    if not token:
        return None
    token = re.sub(r"\s+", "", token)
    if not re.search(r"\d", token):
        return None
    if "," in token and "." in token:
        if token.rfind(",") > token.rfind("."):
            token = token.replace(".", "").replace(",", ".")
        else:
            token = token.replace(",", "")
    elif token.count(",") == 1:
        fractional = token.split(",", 1)[1]
        if len(fractional) == 2:
            token = token.replace(",", ".")
        else:
            token = token.replace(",", "")
    elif token.count(".") > 1:
        token = token.replace(".", "")
    try:
        return float(token)
    except ValueError:
        return None


def _extract_order_camera_total_from_text(text: str) -> int:
    patterns = [
        r"\b(?:total|in total|overall|overall total|grand total)\s*(?:of\s*)?(\d{1,4})\s*(?:cameras?|cams?)\b",
        r"\b(\d{1,4})\s*(?:cameras?|cams?)\s*(?:in\s+total|total|overall)\b",
        r"\b(?:cameras?|cams?|quantity|qty)\s*[:\-]?\s*(\d{1,4})\b",
        r"\b(\d{1,4})\s*(?:checkouts?|cameras?|cams?)\b",
    ]
    matches: list[int] = []
    for pattern in patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            value = _safe_int(match.group(1), default=0)
            if value > 0:
                matches.append(value)
    return max(matches) if matches else 0


def _infer_order_camera_total_from_amount(payment_amount: float, rate_per_camera: float = 778.0) -> int:
    if payment_amount <= 0 or rate_per_camera <= 0:
        return 0
    inferred = payment_amount / rate_per_camera
    rounded = round(inferred)
    if rounded >= 1 and abs(inferred - rounded) <= 0.05:
        return int(rounded)
    return 0


def _extract_purchase_order_metrics(raw_df: pd.DataFrame, text: str) -> tuple[int, float]:
    ordered_cameras = 0
    line_totals: list[tuple[int, float]] = []  # (qty, amount) per detected line item
    seen_row_texts: set[str] = set()
    seen_amounts: set[float] = set()

    summary_label_re = re.compile(
        r"^(total|subtotal|sub[-\s]*total|grand\s*total|amount\s*due|balance\s*due|"
        r"vat|tax|btw|tva|sum|net\s*amount|gross|due|excl\.?\s*vat|incl\.?\s*vat|"
        r"shipping|delivery|discount|payable)\s*[:\-]?\s*$",
        re.IGNORECASE,
    )
    header_only_re = re.compile(r"^(article|description|qty|quantity|unit\s*price|item|price|amount|total)$", re.IGNORECASE)

    def _infer_qty_from_amounts(amounts: list[float]) -> int:
        for total_amount in sorted(amounts, reverse=True):
            for unit_amount in sorted(amounts):
                if unit_amount <= 0 or total_amount <= unit_amount:
                    continue
                ratio = total_amount / unit_amount
                rounded_ratio = round(ratio)
                if rounded_ratio >= 1 and abs(ratio - rounded_ratio) <= 0.05:
                    return int(rounded_ratio)
        return 0

    for row_idx in range(len(raw_df.index)):
        row_values = [_safe_str(raw_df.iat[row_idx, col_idx]).strip() for col_idx in range(len(raw_df.columns))]
        if not any(row_values):
            continue

        # Identify summary rows: rows where the only text content is a summary label
        # (e.g., "Total", "Subtotal", "VAT") plus one amount, and no qty / description.
        non_empty = [v for v in row_values if v]
        text_cells = [v for v in non_empty if _parse_order_amount_token(v) is None and not re.fullmatch(r"\d{1,4}", v)]
        is_header_row = bool(text_cells) and all(header_only_re.match(t) for t in text_cells)
        if is_header_row:
            continue
        is_summary_row = bool(text_cells) and all(summary_label_re.match(t) for t in text_cells)

        normalized_row_text = re.sub(r"\s+", " ", " ".join(row_values).strip()).lower()
        if normalized_row_text in seen_row_texts:
            continue
        seen_row_texts.add(normalized_row_text)

        qty_candidates = [
            _safe_int(cell, default=0)
            for cell in row_values
            if re.fullmatch(r"\d{1,4}", _safe_str(cell).strip())
        ]
        amount_candidates = [
            amount
            for amount in (_parse_order_amount_token(cell) for cell in row_values)
            if amount is not None and amount > 0
        ]
        if not amount_candidates:
            continue
        if is_summary_row:
            # Skip: this is a totals/VAT/subtotal row, not a line item
            continue
        inferred_qty = _infer_qty_from_amounts(amount_candidates) if len(amount_candidates) >= 2 else 0
        if not qty_candidates and not inferred_qty:
            continue

        line_qty = max(qty_candidates) if qty_candidates else inferred_qty
        line_amount = round(max(amount_candidates), 2)
        # Dedupe by amount alone — different line items rarely share the exact same total
        if line_amount in seen_amounts:
            continue
        seen_amounts.add(line_amount)
        line_totals.append((line_qty, line_amount))
        ordered_cameras += line_qty

    payment_amount = sum(amount for _, amount in line_totals)

    if ordered_cameras <= 0:
        patterns = [
            r"(?:installation type|qty|quantity|cameras?|cams?)\s*[:\-]?\s*(\d{1,4})\b",
            r"\b(\d{1,4})\s*(?:checkouts?|cameras?|cams?)\b",
            r"\b(\d{1,4})\s*checkouts?\s*[-+]",
        ]
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                ordered_cameras = max(ordered_cameras, _safe_int(match.group(1), default=0))

    if payment_amount <= 0:
        amount_tokens = [
            amount
            for amount in (_parse_order_amount_token(match.group(1)) for match in re.finditer(r"(?:€|EUR)\s*([\d.,]+)", text, re.IGNORECASE))
            if amount is not None and amount > 0
        ]
        unique_amount_tokens = sorted({round(amount, 2) for amount in amount_tokens}, reverse=True)
        if unique_amount_tokens:
            # Prefer the largest amount as the order total (typically the grand total)
            payment_amount = unique_amount_tokens[0]

    amount_inferred_cams = _infer_order_camera_total_from_amount(payment_amount)
    if amount_inferred_cams > ordered_cameras:
        ordered_cameras = amount_inferred_cams

    return ordered_cameras, round(payment_amount, 2)


def _extract_order_date_from_text(text: str) -> Optional[datetime.datetime]:
    patterns = [
        r"(?:delivery date|date)\s*[:\-]?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
        r"(?:delivery date|date)\s*[:\-]?\s*(\d{4}-\d{2}-\d{2})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            parsed = _parse_optional_datetime(match.group(1))
            if parsed is not None:
                return parsed
    return None


def _guess_order_country(project_name: str) -> str:
    normalized_name = _safe_str(project_name).lower()
    if "delhaize" in normalized_name:
        return "Belgium"
    return ""


def _guess_project_name_from_order_filename(filename: str) -> str:
    stem = _safe_str(Path(filename).stem).strip()
    stem = stem.replace("_", " ")
    stem = re.sub(r"^\d{4}-\d{2}-\d{2,3}\s+", "", stem)
    stem = re.sub(r"^\d{4}-\d{2}-\d+\s+", "", stem)
    stem = re.sub(r"\b(?:revised?|rev|extra(?:\s+pshout)?|pushout)\b.*$", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"\s+", " ", stem).strip(" -_")
    return stem


def _parse_single_project_order_pdf(file_bytes: bytes, filename: str, raw_df: Optional[pd.DataFrame] = None) -> list[dict]:
    text = _extract_project_order_pdf_text(file_bytes)
    project_name = _guess_project_name_from_order_filename(filename)
    if not project_name:
        raise ValueError("Could not infer a project name from this PDF filename")

    num_cams = 0
    payment_amount = 0.0
    if raw_df is not None:
        num_cams, payment_amount = _extract_purchase_order_metrics(raw_df, text)

    country = _guess_order_country(project_name)
    country_match = re.search(r"(?:country|market)\s*[:\-]?\s*([A-Za-z][A-Za-z\s-]{2,40})", text, re.IGNORECASE)
    if country_match:
        country = _safe_str(country_match.group(1)).strip()

    activation_date = _extract_order_date_from_text(text)

    return [{
        "project_name": project_name,
        "country": country,
        "num_cams": num_cams,
        "payment_amount": payment_amount,
        "payment_month": "",
        "installation_year": activation_date.year if activation_date else None,
        "activation_date": activation_date,
        "status": "Active",
        "license_eop": None,
    }]


def _parse_optional_datetime(value) -> Optional[datetime.datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime.combine(value, datetime.time.min)
    try:
        return pd.to_datetime(value).to_pydatetime()
    except Exception:
        return None


def _parse_uploaded_project_order(file_bytes: bytes, filename: str) -> tuple[dict, list[dict]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        raw_df = pd.read_csv(io.BytesIO(file_bytes), header=None)
    elif suffix == ".pdf":
        raw_df = _load_project_order_pdf_as_dataframe(file_bytes)
    elif suffix in {".xlsx", ".xlsm", ".xls"}:
        raw_df = pd.read_excel(io.BytesIO(file_bytes), header=None)
    else:
        raise ValueError("Only PDF, XLSX, XLSM, XLS, and CSV order files are supported")

    header_info = _find_project_order_header_row(raw_df)
    if header_info is None:
        if suffix == ".pdf":
            rows = _parse_single_project_order_pdf(file_bytes, filename, raw_df=raw_df)
            metadata = {
                "filename": filename,
                "row_count": len(rows),
                "columns_found": ["project_name", "num_cams", "payment_amount", "installation_year"],
            }
            return metadata, rows
        raise ValueError("Could not find a project-name column in the uploaded order file")

    header_row_idx, columns = header_info
    rows: list[dict] = []
    explicit_camera_total = _extract_order_camera_total_from_text(_extract_project_order_pdf_text(file_bytes)) if suffix == ".pdf" else 0
    for row_idx in range(header_row_idx + 1, len(raw_df.index)):
        project_name = _safe_str(raw_df.iat[row_idx, columns["project_name"]]).strip()
        if not project_name:
            continue

        activation_date = _parse_optional_datetime(raw_df.iat[row_idx, columns["activation_date"]]) if "activation_date" in columns else None
        install_year = _safe_int(raw_df.iat[row_idx, columns["installation_year"]], default=0) or None
        if install_year is None and activation_date is not None:
            install_year = activation_date.year

        payment_month = normalize_month(_safe_str(raw_df.iat[row_idx, columns["payment_month"]]).strip()) if "payment_month" in columns else ""
        status = _safe_str(raw_df.iat[row_idx, columns["status"]]).strip() if "status" in columns else "Active"

        rows.append({
            "project_name": project_name,
            "country": _safe_str(raw_df.iat[row_idx, columns["country"]]).strip() if "country" in columns else "",
            "num_cams": max(
                _safe_int(raw_df.iat[row_idx, columns["num_cams"]], default=0),
                explicit_camera_total,
                _infer_order_camera_total_from_amount(
                    _safe_float(raw_df.iat[row_idx, columns["payment_amount"]], default=0.0) if "payment_amount" in columns else 0.0
                ),
            ),
            "payment_amount": _safe_float(raw_df.iat[row_idx, columns["payment_amount"]], default=0.0) if "payment_amount" in columns else 0.0,
            "payment_month": payment_month,
            "installation_year": install_year,
            "activation_date": activation_date,
            "status": status or "Active",
            "license_eop": _parse_optional_datetime(raw_df.iat[row_idx, columns["license_eop"]]) if "license_eop" in columns else None,
        })

    metadata = {
        "filename": filename,
        "row_count": len(rows),
        "columns_found": sorted(columns.keys()),
    }
    return metadata, rows


def _default_order_reference(filename: str) -> str:
    stem = _safe_str(Path(filename).stem).strip()
    if not stem:
        return ""
    match = re.search(r"(?:order|po|oc)[-_ ]*([a-z0-9]+)", stem, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return stem


def _archive_uploaded_order_file(file_bytes: bytes, filename: str) -> Path:
    from config.settings import get_data_paths

    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(filename).stem).strip("_") or "order"
    safe_suffix = Path(filename).suffix or ".bin"
    stamp = datetime.datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    archive_dir = get_data_paths()["output_dir"] / "orders"
    archive_dir.mkdir(parents=True, exist_ok=True)
    archive_path = archive_dir / f"{stamp}-{safe_stem}{safe_suffix}"
    archive_path.write_bytes(file_bytes)
    return archive_path


def _resolve_order_archive_path(order: dict) -> Path | None:
    archive_path_text = _safe_str(order.get("source_archive_path")).strip()
    if archive_path_text:
        archive_path = Path(archive_path_text)
        if archive_path.exists():
            return archive_path

    from config.settings import get_data_paths

    source_name = _safe_str(order.get("source_filename")).strip()
    if not source_name:
        return None

    archive_dir = get_data_paths()["output_dir"] / "orders"
    if not archive_dir.exists():
        return None

    source_path = Path(source_name)
    candidate_stems = {
        source_path.stem,
        re.sub(r"[\\/]+", "_", source_name),
        re.sub(r"[^A-Za-z0-9._-]+", "_", source_path.stem).strip("_"),
    }
    candidate_stems = {stem for stem in candidate_stems if stem}

    for stem in candidate_stems:
        exact_matches = sorted(archive_dir.glob(f"*-{stem}{source_path.suffix or '.*'}"))
        if exact_matches:
            return exact_matches[0]

        fallback_matches = sorted(archive_dir.glob(f"*-{stem}*"))
        if fallback_matches:
            return fallback_matches[0]

    return None


def _get_order_pdf_bytes(order: dict) -> tuple[bytes, str] | None:
    """Return (file_bytes, filename) for an order's source PDF, or None."""
    storage_bucket = _safe_str(order.get("pdf_storage_bucket")).strip()
    storage_path = _safe_str(order.get("pdf_storage_path")).strip()
    if storage_bucket and storage_path:
        try:
            data = download_order_pdf_supabase(storage_bucket, storage_path)
            filename = Path(storage_path).name or _safe_str(order.get("source_filename")).strip() or "order.pdf"
            return data, filename
        except Exception as exc:
            logger.warning("Could not download order PDF from Supabase Storage: %s", exc)

    archive_path = _resolve_order_archive_path(order)
    if archive_path is not None:
        try:
            return archive_path.read_bytes(), archive_path.name
        except Exception as exc:
            logger.warning("Could not read archived order PDF: %s", exc)

    return None


def _expand_uploaded_order_sources(uploaded_order_files) -> list[dict]:
    expanded_sources: list[dict] = []
    for uploaded_file in uploaded_order_files:
        source_name = _safe_str(getattr(uploaded_file, "name", "")).strip()
        if not source_name:
            continue
        source_bytes = uploaded_file.getvalue()
        source_suffix = Path(source_name).suffix.lower()

        if source_suffix == ".zip":
            try:
                with zipfile.ZipFile(io.BytesIO(source_bytes)) as archive:
                    for member in archive.infolist():
                        if member.is_dir():
                            continue
                        member_name = Path(member.filename).name
                        member_suffix = Path(member_name).suffix.lower()
                        if member_suffix not in SUPPORTED_ORDER_FILE_SUFFIXES:
                            continue
                        expanded_sources.append({
                            "source_name": member_name,
                            "source_path": member.filename,
                            "file_bytes": archive.read(member),
                            "container_name": source_name,
                        })
            except Exception as exc:
                raise ValueError(f"Could not read ZIP file {source_name}: {exc}")
            continue

        if source_suffix not in SUPPORTED_ORDER_FILE_SUFFIXES:
            raise ValueError(f"Unsupported order file type: {source_name}")

        expanded_sources.append({
            "source_name": source_name,
            "source_path": source_name,
            "file_bytes": source_bytes,
            "container_name": source_name,
        })

    return expanded_sources


def _parse_uploaded_invoice_xlsx(file_bytes: bytes) -> tuple[dict, list[dict]]:
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    worksheet = workbook.worksheets[0]

    invoice_number = _extract_invoice_number(worksheet)
    title, invoice_year = _extract_invoice_title_and_year(worksheet)
    header_info = _find_invoice_header_row(worksheet)
    if header_info is None:
        raise ValueError("Could not find invoice table headers in the uploaded XLSX")

    header_row_idx, columns = header_info

    rows: list[dict] = []
    for row_idx in range(header_row_idx + 1, worksheet.max_row + 1):
        project_name = _safe_str(worksheet.cell(row=row_idx, column=columns["project_name"]).value).strip()
        if not project_name:
            continue
        if project_name.lower() == "license period":
            break

        cameras_number = _safe_int(worksheet.cell(row=row_idx, column=columns["cameras_number"]).value, default=0)
        payment_amount = _safe_float(worksheet.cell(row=row_idx, column=columns["payment_amount"]).value, default=0.0)
        maint_year_col = columns.get("maintenance_year")
        maint_year = ""
        if maint_year_col is not None:
            maint_year = _safe_str(worksheet.cell(row=row_idx, column=maint_year_col).value).strip()
        if not maint_year:
            maint_year = _infer_maintenance_year_label(title)
        if not maint_year and payment_amount <= 0:
            break

        if cameras_number <= 0 and payment_amount <= 0:
            continue

        rows.append({
            "invoice_number": str(invoice_number) if invoice_number is not None else None,
            "project_name": project_name,
            "maintenance_year": maint_year,
            "payment_amount": payment_amount,
            "cameras_number": cameras_number or None,
            "payment_date": None,
            "paid": "No",
            "year": invoice_year,
        })

    metadata = {
        "invoice_number": invoice_number,
        "title": title,
        "year": invoice_year,
        "row_count": len(rows),
        "total_amount": sum(float(row["payment_amount"] or 0) for row in rows),
    }
    return metadata, rows


def _is_missing_supabase_table_error(exc: Exception, table_name: str) -> bool:
    message = str(exc).lower()
    return (
        "could not find the table" in message
        and table_name.lower() in message
    )


def _normalize_order_status(value: str) -> str:
    cleaned = _safe_str(value).strip()
    if not cleaned:
        return "New"
    for option in ORDER_STATUS_OPTIONS:
        if option.lower() == cleaned.lower():
            return option
    return cleaned


def _normalize_country(value: str) -> str:
    """Normalize country codes and names to full country names."""
    cleaned = _safe_str(value).strip()
    if not cleaned:
        return ""
    country_mapping = {
        "bel": "Belgium",
        "ger": "Germany",
        "hol": "Netherlands",
        "lux": "Luxembourg",
        "fra": "France",
        "ita": "Italy",
        "esp": "Spain",
        "gbr": "United Kingdom",
        "deu": "Germany",
        "nld": "Netherlands",
    }
    normalized_key = cleaned.lower()
    return country_mapping.get(normalized_key, cleaned)


def _parse_order_date(value) -> Optional[datetime.date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    try:
        return pd.to_datetime(value).date()
    except Exception:
        return None


def _serialize_order_value(value):
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    return value


def _project_status_from_order_status(order_status: str) -> str:
    normalized = _normalize_order_status(order_status)
    if normalized in {"Active", "Installed"}:
        return "Active"
    return "Offline"


@st.cache_data(ttl=300)
def load_orders_data(source_name: str) -> tuple[list[dict], str]:
    if _is_excel_source(source_name):
        return load_orders_records(), "Local JSON"
    try:
        return load_orders_supabase(), "Supabase"
    except RuntimeError as exc:
        if "Supabase credentials not configured" not in str(exc):
            raise
        return load_orders_records(), "Local JSON"
    except Exception as exc:
        if _is_missing_supabase_table_error(exc, "orders"):
            return load_orders_records(), "Local JSON fallback (orders table missing)"
        raise


def _is_local_orders_source(orders_source_name: str) -> bool:
    return not orders_source_name.startswith("Supabase")


def _next_local_order_id(order_rows: list[dict]) -> int:
    return max((_safe_int(row.get("id"), default=0) for row in order_rows), default=0) + 1


def _create_orders(rows: list[dict], orders_source_name: str) -> int:
    cleaned_rows = [row for row in rows if _safe_str(row.get("project_name")).strip()]
    if not cleaned_rows:
        return 0
    if not _is_local_orders_source(orders_source_name):
        return create_orders_supabase(cleaned_rows)

    entries = load_orders_records()
    next_id = _next_local_order_id(entries)
    timestamp = datetime.datetime.utcnow().isoformat()
    for row in cleaned_rows:
        local_row = {key: _serialize_order_value(value) for key, value in row.items()}
        local_row["id"] = next_id
        local_row["status"] = _normalize_order_status(local_row.get("status", "New"))
        local_row["created_at"] = timestamp
        local_row["updated_at"] = timestamp
        entries.append(local_row)
        next_id += 1
    save_orders_records(entries)
    return len(cleaned_rows)


def _update_order(order_id: int, orders_source_name: str, **fields) -> None:
    if not _is_local_orders_source(orders_source_name):
        update_order_supabase(order_id, **fields)
        return

    entries = load_orders_records()
    updated = False
    for row in entries:
        if _safe_int(row.get("id"), default=0) != int(order_id):
            continue
        for key, value in fields.items():
            row[key] = _serialize_order_value(value)
        row["status"] = _normalize_order_status(row.get("status", "New"))
        row["updated_at"] = datetime.datetime.utcnow().isoformat()
        updated = True
        break
    if not updated:
        raise ValueError(f"Order id={order_id} was not found.")
    save_orders_records(entries)


def _delete_order(order_id: int, orders_source_name: str) -> None:
    if not _is_local_orders_source(orders_source_name):
        delete_order_supabase(order_id)
        return

    entries = load_orders_records()
    remaining = [row for row in entries if _safe_int(row.get("id"), default=0) != int(order_id)]
    save_orders_records(remaining)


def _append_invoice_rows(invoice_number: int, projects, year: int, source_name: str, description: str = None) -> int:
    if _is_excel_source(source_name):
        return _excel_append_invoice(invoice_number=invoice_number, projects=projects, year=year, description=description)
    return _supa_append_invoice(invoice_number=invoice_number, projects=projects, year=year, description=description)


def _push_excel_to_github() -> tuple[bool, str]:
    """Commit the updated Excel file back to GitHub so changes survive app restarts."""
    try:
        gh_cfg = st.secrets.get("github", {})
        token    = gh_cfg.get("token", "")
        repo_name = gh_cfg.get("repo", "zvshtorch-rgb/caddycheck-crm")
        if not token:
            return False, "GitHub token not set in secrets — changes saved locally only."
        try:
            from github import Github
        except ImportError:
            return False, "PyGithub is not installed — changes saved locally only."
        from config.settings import get_data_paths
        data_file = get_data_paths()["projects_file"]
        with open(data_file, "rb") as f:
            content = f.read()
        g = Github(token)
        repo = g.get_repo(repo_name)
        gh_file = repo.get_contents("data/CaddyCheckProjectsInfo.xlsx")
        repo.update_file(
            "data/CaddyCheckProjectsInfo.xlsx",
            "CRM: update data via web app",
            content,
            gh_file.sha,
        )
        return True, "Saved and committed to GitHub."
    except Exception as e:
        return False, f"Local save OK but GitHub commit failed: {e}"


def card(title: str, value: str, css_class: str):
    st.markdown(
        f'<div class="metric-card {css_class}">'
        f'<div class="metric-title">{title}</div>'
        f'<div class="metric-value">{value}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )


# ── Sidebar navigation ────────────────────────────────────────────────────────
st.sidebar.title("📊 CaddyCheck CRM")
st.sidebar.markdown("---")
page = st.sidebar.radio(
    "Navigation",
    ["📊 Dashboard", "❓ Ask Data", "🏗️ Projects", "📦 Orders", "🔐 Licenses", "🧾 Invoice Details", "💸 Debt Report", "📅 Monthly Invoice", "🎫 Tickets", "🏦 Bank Payment", "⚙️ Settings"],
    label_visibility="collapsed",
)
st.sidebar.markdown("---")
role_icon = "✏️" if CAN_EDIT else "👁️"
st.sidebar.caption(f"{role_icon} Logged in as **{ROLES[ROLE]['label']}**")
if st.sidebar.button("Refresh Data"):
    load_data.clear()
    st.cache_data.clear()
    st.rerun()
if st.sidebar.button("Logout"):
    del st.session_state["role"]
    st.rerun()

# Load data
try:
    projects, invoices, debt_summaries, yearly_summary, _data_path = load_data()
    st.sidebar.caption(f"Data: `{_data_path}`")
except Exception as e:
    st.error(f"Failed to load data: {e}")
    st.stop()

page_flash_success = _consume_flash_success(page)
if page_flash_success:
    st.toast(page_flash_success, icon="✅")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
if page == "📊 Dashboard":
    st.title("📊 Dashboard")
    if page_flash_success:
        st.success(page_flash_success)

    # ── Filters ───────────────────────────────────────────────────────────────
    with st.expander("Filters", expanded=True):
        col1, col2, col3, col4 = st.columns(4)
        years = sorted({inv.year for inv in invoices if inv.year}, reverse=True)
        year_opts = ["All"] + [str(y) for y in years]
        sel_year   = col1.selectbox("Year",    year_opts)
        sel_month  = col2.selectbox("Month",   ["All"] + MONTH_ORDER)
        countries  = sorted({p.country for p in projects if p.country})
        sel_country= col3.selectbox("Country", ["All"] + countries)
        sel_status = col4.selectbox("Status",  ["All", "Paid", "Unpaid", "Cancelled"])

    # ── Filter invoices ───────────────────────────────────────────────────────
    proj_country = {p.project_name.lower(): p.country for p in projects}
    month_num = MONTH_ORDER.index(sel_month) + 1 if sel_month != "All" else None

    def filter_invoices(invs):
        result = []
        for inv in invs:
            if sel_year != "All" and inv.year != int(sel_year):
                continue
            if month_num and inv.payment_date and inv.payment_date.month != month_num:
                continue
            if sel_country != "All":
                pc = proj_country.get(inv.project_name.lower().strip(), "")
                if pc != sel_country:
                    continue
            if sel_status == "Paid" and not inv.is_paid():
                continue
            if sel_status == "Unpaid" and not inv.is_unpaid():
                continue
            if sel_status == "Cancelled" and not inv.is_cancelled():
                continue
            result.append(inv)
        return result

    def filter_projects(projs):
        if sel_country != "All":
            return [p for p in projs if p.country == sel_country]
        return projs

    f_inv  = filter_invoices(invoices)
    f_proj = filter_projects(projects)

    # ── Summary cards ─────────────────────────────────────────────────────────
    total_paid   = sum(i.payment_amount for i in f_inv if i.is_paid())
    total_unpaid = sum(i.payment_amount for i in f_inv if i.is_unpaid())
    total_income = total_paid + total_unpaid
    active_count = sum(1 for p in f_proj if p.is_active())

    # Monthly/yearly ref year
    if sel_year != "All":
        ref_year = int(sel_year)
    else:
        paid_years = [i.year for i in f_inv if i.is_paid() and i.year]
        ref_year = max(paid_years) if paid_years else datetime.datetime.now().year

    yearly_val = sum(
        i.payment_amount for i in f_inv
        if i.is_paid() and i.year == ref_year
    )

    def _project_start_year(project) -> Optional[int]:
        if project.activation_date:
            if isinstance(project.activation_date, datetime.datetime):
                return project.activation_date.year
            return int(project.activation_date.year)
        if project.installation_year:
            return int(project.installation_year)
        return None

    # Keep card semantics in sync with "Total Cameras" trend:
    # active cameras accumulated up to the reference year.
    total_cams = 0
    for p in f_proj:
        start_year = _project_start_year(p)
        if p.is_active() and start_year is not None and start_year <= ref_year:
            total_cams += _safe_int(p.num_cams)

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    with c1: card("Total Income",            f"€{total_income:,.0f}", "card-income")
    with c2: card("Total Paid",              f"€{total_paid:,.0f}",   "card-paid")
    with c3: card("Total Debt",              f"€{total_unpaid:,.0f}", "card-debt")
    with c4: card(f"Yearly Income ({ref_year})",  f"€{yearly_val:,.0f}",  "card-yearly")
    with c5: card("Active Projects",         str(active_count),       "card-projects")
    with c6: card("Total Cameras",           str(total_cams),         "card-cameras")

    st.markdown("---")

    # ── Cameras by Project table ───────────────────────────────────────────────
    st.subheader("Cameras by Project")
    sorted_proj = sorted(f_proj, key=lambda p: (0 if p.is_active() else 1, p.project_name))
    proj_df = pd.DataFrame([{
        "Project Name":    _safe_str(p.project_name),
        "Country":         _safe_str(p.country),
        "# Cams":          _safe_int(p.num_cams),
        "Payment Month":   _safe_str(p.payment_month),
        "Install Year":    _safe_str(p.installation_year),
        "Status":          _safe_str(p.status),
    } for p in sorted_proj])
    proj_display_df = proj_df.copy()
    if len(proj_display_df) > 0:
        proj_display_df["# Cams"] = proj_display_df["# Cams"].map(lambda value: _safe_str(value))

    def color_status(val):
        if str(val).strip().lower() == "active":
            return "color: #27AE60; font-weight: bold"
        return "color: #E74C3C"

    proj_table = proj_display_df
    if len(proj_display_df) > 0:
        proj_table = proj_display_df.style.set_properties(subset=["# Cams"], **{"text-align": "left"})
        if "Status" in proj_display_df.columns:
            proj_table = proj_table.map(color_status, subset=["Status"])

    st.dataframe(
        proj_table,
        use_container_width=True,
        hide_index=True,
        height=300,
    )

    st.markdown("---")

    # ── Trend Chart ───────────────────────────────────────────────────────────
    st.subheader("Trends")
    cc1, cc2, cc3, cc4 = st.columns([2, 2, 1, 1])
    if st.session_state.get("ch_metric") == "Cameras":
        st.session_state["ch_metric"] = "Total Cameras"
    metric     = cc1.selectbox(
        "Show",
        ["Income (Paid)", "Income (All)", "Active Projects", "Total Cameras", "Added Cameras"],
        key="ch_metric",
    )
    resolution = cc2.selectbox("Resolution", ["Yearly", "Monthly"], key="ch_res")
    all_years  = sorted({inv.year for inv in invoices if inv.year})
    if not all_years:
        all_years = [datetime.datetime.now().year]
    if resolution == "Yearly":
        from_yr = cc3.selectbox("From Year", [int(y) for y in all_years], index=0, key="ch_from")
        to_yr   = cc4.selectbox("To Year",   [int(y) for y in all_years], index=len(all_years)-1, key="ch_to")
        if from_yr > to_yr:
            from_yr, to_yr = to_yr, from_yr
        monthly_year = None
    else:
        default_monthly_year = int(sel_year) if sel_year != "All" and int(sel_year) in all_years else all_years[-1]
        monthly_year = cc3.selectbox(
            "Year",
            [int(y) for y in all_years],
            index=[int(y) for y in all_years].index(default_monthly_year),
            key="ch_monthly_year",
        )
        cc4.empty()
        from_yr = monthly_year
        to_yr = monthly_year

    is_income = metric.startswith("Income")
    y_label   = "EUR (€)" if is_income else "Count"

    def _project_start_date(project):
        if project.activation_date:
            if isinstance(project.activation_date, datetime.datetime):
                return project.activation_date.date()
            return project.activation_date
        if project.installation_year:
            return datetime.date(int(project.installation_year), 1, 1)
        return None

    if resolution == "Yearly":
        labels, values = [], []
        for yr in range(from_yr, to_yr + 1):
            labels.append(str(yr))
            if metric == "Income (Paid)":
                v = sum(i.payment_amount for i in invoices if i.is_paid() and i.year == yr)
            elif metric == "Income (All)":
                v = sum(i.payment_amount for i in invoices if i.year == yr)
            elif metric == "Active Projects":
                v = sum(1 for p in projects if p.installation_year and p.installation_year <= yr and p.is_active())
            elif metric == "Added Cameras":
                v = sum(
                    p.num_cams
                    for p in projects
                    if p.is_active() and _project_start_date(p) and _project_start_date(p).year == yr
                )
            else:
                v = sum(
                    p.num_cams
                    for p in projects
                    if p.is_active() and _project_start_date(p) and _project_start_date(p).year <= yr
                )
            values.append(float(v))

        fig = px.bar(
            x=labels, y=values,
            labels={"x": "Year", "y": y_label},
            title=f"{metric} — Yearly ({from_yr}–{to_yr})",
            color_discrete_sequence=["#2980B9"],
        )
        fig.update_traces(hovertemplate="<b>%{x}</b><br>" + y_label + ": %{y:,.0f}<extra></extra>")
        fig.update_layout(showlegend=False, height=380)
        fig.update_xaxes(
            type="category",
            categoryorder="array",
            categoryarray=labels,
            tickmode="array",
            tickvals=labels,
            ticktext=labels,
        )

    else:  # Monthly
        rows = []
        for mo in range(1, 13):
            month_end = datetime.date(monthly_year, mo, calendar.monthrange(monthly_year, mo)[1])
            if metric == "Income (Paid)":
                v = sum(i.payment_amount for i in invoices
                        if i.is_paid() and i.payment_date
                        and i.payment_date.year == monthly_year and i.payment_date.month == mo)
            elif metric == "Income (All)":
                v = sum(i.payment_amount for i in invoices
                        if i.payment_date
                        and i.payment_date.year == monthly_year and i.payment_date.month == mo)
            elif metric == "Active Projects":
                v = sum(1 for p in projects if p.installation_year and p.installation_year <= monthly_year and p.is_active())
            elif metric == "Added Cameras":
                v = sum(
                    p.num_cams
                    for p in projects
                    if p.is_active()
                    and _project_start_date(p)
                    and _project_start_date(p).year == monthly_year
                    and _project_start_date(p).month == mo
                )
            else:
                v = sum(
                    p.num_cams
                    for p in projects
                    if p.is_active() and _project_start_date(p) and _project_start_date(p) <= month_end
                )
            rows.append({"date": datetime.date(monthly_year, mo, 1), "value": float(v)})

        df_line = pd.DataFrame(rows)
        fig = px.line(
            df_line, x="date", y="value",
            labels={"date": "Month", "value": y_label},
            title=f"{metric} — Monthly ({monthly_year})",
            color_discrete_sequence=["#2980B9"],
            markers=True,
        )
        fig.update_traces(hovertemplate="<b>%{x|%b %Y}</b><br>" + y_label + ": %{y:,.0f}<extra></extra>")
        fig.update_layout(height=380)

    fig.update_layout(dragmode=False)
    fig.update_xaxes(fixedrange=True)
    fig.update_yaxes(fixedrange=True)

    st.plotly_chart(
        fig,
        use_container_width=True,
        config={
            "scrollZoom": False,
            "displaylogo": False,
            "doubleClick": "reset",
            "modeBarButtonsToRemove": [
                "zoom2d",
                "pan2d",
                "select2d",
                "lasso2d",
                "zoomIn2d",
                "zoomOut2d",
                "autoScale2d",
            ],
        },
    )


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: ASK DATA
# ══════════════════════════════════════════════════════════════════════════════
elif page == "❓ Ask Data":
    st.title("❓ Ask Data")
    st.caption("Ask questions about projects, invoices, debt, or sent PDF invoices.")

    quick_question_cols = st.columns(4)
    quick_questions = [
        "What is the Y1 debt for 2026?",
        "Show unpaid invoices for AD Denderleeuw",
        "Show invoice 8669",
        "Show sent PDF invoices for 2026",
    ]
    for idx, quick_question in enumerate(quick_questions):
        if quick_question_cols[idx].button(quick_question, key=f"ask_quick_{idx}", use_container_width=True):
            st.session_state["ask_data_question"] = quick_question
            answer_text, answer_df = _answer_data_question(quick_question, projects, invoices, debt_summaries)
            st.session_state["ask_data_answer_text"] = answer_text
            st.session_state["ask_data_answer_df"] = answer_df.to_dict(orient="records") if answer_df is not None else None
            st.rerun()

    with st.form("ask_data_form"):
        question = st.text_area(
            "Question",
            value=st.session_state.get("ask_data_question", ""),
            height=100,
            placeholder="Examples: What is the Y2+ debt for 2026? Show unpaid invoices for AD Denderleeuw. Show invoice 8676. Which projects are billed in April?",
        )
        submitted = st.form_submit_button("Ask")

    if submitted:
        st.session_state["ask_data_question"] = question
        answer_text, answer_df = _answer_data_question(question, projects, invoices, debt_summaries)
        st.session_state["ask_data_answer_text"] = answer_text
        st.session_state["ask_data_answer_df"] = answer_df.to_dict(orient="records") if answer_df is not None else None

    answer_text = st.session_state.get("ask_data_answer_text")
    answer_rows = st.session_state.get("ask_data_answer_df")

    if answer_text:
        st.success(answer_text)
    else:
        st.info(
            "Try: 'What is the Y1 debt for 2026?', 'Show unpaid invoices for Proxy Muizen', "
            "'How many active projects are in Belgium?', 'Show invoice 8676', or 'Show sent PDF invoices'."
        )

    if answer_rows:
        st.dataframe(pd.DataFrame(answer_rows), use_container_width=True, hide_index=True, height=420)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: PROJECTS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🏗️ Projects":
    st.title("🏗️ Projects")
    if page_flash_success:
        st.success(page_flash_success)

    def _parse_project_date(value):
        if value is None or value == "":
            return None
        # pandas NaT is a subclass of datetime.datetime, so detect and reject it.
        try:
            import pandas as _pd
            if value is _pd.NaT or (hasattr(_pd, "isna") and _pd.isna(value)):
                return None
        except Exception:
            pass
        if isinstance(value, datetime.datetime):
            return value
        if isinstance(value, datetime.date):
            return datetime.datetime.combine(value, datetime.time.min)
        try:
            return datetime.datetime.strptime(str(value), "%Y-%m-%d")
        except Exception:
            return None

    def _normalize_project_change_value(field_name: str, value) -> str:
        if field_name == "num_cams":
            return str(_safe_int(value, default=0))
        if field_name == "payment_month":
            return normalize_month(_safe_str(value).strip())
        if field_name == "status":
            return _safe_str(value).strip()
        return _safe_str(value).strip()

    def _build_project_change_entries(before_map: dict, after_projects: list, source_name: str) -> list[dict]:
        tracked_fields = {
            "status": "Status",
            "num_cams": "# Cams",
            "payment_month": "Payment Month",
        }
        entries: list[dict] = []
        for project in after_projects:
            project_name = _safe_str(getattr(project, "project_name", "")).strip()
            if not project_name:
                continue
            before_project = before_map.get(project_name)
            if before_project is None:
                continue
            for field_name, label in tracked_fields.items():
                old_value = _normalize_project_change_value(field_name, getattr(before_project, field_name, ""))
                new_value = _normalize_project_change_value(field_name, getattr(project, field_name, ""))
                if old_value == new_value:
                    continue
                entries.append({
                    "project_name": project_name,
                    "country": _safe_str(getattr(project, "country", "")).strip() or None,
                    "field_name": label,
                    "old_value": old_value or None,
                    "new_value": new_value or None,
                    "source_name": source_name,
                    "notes": "Updated via Projects page",
                })
        return entries

    # Filters
    col1, col2, col3 = st.columns(3)
    countries = sorted({p.country for p in projects if p.country})
    project_statuses = sorted({p.status for p in projects if p.status})
    install_year_options = [""] + [str(year) for year in range(2030, 2013, -1)]
    payment_month_options = [""] + MONTH_ORDER
    sel_country = col1.selectbox("Country", ["All"] + countries, key="proj_country")
    sel_status  = col2.selectbox("Status",  ["All", "Active", "Offline"], key="proj_status")
    search      = col3.text_input("Search project name", key="proj_search")

    filtered = projects
    if sel_country != "All":
        filtered = [p for p in filtered if p.country == sel_country]
    if sel_status == "Active":
        filtered = [p for p in filtered if p.is_active()]
    elif sel_status == "Offline":
        filtered = [p for p in filtered if not p.is_active()]
    if search:
        filtered = [p for p in filtered if search.lower() in p.project_name.lower()]

    sort_row_1, sort_row_2 = st.columns(2)
    sort_options = ["Activation Date", "Install Year", "Project Name", "Country", "# Cams", "Payment Month", "Status", "License EOP"]
    selected_sort_column = sort_row_1.selectbox("Sort by", sort_options, index=0, key="proj_sort_column")
    selected_sort_order = sort_row_2.selectbox("Sort order", ["Descending", "Ascending"], index=0, key="proj_sort_order")

    st.caption(f"Showing {len(filtered)} of {len(projects)} projects")

    df = pd.DataFrame([{
        "_original_project_name": _safe_str(p.project_name),
        "Project Name":    _safe_str(p.project_name),
        "Country":         _safe_str(p.country),
        "# Cams":          _safe_int(p.num_cams),
        "Payment Month":   _safe_str(p.payment_month),
        "Install Year":    _safe_str(p.installation_year),
        "Activation Date": p.activation_date.date() if p.activation_date else None,
        "Status":          _safe_str(p.status),
        "License EOP":     p.license_eop.date() if p.license_eop else None,
    } for p in filtered])

    if not df.empty:
        df["Activation Date"] = pd.to_datetime(df["Activation Date"], errors="coerce")
        df["License EOP"] = pd.to_datetime(df["License EOP"], errors="coerce")
        ascending = selected_sort_order == "Ascending"
        sort_column = selected_sort_column
        if sort_column == "Install Year":
            # Keep UI values as strings for SelectboxColumn, but sort numerically.
            df["_sort_install_year"] = pd.to_numeric(df["Install Year"], errors="coerce")
            sort_column = "_sort_install_year"
        if sort_column in df.columns:
            df = df.sort_values(
                by=sort_column,
                ascending=ascending,
                na_position="last",
                kind="mergesort",
            )
        if "_sort_install_year" in df.columns:
            df = df.drop(columns=["_sort_install_year"])
        df["Activation Date"] = df["Activation Date"].dt.date
        df["License EOP"] = df["License EOP"].dt.date

    def color_status(val):
        if str(val).strip().lower() == "active":
            return "color: #27AE60; font-weight: bold"
        return "color: #E74C3C"

    if CAN_EDIT:
        st.info("✏️ Admin mode: you can edit cells directly. Click **Save Changes** when done.")

        with st.expander("📥 Import Project Order", expanded=False):
            uploaded_project_order = st.file_uploader(
                "Upload project order (PDF, XLSX, or CSV)",
                type=["pdf", "xlsx", "xlsm", "xls", "csv"],
                key="project_order_upload",
                help="Upload a text-based order PDF or spreadsheet to add new projects into the Projects list.",
            )
            if uploaded_project_order is not None:
                try:
                    import_meta, import_rows = _parse_uploaded_project_order(
                        uploaded_project_order.getvalue(),
                        uploaded_project_order.name,
                    )
                    preview_df = pd.DataFrame([
                        {
                            "Project Name": row["project_name"],
                            "Country": row["country"],
                            "# Cams": row["num_cams"],
                            "Payment Month": row["payment_month"],
                            "Install Year": row["installation_year"] or "",
                            "Activation Date": row["activation_date"].date().isoformat() if row["activation_date"] else "",
                            "Status": row["status"],
                        }
                        for row in import_rows
                    ])
                    st.caption(
                        f"Parsed {import_meta['row_count']} row(s) from {import_meta['filename']}. "
                        f"Detected columns: {', '.join(import_meta['columns_found']) or 'project name only'}."
                    )
                    if preview_df.empty:
                        st.warning("No project rows were found in the uploaded order file.")
                    else:
                        st.dataframe(preview_df, use_container_width=True, hide_index=True, height=260)

                        normalized_import_names = [row["project_name"].strip().lower() for row in import_rows]
                        duplicate_import_names = sorted({
                            row["project_name"]
                            for row in import_rows
                            if normalized_import_names.count(row["project_name"].strip().lower()) > 1
                        })
                        existing_project_names = {
                            _normalize_project_name_key(project.project_name)
                            for project in projects
                            if _safe_str(project.project_name).strip()
                        }
                        new_project_rows = [
                            row for row in import_rows
                            if not _project_name_matches(row["project_name"], existing_project_names)
                        ]
                        skipped_existing = len(import_rows) - len(new_project_rows)

                        if duplicate_import_names:
                            st.error(
                                "The uploaded order contains duplicate project names: "
                                + ", ".join(duplicate_import_names)
                            )
                        elif not new_project_rows:
                            st.info("All uploaded projects already exist in the Projects list.")
                        else:
                            if skipped_existing:
                                st.info(f"{skipped_existing} existing project row(s) will be skipped.")
                            if st.button("Add New Projects From Order", key="import_project_order_btn", type="primary"):
                                from models.project import Project as ProjectModel

                                added_count = 0
                                for row in new_project_rows:
                                    projects.append(ProjectModel(
                                        project_name=row["project_name"],
                                        country=row["country"],
                                        num_cams=row["num_cams"],
                                        payment_month=row["payment_month"],
                                        installation_year=row["installation_year"],
                                        activation_date=row["activation_date"],
                                        status=row["status"],
                                        license_eop=row["license_eop"],
                                    ))
                                    added_count += 1
                                try:
                                    _save_projects(projects, _data_path)
                                    load_data.clear()
                                    st.session_state["_flash_success"] = (
                                        f"Imported {added_count} new project(s) from the uploaded order."
                                        + (f" Skipped {skipped_existing} existing row(s)." if skipped_existing else "")
                                    )
                                    st.session_state["_flash_success_page"] = "🏗️ Projects"
                                    st.rerun()
                                except Exception as exc:
                                    st.error(f"Import failed: {exc}")
                except Exception as exc:
                    st.error(f"Failed to parse uploaded project order: {exc}")

        proj_control_col1, proj_control_col2, _proj_control_spacer = st.columns([1, 1, 3])
        with proj_control_col1:
            if st.button("➕ Add New Project", key="btn_add_proj"):
                st.session_state["add_proj_row"] = 1
        with proj_control_col2:
            save_projects_clicked = st.button("💾 Save Changes", key="save_projects_top")

        _empty_proj = {"_original_project_name": "", "Project Name": "", "Country": "", "# Cams": 0,
                       "Payment Month": "", "Install Year": "",
                       "Activation Date": None, "Status": "Active", "License EOP": None}
        n_new = st.session_state.get("add_proj_row", 0)
        if n_new:
            empty_rows = pd.DataFrame([_empty_proj] * n_new)
            df_edit = pd.concat([empty_rows, df.reset_index(drop=True)], ignore_index=True)
        else:
            df_edit = df.reset_index(drop=True)

        edited_df = st.data_editor(
            df_edit,
            use_container_width=True,
            height=600,
            num_rows="dynamic",
            column_config={
                "_original_project_name": None,
                "Country": st.column_config.SelectboxColumn(
                    "Country",
                    options=[""] + countries,
                ),
                "Install Year": st.column_config.SelectboxColumn(
                    "Install Year",
                    options=install_year_options,
                ),
                "Payment Month": st.column_config.SelectboxColumn(
                    "Payment Month",
                    options=payment_month_options,
                ),
                "Activation Date": st.column_config.DateColumn(
                    "Activation Date",
                    format="YYYY-MM-DD",
                ),
                "Status": st.column_config.SelectboxColumn(
                    "Status",
                    options=([""] + project_statuses) if project_statuses else ["", "Active", "Offline"],
                ),
                "License EOP": st.column_config.DateColumn(
                    "License EOP",
                    format="YYYY-MM-DD",
                ),
            },
            key="proj_editor",
        )
        if save_projects_clicked:
            from models.project import Project as ProjectModel
            original_project_map = {
                _safe_str(p.project_name).strip(): p
                for p in projects
                if _safe_str(p.project_name).strip()
            }
            before_projects_map = {
                name: copy.deepcopy(project)
                for name, project in original_project_map.items()
            }
            visible_original_names = {
                _safe_str(p.project_name).strip()
                for p in filtered
                if _safe_str(p.project_name).strip()
            }
            preserved_original_names = set()
            projects_to_save = []
            delete_project_names = []
            renamed_projects = {}
            new_count = 0
            for _, row in edited_df.iterrows():
                name = _safe_str(row.get("Project Name", "")).strip()
                if not name:
                    continue
                original_name = _safe_str(row.get("_original_project_name", "")).strip()
                if original_name:
                    preserved_original_names.add(original_name)
                if original_name and original_name in original_project_map:
                    p = original_project_map[original_name]
                else:
                    p = ProjectModel(project_name=name)
                    new_count += 1
                old_name = _safe_str(p.project_name).strip()
                p.project_name = name
                p.country           = _safe_str(row.get("Country", ""))
                p.num_cams          = _safe_int(row.get("# Cams", 0))
                p.payment_month     = _safe_str(row.get("Payment Month", ""))
                p.installation_year = _safe_int(row.get("Install Year")) or None
                p.status            = _safe_str(row.get("Status", ""))
                p.activation_date   = _parse_project_date(row.get("Activation Date"))
                p.license_eop       = _parse_project_date(row.get("License EOP"))
                projects_to_save.append(p)
                if original_name and old_name and old_name != name:
                    delete_project_names.append(old_name)
                    renamed_projects[old_name] = name

            removed_visible_names = visible_original_names - preserved_original_names
            delete_project_names.extend(sorted(name for name in removed_visible_names if name))

            remaining_projects = [
                p for p in projects
                if _safe_str(p.project_name).strip() not in visible_original_names
            ]
            projects[:] = remaining_projects + projects_to_save
            try:
                project_change_entries = _build_project_change_entries(
                    before_projects_map,
                    projects_to_save,
                    _data_path,
                )
                _save_projects(projects, _data_path)
                _rename_invoice_project_names(renamed_projects, _data_path)
                _delete_projects(sorted({name for name in delete_project_names if name}), _data_path)
                for entry in project_change_entries:
                    append_project_change_log(entry)
                load_data.clear()
                st.session_state.pop("add_proj_row", None)
                st.session_state.pop("proj_editor", None)
                msg = f"Saved! {new_count} new project(s) added." if new_count else "Projects saved successfully!"
                st.session_state["_flash_success"] = msg
                st.session_state["_flash_success_page"] = "🏗️ Projects"
                st.rerun()
            except Exception as e:
                st.error(f"Save failed: {e}")

        st.markdown("---")
        st.subheader("Delete Project")
        project_delete_options = sorted({_safe_str(p.project_name).strip() for p in projects if _safe_str(p.project_name).strip()})
        if project_delete_options:
            with st.form("delete_project_form"):
                selected_project_to_delete = st.selectbox(
                    "Project to delete",
                    project_delete_options,
                    key="project_delete_name",
                )
                delete_project_btn = st.form_submit_button("🗑️ Delete Project", type="secondary")

            if delete_project_btn:
                try:
                    deleted = _delete_projects([selected_project_to_delete], _data_path)
                    load_data.clear()
                    st.session_state["_flash_success"] = (
                        f"Deleted {deleted} project row(s) for {selected_project_to_delete}."
                    )
                    st.session_state["_flash_success_page"] = "🏗️ Projects"
                    st.rerun()
                except Exception as e:
                    st.error(f"Delete failed: {e}")

        st.markdown("---")
        st.subheader("Merge / Rename Project")
        project_merge_options = sorted({_safe_str(p.project_name).strip() for p in projects if _safe_str(p.project_name).strip()})
        if project_merge_options:
            with st.form("merge_project_form"):
                merge_source_name = st.selectbox(
                    "Old project name",
                    project_merge_options,
                    key="project_merge_source",
                )
                merge_target_name = st.text_input(
                    "New / target project name",
                    key="project_merge_target",
                    placeholder="Proxy Kluisbergen",
                )
                merge_project_btn = st.form_submit_button("Rename Across Projects + Invoices", type="primary")

            if merge_project_btn:
                target_name = _safe_str(merge_target_name).strip()
                if not target_name:
                    st.error("Enter the new / target project name.")
                elif target_name == merge_source_name:
                    st.error("Choose a different target name.")
                else:
                    try:
                        source_project = next(
                            (project for project in projects if _safe_str(project.project_name).strip() == merge_source_name),
                            None,
                        )
                        if source_project is None:
                            st.error(f"Project {merge_source_name} was not found.")
                        else:
                            target_exists = any(
                                _safe_str(project.project_name).strip() == target_name
                                for project in projects
                            )
                            if not target_exists:
                                source_project.project_name = target_name
                                _save_projects(projects, _data_path)
                            _rename_invoice_project_names({merge_source_name: target_name}, _data_path)
                            _delete_projects([merge_source_name], _data_path)
                            load_data.clear()
                            st.session_state["_flash_success"] = (
                                f"Renamed {merge_source_name} to {target_name} across projects and invoices."
                            )
                            st.session_state["_flash_success_page"] = "🏗️ Projects"
                            st.rerun()
                    except Exception as e:
                        st.error(f"Merge/rename failed: {e}")

        st.markdown("---")
        st.subheader("Project Change Log")
        project_change_rows = load_project_change_log()
        if project_change_rows:
            log_rows = []
            for row in project_change_rows:
                changed_at = _parse_optional_datetime(row.get("changed_at"))
                log_rows.append({
                    "Changed At": changed_at.strftime("%Y-%m-%d %H:%M") if changed_at else _safe_str(row.get("changed_at")),
                    "Project": _safe_str(row.get("project_name")),
                    "Country": _safe_str(row.get("country")),
                    "Field": _safe_str(row.get("field_name")),
                    "Old Value": _safe_str(row.get("old_value")),
                    "New Value": _safe_str(row.get("new_value")),
                    "Source": _safe_str(row.get("source_name")),
                })

            project_change_df = pd.DataFrame(log_rows)
            pcf1, pcf2, pcf3, pcf4 = st.columns(4)
            field_options = ["All"] + sorted({row["Field"] for row in log_rows if row["Field"]})
            country_options = ["All"] + sorted({row["Country"] for row in log_rows if row["Country"]})
            selected_field = pcf1.selectbox("Field", field_options, key="project_change_field")
            selected_country = pcf2.selectbox("Country", country_options, key="project_change_country")
            selected_project_search = pcf3.text_input("Search project", key="project_change_search")
            selected_value_search = pcf4.text_input("Search value", key="project_change_value_search")

            if selected_field != "All":
                project_change_df = project_change_df[project_change_df["Field"] == selected_field]
            if selected_country != "All":
                project_change_df = project_change_df[project_change_df["Country"] == selected_country]
            if selected_project_search.strip():
                needle = selected_project_search.strip().lower()
                project_change_df = project_change_df[
                    project_change_df["Project"].str.lower().str.contains(needle, na=False)
                ]
            if selected_value_search.strip():
                needle = selected_value_search.strip().lower()
                project_change_df = project_change_df[
                    project_change_df["Old Value"].str.lower().str.contains(needle, na=False)
                    | project_change_df["New Value"].str.lower().str.contains(needle, na=False)
                ]

            st.dataframe(
                project_change_df,
                use_container_width=True,
                hide_index=True,
                height=280,
            )
        else:
            st.info("No project changes are logged yet.")
    else:
        st.dataframe(
            df.style.map(color_status, subset=["Status"]) if "Status" in df.columns and len(df) > 0 else df,
            use_container_width=True,
            height=600,
        )

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: ORDERS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📦 Orders":
    st.title("📦 Orders")
    if page_flash_success:
        st.success(page_flash_success)

    try:
        orders, orders_source_name = load_orders_data(_data_path)
    except Exception as exc:
        st.error(f"Failed to load orders: {exc}")
        st.stop()

    if orders_source_name != "Supabase":
        st.info(
            "Orders are currently stored in local JSON fallback mode. "
            "Run the SQL in migrations/create_orders.sql to store orders centrally in Supabase."
        )

    project_name_keys = {
        _normalize_project_name_key(canonical_project_name(project.project_name))
        for project in projects
        if _safe_str(project.project_name).strip()
    } | {
        _normalize_project_name_key(project.project_name)
        for project in projects
        if _safe_str(project.project_name).strip()
    }
    project_lookup = {
        _normalize_project_name_key(project.project_name): project
        for project in projects
        if _safe_str(project.project_name).strip()
    }
    project_name_choices = [
        _safe_str(project.project_name).strip()
        for project in projects
        if _safe_str(project.project_name).strip()
    ]
    install_year_options = [""] + [str(year) for year in range(2030, 2013, -1)]

    total_orders = len(orders)
    open_orders = sum(
        1 for order in orders
        if _normalize_order_status(order.get("status", "")) in {"New", "Ordered", "In Progress"}
    )
    missing_project_orders = [
        order for order in orders
        if not _project_name_matches(order.get("project_name"), project_name_keys)
    ]
    total_ordered_cameras = sum(_safe_int(order.get("ordered_cameras"), default=0) for order in orders)

    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("Total Orders", total_orders)
    mc2.metric("Open Orders", open_orders)
    mc3.metric("Missing Projects", len(missing_project_orders))
    mc4.metric("Ordered Cameras", total_ordered_cameras)

    if CAN_EDIT:
        with st.expander("📥 Import Orders", expanded=False):
            uploaded_order_files = st.file_uploader(
                "Upload order file(s) (PDF, XLSX, CSV, or ZIP)",
                type=["pdf", "xlsx", "xlsm", "xls", "csv", "zip"],
                key="orders_upload",
                accept_multiple_files=True,
                help="Upload one or more order files or ZIP archives. Each parsed row becomes a tracked order record.",
            )
            if uploaded_order_files:
                try:
                    expanded_order_sources = _expand_uploaded_order_sources(uploaded_order_files)
                except Exception as exc:
                    st.error(f"Could not read uploaded order files: {exc}")
                    expanded_order_sources = []

                multi_source_import = len(expanded_order_sources) > 1
                default_order_ref = (
                    _default_order_reference(expanded_order_sources[0]["source_name"])
                    if len(expanded_order_sources) == 1 else ""
                )
                ic1, ic2, ic3 = st.columns(3)
                if multi_source_import:
                    ic1.caption("Review and edit the Order Ref column below before importing.")
                    import_order_ref = ""
                else:
                    import_order_ref = ic1.text_input("Order reference", value=default_order_ref, key="orders_import_ref")
                import_order_date = ic2.date_input("Order date", value=datetime.date.today(), key="orders_import_date")
                import_order_status = ic3.selectbox(
                    "Imported status",
                    ORDER_STATUS_OPTIONS,
                    index=ORDER_STATUS_OPTIONS.index("Ordered"),
                    key="orders_import_status",
                )

                parsed_sources = []
                parse_errors = []
                preview_rows = []
                for source_index, order_source in enumerate(expanded_order_sources):
                    try:
                        import_meta, import_rows = _parse_uploaded_project_order(
                            order_source["file_bytes"],
                            order_source["source_name"],
                        )
                        order_reference = (
                            _default_order_reference(order_source["source_name"])
                            if multi_source_import else _safe_str(import_order_ref).strip()
                        )
                        source_key = f"{order_source['container_name']}::{order_source['source_path']}"
                        archive_name = re.sub(r"[\\/]+", "_", _safe_str(order_source["source_path"]).strip()) or order_source["source_name"]
                        parsed_sources.append({
                            "source_key": source_key,
                            "source_name": order_source["source_name"],
                            "source_path": order_source["source_path"],
                            "container_name": order_source["container_name"],
                            "archive_name": archive_name,
                            "file_bytes": order_source["file_bytes"],
                            "meta": import_meta,
                            "rows": import_rows,
                            "default_order_reference": order_reference,
                        })
                        for row_index, row in enumerate(import_rows):
                            existing_project = project_lookup.get(row["project_name"].strip().lower())
                            resolved_country = row["country"] or (existing_project.country if existing_project else "") or _guess_order_country(row["project_name"])
                            resolved_ordered_cams = row["num_cams"] or (_safe_int(existing_project.num_cams, default=0) if existing_project else 0)
                            resolved_payment_month = row["payment_month"] or (_safe_str(existing_project.payment_month).strip() if existing_project else "")
                            resolved_install_year = row["installation_year"] or (existing_project.installation_year if existing_project else None)
                            preview_rows.append({
                                "_preview_id": f"{source_key}::{row_index}",
                                "_source_key": source_key,
                                "File": order_source["source_path"],
                                "Order Ref": order_reference,
                                "Project": row["project_name"],
                                "Country": resolved_country,
                                "Ordered Cams": resolved_ordered_cams,
                                "Payment Amount": _safe_float(row.get("payment_amount"), default=0.0),
                                "Payment Month": resolved_payment_month,
                                "Install Year": resolved_install_year or "",
                                "Requested Activation": row["activation_date"].date().isoformat() if row["activation_date"] else "",
                            })
                    except Exception as exc:
                        parse_errors.append(f"{order_source['source_path']}: {exc}")

                if parse_errors:
                    st.warning("Some files could not be parsed:\n" + "\n".join(parse_errors))

                if parsed_sources:
                    parsed_file_count = len(parsed_sources)
                    parsed_row_count = sum(file_info["meta"]["row_count"] for file_info in parsed_sources)
                    detected_columns = sorted({
                        column_name
                        for file_info in parsed_sources
                        for column_name in file_info["meta"]["columns_found"]
                    })
                    st.caption(
                        f"Parsed {parsed_row_count} row(s) from {parsed_file_count} file(s). "
                        f"Detected columns: {', '.join(detected_columns) or 'project name only'}."
                    )
                    preview_df = pd.DataFrame(preview_rows)
                    if preview_df.empty:
                        st.warning("No order rows were found in the uploaded files.")
                    else:
                        existing_project_names = sorted({
                            _safe_str(p.project_name).strip()
                            for p in projects
                            if _safe_str(p.project_name).strip()
                        })
                        if existing_project_names:
                            with st.expander("Existing project names (for reference)"):
                                st.write(", ".join(existing_project_names))
                        reviewed_preview_df = st.data_editor(
                            preview_df,
                            use_container_width=True,
                            hide_index=True,
                            height=260,
                            column_config={
                                "_preview_id": None,
                                "_source_key": None,
                                "Project": st.column_config.TextColumn("Project", required=True),
                                "Ordered Cams": st.column_config.NumberColumn("Ordered Cams", min_value=0, step=1),
                                "Payment Amount": st.column_config.NumberColumn("Payment Amount", min_value=0.0, step=1.0, format="€ %.2f"),
                            },
                            disabled=[
                                "File",
                            ],
                            key="orders_import_review",
                        )
                        if reviewed_preview_df["Order Ref"].astype(str).str.strip().eq("").any():
                            st.error("Every imported row needs an Order Ref before import.")
                        else:
                            existing_order_keys = {
                                (
                                    _safe_str(order.get("order_number")).strip().lower(),
                                    _safe_str(order.get("project_name")).strip().lower(),
                                )
                                for order in orders
                                if _safe_str(order.get("project_name")).strip()
                            }
                            source_lookup = {
                                file_info["source_key"]: file_info
                                for file_info in parsed_sources
                            }
                            reviewed_rows_by_id = {
                                _safe_str(row.get("_preview_id")).strip(): row
                                for row in reviewed_preview_df.to_dict(orient="records")
                            }
                            import_rows_to_create = []
                            skipped_existing = 0
                            for file_info in parsed_sources:
                                for row_index, row in enumerate(file_info["rows"]):
                                    preview_id = f"{file_info['source_key']}::{row_index}"
                                    reviewed_row = reviewed_rows_by_id.get(preview_id, {})
                                    order_ref_value = _safe_str(reviewed_row.get("Order Ref", file_info["default_order_reference"])) .strip()
                                    if not order_ref_value:
                                        continue
                                    edited_project_name = _safe_str(reviewed_row.get("Project", row["project_name"])).strip() or row["project_name"]
                                    project_key = edited_project_name.strip().lower()
                                    record_key = (order_ref_value.lower(), project_key)
                                    if record_key in existing_order_keys:
                                        skipped_existing += 1
                                        continue
                                    existing_order_keys.add(record_key)
                                    import_rows_to_create.append({
                                        "order_number": order_ref_value,
                                        "project_name": edited_project_name,
                                        "country": _safe_str(reviewed_row.get("Country", row["country"])).strip(),
                                        "ordered_cameras": _safe_int(reviewed_row.get("Ordered Cams", row["num_cams"]), default=0),
                                        "payment_amount": _safe_float(reviewed_row.get("Payment Amount", row.get("payment_amount", 0.0)), default=0.0),
                                        "payment_month": _safe_str(reviewed_row.get("Payment Month", row["payment_month"])).strip(),
                                        "installation_year": _safe_int(reviewed_row.get("Install Year", row["installation_year"])),
                                        "order_date": import_order_date,
                                        "requested_activation_date": _parse_order_date(reviewed_row.get("Requested Activation", row["activation_date"])),
                                        "status": import_order_status,
                                        "notes": f"Imported from {file_info['source_path']}",
                                        "source_filename": file_info["source_path"],
                                        "_archive_name": file_info["archive_name"],
                                        "_source_key": file_info["source_key"],
                                    })

                            if not import_rows_to_create:
                                st.info("All parsed order rows already exist for these order references.")
                            else:
                                if skipped_existing:
                                    st.info(f"{skipped_existing} existing order row(s) will be skipped.")
                                if multi_source_import:
                                    st.info("You can edit the Order Ref values in the review table before importing.")
                                if st.button("Import Order Rows", type="primary", key="import_orders_btn"):
                                    try:
                                        archive_paths = {}
                                        storage_meta_by_key = {}
                                        for file_info in parsed_sources:
                                            archive_paths[file_info["source_key"]] = _archive_uploaded_order_file(
                                                file_info["file_bytes"],
                                                file_info["archive_name"],
                                            )
                                            try:
                                                storage_meta_by_key[file_info["source_key"]] = upload_order_pdf_supabase(
                                                    file_info["file_bytes"],
                                                    file_info["archive_name"],
                                                )
                                            except Exception as exc:
                                                logger.warning(
                                                    "Could not upload order PDF to Supabase Storage: %s",
                                                    exc,
                                                )
                                        for row in import_rows_to_create:
                                            source_key = row.pop("_source_key")
                                            row["source_archive_path"] = str(archive_paths[source_key])
                                            storage_meta = storage_meta_by_key.get(source_key)
                                            if storage_meta:
                                                row["pdf_storage_bucket"] = storage_meta.get("pdf_storage_bucket")
                                                row["pdf_storage_path"] = storage_meta.get("pdf_storage_path")
                                            row.pop("_archive_name", None)
                                        created_count = _create_orders(import_rows_to_create, orders_source_name)
                                        load_orders_data.clear()
                                        st.session_state["_flash_success"] = (
                                            f"Imported {created_count} order row(s) from {parsed_file_count} file(s)."
                                            + (f" Skipped {skipped_existing} existing row(s)." if skipped_existing else "")
                                        )
                                        st.session_state["_flash_success_page"] = "📦 Orders"
                                        st.rerun()
                                    except Exception as exc:
                                        st.error(f"Import failed: {exc}")

        with st.expander("📎 Backfill Source PDFs", expanded=False):
            st.caption(
                "Upload one or more PDFs (or a ZIP). Each file is matched to an existing "
                "order row by source filename or by the order reference / project name in the filename, "
                "then uploaded to Supabase Storage."
            )
            backfill_files = st.file_uploader(
                "Upload PDFs to attach to existing orders",
                type=["pdf", "zip"],
                key="orders_backfill_upload",
                accept_multiple_files=True,
            )
            if backfill_files and st.button("Attach PDFs To Matching Orders", type="primary", key="orders_backfill_btn"):
                expanded: list[tuple[str, bytes]] = []
                for uploaded_file in backfill_files:
                    name = _safe_str(getattr(uploaded_file, "name", "")).strip()
                    if not name:
                        continue
                    data = uploaded_file.getvalue()
                    if Path(name).suffix.lower() == ".zip":
                        try:
                            with zipfile.ZipFile(io.BytesIO(data)) as archive:
                                for member in archive.infolist():
                                    if member.is_dir():
                                        continue
                                    member_name = Path(member.filename).name
                                    if Path(member_name).suffix.lower() == ".pdf":
                                        expanded.append((member_name, archive.read(member)))
                        except Exception as exc:
                            st.error(f"Could not read ZIP {name}: {exc}")
                    elif Path(name).suffix.lower() == ".pdf":
                        expanded.append((name, data))

                if not expanded:
                    st.warning("No PDF files found to attach.")
                else:
                    def _norm(text: str) -> str:
                        return re.sub(r"[^a-z0-9]+", "", _safe_str(text).lower())

                    matched_count = 0
                    skipped: list[str] = []
                    failed: list[str] = []
                    used_order_ids: set[int] = set()
                    for filename, file_bytes in expanded:
                        stem_norm = _norm(Path(filename).stem)
                        match_order = None
                        for order in orders:
                            order_id = _safe_int(order.get("id"), default=0)
                            if order_id in used_order_ids:
                                continue
                            source_norm = _norm(Path(_safe_str(order.get("source_filename"))).stem)
                            if source_norm and source_norm == stem_norm:
                                match_order = order
                                break
                        if match_order is None:
                            for order in orders:
                                order_id = _safe_int(order.get("id"), default=0)
                                if order_id in used_order_ids:
                                    continue
                                order_ref_norm = _norm(order.get("order_number"))
                                project_norm = _norm(order.get("project_name"))
                                if order_ref_norm and order_ref_norm in stem_norm and project_norm and project_norm in stem_norm:
                                    match_order = order
                                    break
                        if match_order is None:
                            skipped.append(filename)
                            continue
                        try:
                            storage_meta = upload_order_pdf_supabase(file_bytes, filename)
                            _update_order(
                                _safe_int(match_order.get("id"), default=0),
                                orders_source_name,
                                pdf_storage_bucket=storage_meta.get("pdf_storage_bucket"),
                                pdf_storage_path=storage_meta.get("pdf_storage_path"),
                                source_filename=_safe_str(match_order.get("source_filename")).strip() or filename,
                            )
                            used_order_ids.add(_safe_int(match_order.get("id"), default=0))
                            matched_count += 1
                        except Exception as exc:
                            failed.append(f"{filename}: {exc}")

                    if matched_count:
                        load_orders_data.clear()
                        st.session_state["_flash_success"] = f"Attached {matched_count} PDF(s) to existing orders."
                        st.session_state["_flash_success_page"] = "📦 Orders"
                    if skipped:
                        st.warning("No matching order found for: " + ", ".join(skipped))
                    if failed:
                        st.error("Errors:\n" + "\n".join(failed))
                    if matched_count:
                        st.rerun()

        with st.expander("➕ New Order", expanded=False):
            with st.form("new_order_form"):
                nc1, nc2, nc3 = st.columns(3)
                new_order_number = nc1.text_input("Order reference", key="new_order_number")
                new_project_name = nc2.text_input("Project name", key="new_order_project")
                new_country_options = [""] + sorted({_normalize_country(_safe_str(order.get("country")).strip()) for order in orders if _safe_str(order.get("country")).strip()} | {_normalize_country(_safe_str(p.country).strip()) for p in projects if _safe_str(p.country).strip()})
                new_country = nc3.selectbox("Country", new_country_options, key="new_order_country")

                nc4, nc5, nc6 = st.columns(3)
                new_ordered_cameras = nc4.number_input("Ordered cameras", min_value=0, step=1, key="new_order_cameras")
                new_payment_amount = nc5.number_input("Payment amount", min_value=0.0, step=1.0, key="new_order_amount")
                new_payment_month = nc6.selectbox("Payment month", [""] + MONTH_ORDER, key="new_order_payment_month")

                nc7, nc8, nc9 = st.columns(3)
                new_installation_year = nc7.selectbox("Install year", install_year_options, key="new_order_install_year")
                new_order_date = nc8.date_input("Order date", value=datetime.date.today(), key="new_order_date")
                new_has_activation = nc9.checkbox("Set requested activation", value=False, key="new_order_has_activation")

                new_requested_activation = st.date_input(
                    "Requested activation date",
                    value=datetime.date.today(),
                    disabled=not new_has_activation,
                    key="new_order_activation",
                )

                new_status = st.selectbox("Status", ORDER_STATUS_OPTIONS, index=ORDER_STATUS_OPTIONS.index("Ordered"), key="new_order_status")
                new_notes = st.text_area("Notes", key="new_order_notes", height=90)
                create_order_btn = st.form_submit_button("Create Order", type="primary")

            if create_order_btn:
                if not _safe_str(new_project_name).strip():
                    st.error("Project name is required.")
                else:
                    try:
                        created_count = _create_orders([
                            {
                                "order_number": _safe_str(new_order_number).strip(),
                                "project_name": _safe_str(new_project_name).strip(),
                                "country": _safe_str(new_country).strip(),
                                "ordered_cameras": int(new_ordered_cameras),
                                "payment_amount": float(new_payment_amount),
                                "payment_month": _safe_str(new_payment_month).strip(),
                                "installation_year": _safe_int(new_installation_year) or None,
                                "order_date": new_order_date,
                                "requested_activation_date": new_requested_activation if new_has_activation else None,
                                "status": new_status,
                                "notes": _safe_str(new_notes),
                                "source_filename": "",
                                "source_archive_path": "",
                            }
                        ], orders_source_name)
                        load_orders_data.clear()
                        st.session_state["_flash_success"] = f"Created {created_count} order row(s)."
                        st.session_state["_flash_success_page"] = "📦 Orders"
                        st.rerun()
                    except Exception as exc:
                        st.error(f"Create failed: {exc}")

    st.markdown("---")
    with st.expander("🔍 Filters", expanded=True):
        fc1, fc2, fc3, fc4 = st.columns(4)
        status_options = ["All"] + sorted({_normalize_order_status(order.get("status", "")) for order in orders if _safe_str(order.get("status"))})
        country_options = ["All"] + sorted({_safe_str(order.get("country")).strip() for order in orders if _safe_str(order.get("country")).strip()})
        order_status_filter = fc1.selectbox("Status", status_options, key="orders_filter_status")
        order_country_filter = fc2.selectbox("Country", country_options, key="orders_filter_country")
        order_search = fc3.text_input("Search project / order", key="orders_filter_search")
        missing_only = fc4.checkbox("Only missing projects", key="orders_filter_missing")

    filtered_orders = orders
    if order_status_filter != "All":
        filtered_orders = [order for order in filtered_orders if _normalize_order_status(order.get("status", "")) == order_status_filter]
    if order_country_filter != "All":
        filtered_orders = [order for order in filtered_orders if _safe_str(order.get("country")).strip() == order_country_filter]
    if order_search.strip():
        order_search_lower = order_search.lower().strip()
        filtered_orders = [
            order for order in filtered_orders
            if order_search_lower in _safe_str(order.get("project_name")).lower()
            or order_search_lower in _safe_str(order.get("order_number")).lower()
        ]
    if missing_only:
        filtered_orders = [
            order for order in filtered_orders
            if not _project_name_matches(order.get("project_name"), project_name_keys)
        ]

    st.caption(f"Showing {len(filtered_orders)} of {len(orders)} order row(s)")
    if filtered_orders:
        order_pdf_links: dict[int, str | None] = {}
        for order in filtered_orders:
            order_id = _safe_int(order.get("id"), default=0)
            storage_bucket = _safe_str(order.get("pdf_storage_bucket")).strip()
            storage_path = _safe_str(order.get("pdf_storage_path")).strip()
            link: str | None = None
            if storage_bucket and storage_path:
                link = create_order_pdf_signed_url_supabase(storage_bucket, storage_path)
            order_pdf_links[order_id] = link

        orders_df = pd.DataFrame([
            {
                "Order": _safe_str(order.get("order_number")),
                "Project": _safe_str(order.get("project_name")),
                "Suggested Match": _suggest_best_project_match(order.get("project_name"), project_name_choices)[0],
                "Country": _safe_str(order.get("country")),
                "Ordered Cams": _safe_int(order.get("ordered_cameras"), default=0),
                "Payment Amount": _safe_float(order.get("payment_amount"), default=0.0),
                "Install Year": _safe_int(order.get("installation_year"), default=0) or "",
                "Order Date": (_parse_order_date(order.get("order_date")) or ""),
                "Requested Activation": (_parse_order_date(order.get("requested_activation_date")) or ""),
                "Status": _normalize_order_status(order.get("status", "")),
                "Project Exists": "Yes" if _project_name_matches(order.get("project_name"), project_name_keys) else "No",
                "Source PDF": order_pdf_links.get(_safe_int(order.get("id"), default=0)) or "",
            }
            for order in filtered_orders
        ])
        st.dataframe(
            orders_df,
            use_container_width=True,
            hide_index=True,
            height=340,
            column_config={
                "Source PDF": st.column_config.LinkColumn(
                    "Source PDF",
                    help="Click to open the originally uploaded order PDF.",
                    display_text="Download",
                ),
            },
        )
    else:
        st.info("No orders match the current filters.")

    if CAN_EDIT and filtered_orders:
        st.markdown("---")
        st.subheader("Create Missing Projects")
        missing_order_options = {
            f"#{_safe_int(order.get('id'), default=0)} | {_safe_str(order.get('order_number')) or 'No Ref'} | {_safe_str(order.get('project_name'))}": order
            for order in filtered_orders
            if not _project_name_matches(order.get("project_name"), project_name_keys)
        }
        if missing_order_options:
            selected_missing_orders = st.multiselect(
                "Order rows to create as projects",
                list(missing_order_options.keys()),
                key="orders_create_projects",
            )
            if st.button("Create Missing Projects From Selected Orders", type="primary", key="orders_create_projects_btn"):
                if not selected_missing_orders:
                    st.error("Select at least one order row.")
                else:
                    from models.project import Project as ProjectModel

                    existing_project_keys = {
                        _normalize_project_name_key(project.project_name)
                        for project in projects
                        if _safe_str(project.project_name).strip()
                    }
                    added_projects = 0
                    for label in selected_missing_orders:
                        order = missing_order_options[label]
                        project_name = _safe_str(order.get("project_name")).strip()
                        if not project_name or _project_name_matches(project_name, existing_project_keys):
                            continue
                        activation_date = _parse_order_date(order.get("requested_activation_date"))
                        projects.append(ProjectModel(
                            project_name=project_name,
                            country=_safe_str(order.get("country")).strip(),
                            num_cams=_safe_int(order.get("ordered_cameras"), default=0),
                            payment_month=_safe_str(order.get("payment_month")).strip(),
                            installation_year=_safe_int(order.get("installation_year")) or None,
                            activation_date=(
                                datetime.datetime.combine(activation_date, datetime.time.min)
                                if activation_date else None
                            ),
                            status=_project_status_from_order_status(order.get("status", "")),
                        ))
                        existing_project_keys.add(_normalize_project_name_key(project_name))
                        added_projects += 1

                    try:
                        _save_projects(projects, _data_path)
                        load_data.clear()
                        st.session_state["_flash_success"] = f"Created {added_projects} missing project(s) from orders."
                        st.session_state["_flash_success_page"] = "📦 Orders"
                        st.rerun()
                    except Exception as exc:
                        st.error(f"Project creation failed: {exc}")
        else:
            st.info("All filtered orders are already linked to projects.")

        st.markdown("---")
        st.subheader("Update Order")
        order_options = {
            f"#{_safe_int(order.get('id'), default=0)} | {_safe_str(order.get('order_number')) or 'No Ref'} | {_safe_str(order.get('project_name'))}": order
            for order in filtered_orders
        }
        selected_order_label = st.selectbox("Select order row", list(order_options.keys()), key="orders_selected_row")
        selected_order = order_options[selected_order_label]
        selected_order_id = _safe_int(selected_order.get("id"), default=0)
        field_key_suffix = f"_{selected_order_id}"

        current_order_date = _parse_order_date(selected_order.get("order_date")) or datetime.date.today()
        current_activation_date = _parse_order_date(selected_order.get("requested_activation_date"))
        current_payment_month = _safe_str(selected_order.get("payment_month")).strip()
        current_installation_year = _safe_str(selected_order.get("installation_year")).strip()
        current_payment_amount = _safe_float(selected_order.get("payment_amount"), default=0.0)
        current_status = _normalize_order_status(selected_order.get("status", ""))
        if current_status not in ORDER_STATUS_OPTIONS:
            current_status = ORDER_STATUS_OPTIONS[0]
        exact_project_match = _get_exact_existing_project_match(selected_order.get("project_name"), project_name_choices)
        if exact_project_match:
            st.caption(f"Exact project match found: {exact_project_match}")
        else:
            st.caption("No exact project match found. Use the dropdown below to pick a real project name.")

        with st.form("update_order_form"):
            uc1, uc2, uc3 = st.columns(3)
            upd_order_number = uc1.text_input("Order reference", value=_safe_str(selected_order.get("order_number")), key=f"upd_order_number{field_key_suffix}")
            upd_project_name = uc2.text_input("Project name", value=_safe_str(selected_order.get("project_name")), key=f"upd_order_project{field_key_suffix}")
            upd_country_options = [""] + sorted({_normalize_country(_safe_str(order.get("country")).strip()) for order in orders if _safe_str(order.get("country")).strip()} | {_normalize_country(_safe_str(p.country).strip()) for p in projects if _safe_str(p.country).strip()})
            current_country = _normalize_country(_safe_str(selected_order.get("country")).strip())
            upd_country_index = upd_country_options.index(current_country) if current_country in upd_country_options else 0
            upd_country = uc3.selectbox("Country", upd_country_options, index=upd_country_index, key=f"upd_order_country{field_key_suffix}")
            exact_project_picker_options = [""] + project_name_choices
            exact_project_index = 1 if exact_project_match and exact_project_match in project_name_choices else 0
            upd_project_match = st.selectbox(
                "Existing project name (exact match only)",
                exact_project_picker_options,
                index=exact_project_index,
                help="This dropdown only shows project names that already exist in the Projects list.",
                key=f"upd_order_project_match{field_key_suffix}",
            )

            uc4, uc5, uc6 = st.columns(3)
            upd_ordered_cameras = uc4.number_input(
                "Ordered cameras",
                min_value=0,
                step=1,
                value=max(0, _safe_int(selected_order.get("ordered_cameras"), default=0)),
                key=f"upd_order_cameras{field_key_suffix}",
            )
            upd_payment_amount = uc5.number_input(
                "Payment amount",
                min_value=0.0,
                step=1.0,
                value=current_payment_amount,
                key=f"upd_order_amount{field_key_suffix}",
            )
            upd_payment_month = uc6.selectbox(
                "Payment month",
                [""] + MONTH_ORDER,
                index=([""] + MONTH_ORDER).index(current_payment_month) if current_payment_month in MONTH_ORDER else 0,
                key=f"upd_order_payment_month{field_key_suffix}",
            )

            uc7, uc8, uc9 = st.columns(3)
            upd_installation_year = uc7.selectbox(
                "Install year",
                install_year_options,
                index=install_year_options.index(current_installation_year) if current_installation_year in install_year_options else 0,
                key=f"upd_order_install_year{field_key_suffix}",
            )
            upd_order_date = uc8.date_input("Order date", value=current_order_date, key=f"upd_order_date{field_key_suffix}")
            upd_has_activation = uc9.checkbox(
                "Set requested activation",
                value=current_activation_date is not None,
                key=f"upd_order_has_activation{field_key_suffix}",
            )
            upd_requested_activation = st.date_input(
                "Requested activation date",
                value=current_activation_date or datetime.date.today(),
                disabled=not upd_has_activation,
                key=f"upd_order_activation{field_key_suffix}",
            )

            upd_status = st.selectbox(
                "Status",
                ORDER_STATUS_OPTIONS,
                index=ORDER_STATUS_OPTIONS.index(current_status),
                key=f"upd_order_status{field_key_suffix}",
            )
            upd_notes = st.text_area("Notes", value=_safe_str(selected_order.get("notes")), height=100, key=f"upd_order_notes{field_key_suffix}")
            form_col1, form_col2 = st.columns([3, 1])
            update_order_btn = form_col1.form_submit_button("💾 Update Order", type="primary")
            delete_order_btn = form_col2.form_submit_button("🗑️ Delete", type="secondary")

        if update_order_btn:
            final_project_name = _safe_str(upd_project_match).strip() or _safe_str(upd_project_name).strip()
            if not final_project_name:
                st.error("Project name is required.")
            else:
                try:
                    _update_order(
                        _safe_int(selected_order.get("id"), default=0),
                        orders_source_name,
                        order_number=_safe_str(upd_order_number).strip(),
                        project_name=final_project_name,
                        country=_safe_str(upd_country).strip(),
                        ordered_cameras=int(upd_ordered_cameras),
                        payment_amount=float(upd_payment_amount),
                        payment_month=_safe_str(upd_payment_month).strip(),
                        installation_year=_safe_int(upd_installation_year) or None,
                        order_date=upd_order_date,
                        requested_activation_date=upd_requested_activation if upd_has_activation else None,
                        status=upd_status,
                        notes=_safe_str(upd_notes),
                    )
                    load_orders_data.clear()
                    st.session_state["_flash_success"] = "Order updated successfully."
                    st.session_state["_flash_success_page"] = "📦 Orders"
                    st.rerun()
                except Exception as exc:
                    st.error(f"Update failed: {exc}")

        if delete_order_btn:
            try:
                _delete_order(_safe_int(selected_order.get("id"), default=0), orders_source_name)
                load_orders_data.clear()
                st.session_state["_flash_success"] = "Order deleted successfully."
                st.session_state["_flash_success_page"] = "📦 Orders"
                st.rerun()
            except Exception as exc:
                st.error(f"Delete failed: {exc}")

        order_pdf = _get_order_pdf_bytes(selected_order)
        if order_pdf is not None:
            file_bytes, file_name = order_pdf
            st.download_button(
                "Download Source Order File",
                data=file_bytes,
                file_name=file_name,
                mime="application/octet-stream",
                key="download_order_source",
            )
        elif _safe_str(selected_order.get("source_filename")).strip():
            st.caption(f"Source file: {_safe_str(selected_order.get('source_filename')).strip()}")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: LICENSES
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🔐 Licenses":
    st.title("🔐 Licenses")
    if page_flash_success:
        st.success(page_flash_success)

    today = datetime.date.today()
    next_month = today.month % 12 + 1
    next_month_year = today.year if today.month < 12 else today.year + 1

    license_rows = []
    for project in projects:
        license_date = _project_license_date(project)
        license_rows.append({
            "Project": _safe_str(project.project_name),
            "Country": _safe_str(project.country),
            "Cameras": _safe_int(project.num_cams),
            "Status": _safe_str(project.status),
            "License EOP": license_date.strftime("%Y-%m-%d") if license_date else "",
            "License Status": _license_status(project, today),
        })

    next_month_rows = [
        row for row in license_rows
        if row["License Status"] == "Update Next Month"
    ]
    active_license_count = sum(1 for row in license_rows if row["License Status"] == "Active")
    missing_license_count = sum(1 for row in license_rows if row["License Status"] == "Missing")
    expired_license_count = sum(1 for row in license_rows if row["License Status"] == "Expired")
    update_next_month_count = sum(1 for row in license_rows if row["License Status"] == "Update Next Month")

    lc1, lc2, lc3, lc4 = st.columns(4)
    lc1.metric("Active licenses", active_license_count)
    lc2.metric("Need update next month", update_next_month_count)
    lc3.metric("Expired / missing", expired_license_count + missing_license_count)
    lc4.metric("Total tracked projects", len(license_rows))

    with st.expander("Filters", expanded=True):
        lf1, lf2, lf3 = st.columns(3)
        license_country = lf1.selectbox(
            "Country",
            ["All"] + sorted({row["Country"] for row in license_rows if row["Country"]}),
            key="license_country",
        )
        license_status = lf2.selectbox(
            "License Status",
            ["All", "Active", "Update Next Month", "Expired", "Missing", "Offline"],
            key="license_status",
        )
        license_search = lf3.text_input("Search project", key="license_search")

    filtered_license_rows = [
        row for row in license_rows
        if (license_country == "All" or row["Country"] == license_country)
        and (license_status == "All" or row["License Status"] == license_status)
        and (not license_search.strip() or license_search.lower() in row["Project"].lower())
    ]

    st.subheader(f"Projects Needing Update in {calendar.month_name[next_month]} {next_month_year} (offline excluded)")
    if next_month_rows:
        st.dataframe(
            pd.DataFrame(next_month_rows)[["Project", "Country", "Cameras", "License EOP", "Status"]],
            use_container_width=True,
            hide_index=True,
            height=220,
        )
    else:
        st.info(f"No non-offline projects currently expire in {calendar.month_name[next_month]} {next_month_year}.")

    if CAN_EDIT and license_rows:
        st.markdown("---")
        st.subheader("Add or Extend License EOP")
        project_names = sorted(row["Project"] for row in license_rows)
        with st.form("license_update_form"):
            lu1, lu2 = st.columns(2)
            selected_project_name = lu1.selectbox("Project", project_names, key="license_project")
            selected_project = next((project for project in projects if project.project_name == selected_project_name), None)
            current_license_date = _project_license_date(selected_project) if selected_project else None
            extend_action = lu2.selectbox(
                "Action",
                ["Set exact date", "Extend by 1 month", "Extend by 12 months"],
                key="license_action",
            )
            base_license_date = current_license_date or today
            new_license_date = st.date_input(
                "New License EOP",
                value=_add_months(base_license_date, 12),
                key="license_new_date",
            )
            if current_license_date:
                st.caption(f"Current License EOP: {current_license_date.strftime('%Y-%m-%d')}")
            else:
                st.caption("Current License EOP: not set")
            submit_license = st.form_submit_button("Save License Update")

        if submit_license and selected_project is not None:
            previous_license_date = current_license_date
            if extend_action == "Set exact date":
                target_license_date = new_license_date
            elif extend_action == "Extend by 1 month":
                target_license_date = _add_months(max(base_license_date, today), 1)
            else:
                target_license_date = _add_months(max(base_license_date, today), 12)

            selected_project.license_eop = datetime.datetime.combine(target_license_date, datetime.time.min)
            try:
                _save_projects(projects, _data_path)
                append_license_change_log({
                    "project_name": selected_project.project_name,
                    "country": selected_project.country,
                    "old_license_eop": previous_license_date.isoformat() if previous_license_date else None,
                    "new_license_eop": target_license_date.isoformat(),
                    "action": extend_action,
                    "source_name": _data_path,
                })
                load_data.clear()
                # Reset License Status filter to "All" so the updated project is visible
                # (it may move from "Expired" to "Active" and would otherwise be hidden).
                st.session_state["license_status"] = "All"
                st.session_state["_flash_success"] = (
                    f"License EOP updated for {selected_project.project_name}: {target_license_date.strftime('%Y-%m-%d')}"
                )
                st.session_state["_flash_success_page"] = "🔐 Licenses"
                st.rerun()
            except Exception as exc:
                st.error(f"Failed to save license update: {exc}")

    st.markdown("---")
    st.subheader("All Project Licenses")
    license_table_columns = [
        "Project",
        "Country",
        "Cameras",
        "Project Status",
        "License EOP",
        "License Status",
    ]
    license_table_df = pd.DataFrame([
        {
            "Project": row.get("Project", ""),
            "Country": row.get("Country", ""),
            "Cameras": row.get("Cameras", 0),
            "Project Status": row.get("Status", ""),
            "License EOP": row.get("License EOP", ""),
            "License Status": row.get("License Status", ""),
        }
        for row in filtered_license_rows
    ], columns=license_table_columns)

    def color_license_status(value):
        key = str(value).strip().lower()
        if key == "active":
            return "color: #27AE60; font-weight: bold"
        if key == "update next month":
            return "color: #F39C12; font-weight: bold"
        if key == "expired":
            return "color: #E74C3C; font-weight: bold"
        if key == "missing":
            return "color: #7F8C8D; font-weight: bold"
        if key == "offline":
            return "color: #5D6D7E; font-weight: bold"
        return ""

    if "License Status" in license_table_df.columns:
        license_table_display = license_table_df.style.map(
            color_license_status,
            subset=["License Status"],
        )
    else:
        license_table_display = license_table_df

    st.dataframe(
        license_table_display,
        use_container_width=True,
        hide_index=True,
        height=480,
    )

    st.markdown("---")
    st.subheader("License Change Log")
    license_change_rows = load_license_change_log()
    if license_change_rows:
        history_rows = []
        for row in license_change_rows:
            changed_at = _parse_optional_datetime(row.get("changed_at"))
            history_rows.append({
                "Changed At": changed_at.strftime("%Y-%m-%d %H:%M") if changed_at else _safe_str(row.get("changed_at")),
                "Month": changed_at.strftime("%B %Y") if changed_at else "",
                "Project": _safe_str(row.get("project_name")),
                "Country": _safe_str(row.get("country")),
                "Old License EOP": _safe_str(row.get("old_license_eop")),
                "New License EOP": _safe_str(row.get("new_license_eop")),
                "Action": _safe_str(row.get("action")),
                "Source": _safe_str(row.get("source_name")),
            })

        history_df = pd.DataFrame(history_rows)
        month_choices = ["All"] + sorted({row["Month"] for row in history_rows if row["Month"]}, reverse=True)
        selected_month = st.selectbox("History month", month_choices, key="license_history_month")
        preset_options = ["All", "Previous month only"]
        selected_preset = st.selectbox("History preset", preset_options, key="license_history_preset")

        if selected_preset == "Previous month only":
            previous_month_date = today.replace(day=1) - datetime.timedelta(days=1)
            previous_month_label = previous_month_date.strftime("%B %Y")
            history_df = history_df[history_df["Month"] == previous_month_label]
        if selected_month != "All":
            history_df = history_df[history_df["Month"] == selected_month]

        visible_history_df = history_df[["Changed At", "Project", "Country", "Old License EOP", "New License EOP", "Action", "Source"]]
        st.download_button(
            "Download CSV",
            data=visible_history_df.to_csv(index=False).encode("utf-8"),
            file_name="license_change_log.csv",
            mime="text/csv",
        )
        st.dataframe(
            visible_history_df,
            use_container_width=True,
            hide_index=True,
            height=320,
        )
    else:
        st.info("No license changes have been logged yet.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: INVOICE DETAILS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🧾 Invoice Details":
    st.title("🧾 Invoice Details")
    if page_flash_success:
        st.success(page_flash_success)

    with st.expander("🔍 Filters", expanded=False):
        col1, col2, col3, col4 = st.columns(4)
        years = sorted({inv.year for inv in invoices if inv.year}, reverse=True)
        sel_year    = col1.selectbox("Year",    ["All"] + [str(y) for y in years], key="inv_year")
        sel_paid    = col2.selectbox("Paid Status", ["All", "Paid", "Unpaid", "Cancelled"], key="inv_paid")
        countries   = sorted({p.country for p in projects if p.country})
        sel_country = col3.selectbox("Country", ["All"] + countries, key="inv_country")
        search      = col4.text_input("Search project name", key="inv_search")

        col5, col6, col7, col8 = st.columns(4)
        maint_years = sorted({inv.maintenance_year for inv in invoices if inv.maintenance_year})
        sel_maint     = col5.selectbox("Maint. Year", ["All"] + maint_years, key="inv_maint")
        inv_no_search = col6.text_input("Invoice #", key="inv_no_search")
        amt_min       = col7.number_input("Min Amount (€)", min_value=0, value=0, step=100, key="inv_amt_min")
        amt_max       = col8.number_input("Max Amount (€)", min_value=0, value=0, step=100, key="inv_amt_max",
                                          help="0 = no limit")

    proj_country = {p.project_name.lower(): p.country for p in projects}

    filtered_inv = invoices
    if sel_year != "All":
        filtered_inv = [i for i in filtered_inv if i.year == int(sel_year)]
    if sel_paid == "Paid":
        filtered_inv = [i for i in filtered_inv if i.is_paid()]
    elif sel_paid == "Unpaid":
        filtered_inv = [i for i in filtered_inv if i.is_unpaid()]
    elif sel_paid == "Cancelled":
        filtered_inv = [i for i in filtered_inv if i.is_cancelled()]
    if sel_country != "All":
        filtered_inv = [i for i in filtered_inv
                        if proj_country.get(i.project_name.lower().strip()) == sel_country]
    if search:
        filtered_inv = [i for i in filtered_inv if search.lower() in i.project_name.lower()]
    if sel_maint != "All":
        filtered_inv = [i for i in filtered_inv if _safe_str(i.maintenance_year) == sel_maint]
    if inv_no_search.strip():
        filtered_inv = [i for i in filtered_inv if inv_no_search.strip() in _safe_str(i.invoice_number)]
    if amt_min > 0:
        filtered_inv = [i for i in filtered_inv if i.payment_amount >= amt_min]
    if amt_max > 0:
        filtered_inv = [i for i in filtered_inv if i.payment_amount <= amt_max]

    deduped_filtered_inv = []
    seen_invoice_project_keys = set()
    for inv in filtered_inv:
        invoice_number_key = _safe_int(inv.invoice_number) or 0
        project_key = _normalize_project_name_key(canonical_project_name(_safe_str(inv.project_name).strip()))
        dedupe_key = (invoice_number_key, project_key)
        if project_key and dedupe_key in seen_invoice_project_keys:
            continue
        seen_invoice_project_keys.add(dedupe_key)
        deduped_filtered_inv.append(inv)
    if len(deduped_filtered_inv) != len(filtered_inv):
        filtered_inv = deduped_filtered_inv

    # Summary row
    total_all_f    = sum(i.payment_amount for i in filtered_inv)
    total_paid_f   = sum(i.payment_amount for i in filtered_inv if i.is_paid())
    total_unpaid_f = sum(i.payment_amount for i in filtered_inv if i.is_unpaid())
    total_cancel_f = sum(i.payment_amount for i in filtered_inv if i.is_cancelled())
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Invoices shown",    len(filtered_inv))
    c2.metric("Total Amount",      f"€{total_all_f:,.0f}")
    c3.metric("Total Paid",        f"€{total_paid_f:,.0f}")
    c4.metric("Total Unpaid",      f"€{total_unpaid_f:,.0f}")
    c5.metric("Total Cancelled",   f"€{total_cancel_f:,.0f}")

    st.caption(f"Showing {len(filtered_inv)} of {len(invoices)} invoices — use filters above to narrow results")

    saved_invoice_numbers = sorted(
        {
            int(float(inv.invoice_number))
            for inv in invoices
            if inv.invoice_number not in (None, "")
        },
        reverse=True,
    )

    with st.expander("🧾 Download Invoice PDF From Ledger", expanded=False):
        st.caption("Builds a PDF from the saved invoice rows only. This does not change any data in Supabase.")
        if not saved_invoice_numbers:
            st.info("No saved invoice numbers were found.")
        else:
            default_invoice_number = saved_invoice_numbers[0]
            if inv_no_search.strip().isdigit():
                searched_invoice_number = int(inv_no_search.strip())
                if searched_invoice_number in saved_invoice_numbers:
                    default_invoice_number = searched_invoice_number
            selected_invoice_number = st.selectbox(
                "Saved invoice number",
                saved_invoice_numbers,
                index=saved_invoice_numbers.index(default_invoice_number),
                key="invoice_pdf_from_ledger_number",
            )
            pdf_description = st.text_input(
                "PDF description",
                value="Iretailcheck - Maintenance Adjustment",
                key="invoice_pdf_from_ledger_description",
                help="Only affects the PDF text. It does not change the ledger rows.",
            )
            invoice_rows_for_pdf = [
                inv for inv in invoices
                if _safe_int(inv.invoice_number) == int(selected_invoice_number)
            ]
            if invoice_rows_for_pdf:
                try:
                    from services.pdf_service import generate_invoice_pdf_from_rows

                    invoice_pdf_bytes = generate_invoice_pdf_from_rows(
                        invoice_rows=invoice_rows_for_pdf,
                        invoice_number=int(selected_invoice_number),
                        description=_safe_str(pdf_description).strip() or None,
                    )
                    st.download_button(
                        label="Download PDF Invoice",
                        data=invoice_pdf_bytes,
                        file_name=f"CC_inv_{int(selected_invoice_number)}.pdf",
                        mime="application/pdf",
                        key="download_invoice_pdf_from_ledger",
                    )
                except Exception as exc:
                    st.error(f"PDF generation failed: {exc}")
            else:
                st.warning("No saved ledger rows were found for the selected invoice number.")

    _invoice_columns = [
        "Invoice #",
        "Project",
        "Maint. Year",
        "Amount (€)",
        "Cameras",
        "Payment Date",
        "Paid",
        "Year",
        "Description",
    ]

    invoice_index_map = {id(inv): idx for idx, inv in enumerate(invoices)}

    df_inv = pd.DataFrame(
        [
            {
                "_invoice_index": invoice_index_map[id(i)],
                "Invoice #": _safe_int(i.invoice_number) or None,
                "Project": canonical_project_name(_safe_str(i.project_name)),
                "Maint. Year": _safe_str(i.maintenance_year),
                "Amount (€)": _safe_float(i.payment_amount),
                "Cameras": _safe_int(i.cameras_number),
                "Payment Date": i.payment_date.date() if i.payment_date else None,
                "Paid": _safe_str(i.paid),
                "Year": _safe_str(_safe_int(i.year) or ""),
                "Description": _safe_str(getattr(i, "description", "")) if getattr(i, "description", "") else "",
            }
            for i in filtered_inv
        ],
        columns=["_invoice_index", *_invoice_columns],
    )

    # Apply dynamic sort (sort controls defined above in CAN_EDIT block, but sort applies to df_inv always)
    _inv_sort_col = st.session_state.get("inv_sort_column", "Year")
    _inv_sort_asc = st.session_state.get("inv_sort_order", "Descending") == "Ascending"
    if _inv_sort_col in df_inv.columns:
        df_inv = df_inv.sort_values(
            by=_inv_sort_col,
            ascending=_inv_sort_asc,
            na_position="last",
            kind="mergesort",
        ).reset_index(drop=True)

    def color_paid(val):
        v = str(val).strip().lower()
        if v == "yes":     return "color: #27AE60; font-weight: bold"
        if v == "no":      return "color: #E74C3C"
        if v == "cancelled": return "color: #F39C12"
        return ""

    def _render_invoice_bottom_summary(table_df: pd.DataFrame):
        if table_df is None or table_df.empty:
            return

        total_amount = sum(_safe_float(value) for value in table_df.get("Amount (€)", []))
        total_cameras = sum(_safe_int(value) for value in table_df.get("Cameras", []))
        total_projects = len(
            {
                canonical_project_name(_safe_str(value).strip())
                for value in table_df.get("Project", [])
                if _safe_str(value).strip()
            }
        )
        invoice_numbers = sorted(
            {
                _safe_int(value)
                for value in table_df.get("Invoice #", [])
                if _safe_int(value)
            }
        )
        summary_invoice_number = invoice_numbers[0] if len(invoice_numbers) == 1 else ""

        summary_row_df = pd.DataFrame([
            {
                "Invoice #": str(summary_invoice_number) if summary_invoice_number else "",
                "Project": f"TOTAL ({total_projects} projects)",
                "Maint. Year": "",
                "Amount (€)": f"{total_amount:,.0f}",
                "Cameras": str(total_cameras),
                "Payment Date": "",
                "Paid": "",
                "Year": "",
                "Description": "",
            }
        ])
        st.dataframe(
            summary_row_df,
            use_container_width=True,
            hide_index=True,
            height=70,
        )
        if len(deduped_filtered_inv) != len(table_df):
            st.caption(f"{len(table_df) - len(deduped_filtered_inv)} duplicate project row(s) were collapsed in this view.")

    if CAN_EDIT:
        st.info("✏️ Admin mode: you can edit cells directly. Click **Save Changes** when done.")

        def _parse_invoice_date(value):
            if value in (None, ""):
                return None
            if isinstance(value, datetime.datetime):
                return value
            if isinstance(value, datetime.date):
                return datetime.datetime.combine(value, datetime.time.min)
            try:
                return datetime.datetime.strptime(str(value), "%Y-%m-%d")
            except Exception:
                return None

        def _invoice_row_has_values(row) -> bool:
            return any([
                row.get("Invoice #") not in (None, ""),
                _safe_str(row.get("Maint. Year", "")).strip(),
                _safe_float(row.get("Amount (€)", 0)) != 0,
                _safe_int(row.get("Cameras", 0)) != 0,
                row.get("Payment Date") not in (None, ""),
                _safe_str(row.get("Paid", "")).strip() not in ("", "No"),
                _safe_str(row.get("Year", "")).strip(),
            ])

        blank_project_invoices = [inv for inv in invoices if not _safe_str(inv.project_name).strip()]
        if blank_project_invoices:
            st.warning(
                f"Found {len(blank_project_invoices)} invoice row(s) with an empty project name. "
                "Remove them before saving other invoice changes."
            )
            if st.button("🧹 Remove Blank Project Invoice Rows", key="remove_blank_project_invoices"):
                invoices[:] = [inv for inv in invoices if _safe_str(inv.project_name).strip()]
                try:
                    _save_invoices(invoices, _data_path)
                    load_data.clear()
                    st.session_state["_flash_success"] = "Blank-project invoice rows were removed."
                    st.session_state["_flash_success_page"] = "🧾 Invoice Details"
                    st.rerun()
                except Exception as exc:
                    st.error(f"Failed to remove blank-project invoice rows: {exc}")

        with st.expander("📥 Import Invoice XLSX", expanded=False):
            if _is_excel_source(_data_path):
                st.warning("XLSX import is only available when the app is connected to Supabase.")
            else:
                uploaded_invoice_xlsx = st.file_uploader(
                    "Upload invoice XLSX",
                    type=["xlsx"],
                    key="invoice_xlsx_upload",
                    help="Upload a generated invoice workbook and replace that invoice number in Supabase.",
                )
                if uploaded_invoice_xlsx is not None:
                    try:
                        import_meta, import_rows = _parse_uploaded_invoice_xlsx(uploaded_invoice_xlsx.getvalue())
                        inferred_invoice_number = import_meta["invoice_number"] or 0
                        target_invoice_number = st.number_input(
                            "Invoice number to replace",
                            min_value=1,
                            step=1,
                            value=max(1, inferred_invoice_number),
                            key="invoice_import_target_number",
                        )
                        st.caption(
                            f"Parsed {import_meta['row_count']} rows from '{import_meta['title'] or uploaded_invoice_xlsx.name}' "
                            f"for year {import_meta['year'] or 'unknown'} with total €{import_meta['total_amount']:,.0f}."
                        )
                        preview_df = pd.DataFrame([
                            {
                                "Project": row["project_name"],
                                "Maint. Year": row["maintenance_year"],
                                "Amount (€)": row["payment_amount"],
                                "Cameras": row["cameras_number"],
                                "Year": row["year"],
                            }
                            for row in import_rows
                        ])
                        st.dataframe(preview_df, use_container_width=True, height=250)

                        if preview_df.empty:
                            st.warning("No invoice rows were found in the uploaded XLSX.")
                        else:
                            blank_import_rows = [row for row in import_rows if not _safe_str(row.get("project_name", "")).strip()]
                            project_counts = preview_df["Project"].value_counts()
                            duplicate_projects = project_counts[project_counts > 1]
                            if blank_import_rows:
                                st.error("The uploaded file contains invoice rows with an empty project name. Import was blocked.")
                            elif not duplicate_projects.empty:
                                st.error(
                                    "The uploaded file contains duplicate project rows: "
                                    + ", ".join(duplicate_projects.index.tolist())
                                )
                            elif st.button("Replace Invoice From XLSX", key="replace_invoice_from_xlsx"):
                                replace_invoice_rows(target_invoice_number, import_rows)
                                load_data.clear()
                                st.session_state["_flash_success"] = (
                                    f"Invoice {target_invoice_number} replaced from uploaded XLSX with {len(import_rows)} row(s)."
                                )
                                st.session_state["_flash_success_page"] = "🧾 Invoice Details"
                                st.rerun()
                    except Exception as exc:
                        st.error(f"Failed to parse uploaded invoice XLSX: {exc}")

        control_col1, control_col2, _control_spacer = st.columns([1, 1, 3])
        with control_col1:
            if st.button("➕ Add New Invoice", key="btn_add_inv"):
                st.session_state["add_inv_row"] = st.session_state.get("add_inv_row", 0) + 1
        with control_col2:
            save_invoice_clicked = st.button("💾 Save Changes", key="save_invoices_top")

        inv_sort_col1, inv_sort_col2 = st.columns(2)
        inv_sort_options = ["Invoice #", "Year", "Project", "Amount (€)", "Cameras", "Maint. Year", "Payment Date", "Paid"]
        inv_selected_sort = inv_sort_col1.selectbox("Sort by", inv_sort_options, index=1, key="inv_sort_column")
        inv_selected_order = inv_sort_col2.selectbox("Sort order", ["Descending", "Ascending"], index=0, key="inv_sort_order")

        invoice_project_options = [""] + sorted(
            {
                canonical_project_name(_safe_str(p.project_name).strip())
                for p in projects
                if canonical_project_name(_safe_str(p.project_name).strip())
            }
            | {
                canonical_project_name(_safe_str(inv.project_name).strip())
                for inv in invoices
                if canonical_project_name(_safe_str(inv.project_name).strip())
            }
        )
        invoice_paid_options = ["No", "Yes", "cancelled"]
        invoice_year_options = [""] + [str(year) for year in range(datetime.date.today().year + 1, 2012, -1)]
        invoice_maint_options = []
        for label in maint_years + [f"Y{i}" for i in range(1, 11)] + ["Paid Trial-0.5Y"]:
            if label and label not in invoice_maint_options:
                invoice_maint_options.append(label)

        next_invoice_number = _get_next_invoice_number(invoices, _data_path)

        _empty_inv = {"_invoice_index": None, "Invoice #": None, "Project": "", "Maint. Year": "Y1",
                      "Amount (€)": 0.0, "Cameras": 0,
                      "Payment Date": None, "Paid": "No", "Year": str(datetime.date.today().year),
                      "Description": "Complementary"}
        n_new_inv = st.session_state.get("add_inv_row", 0)
        if n_new_inv:
            empty_rows = pd.DataFrame([
                {
                    **_empty_inv,
                    "Invoice #": next_invoice_number + row_offset,
                }
                for row_offset in range(n_new_inv)
            ])
            df_inv_edit = pd.concat([empty_rows, df_inv.reset_index(drop=True)], ignore_index=True)
        else:
            df_inv_edit = df_inv.reset_index(drop=True)

        edited_inv = st.data_editor(
            df_inv_edit,
            use_container_width=True,
            height=550,
            num_rows="dynamic",
            column_config={
                "_invoice_index": None,
                "Invoice #": st.column_config.NumberColumn(
                    "Invoice #",
                    min_value=0,
                    step=1,
                    format="%d",
                ),
                "Project": st.column_config.SelectboxColumn(
                    "Project",
                    options=invoice_project_options,
                ),
                "Maint. Year": st.column_config.SelectboxColumn(
                    "Maint. Year",
                    options=invoice_maint_options,
                ),
                "Payment Date": st.column_config.DateColumn(
                    "Payment Date",
                    format="YYYY-MM-DD",
                ),
                "Paid": st.column_config.SelectboxColumn(
                    "Paid",
                    options=invoice_paid_options,
                ),
                "Year": st.column_config.SelectboxColumn(
                    "Year",
                    options=invoice_year_options,
                ),
                "Description": st.column_config.TextColumn(
                    "Description",
                ),
            },
            key="inv_editor",
        )
        _render_invoice_bottom_summary(edited_inv)
        if save_invoice_clicked:
            invalid_editor_rows = []
            for row_idx, row in edited_inv.iterrows():
                project = _safe_str(row.get("Project", "")).strip()
                if not project and _invoice_row_has_values(row):
                    invalid_editor_rows.append(row_idx + 1)

            if invalid_editor_rows:
                st.error(
                    "Cannot save invoice rows with invoice data but no project name. "
                    f"Fix or remove row(s): {', '.join(str(idx) for idx in invalid_editor_rows[:10])}"
                )
                st.stop()

            from models.invoice import Invoice as InvoiceModel
            inv_map = {(i.invoice_number, i.project_name.strip().lower()): i for i in invoices if i.invoice_number}
            # Secondary lookup for invoices without invoice numbers — match by (project, maint_year, year)
            no_inv_map = {
                (i.project_name.strip().lower(), str(i.maintenance_year).strip(), str(i.year) if i.year else ""): i
                for i in invoices if not i.invoice_number
            }
            new_count = 0
            for _, row in edited_inv.iterrows():
                project = _safe_str(row.get("Project", "")).strip()
                if not project:
                    continue
                raw_invoice_index = row.get("_invoice_index", "")
                existing_invoice_index = _safe_int(raw_invoice_index, default=-1)
                inv = None
                if existing_invoice_index >= 0 and existing_invoice_index < len(invoices):
                    inv = invoices[existing_invoice_index]
                inv_no_str = _safe_str(row.get("Invoice #", "")).strip()
                try:
                    inv_no = float(inv_no_str) if inv_no_str else None
                except Exception:
                    inv_no = None
                maint_year = _safe_str(row.get("Maint. Year", "")).strip()
                inv_year = str(_safe_int(row.get("Year")) or "")
                if inv is None and inv_no:
                    inv = inv_map.get((inv_no, project.lower()))
                elif inv is None:
                    inv = no_inv_map.get((project.lower(), maint_year, inv_year))
                if inv is None:
                    # Truly new invoice row
                    inv = InvoiceModel(
                        invoice_number=inv_no,
                        project_name=project,
                        maintenance_year=maint_year,
                        payment_amount=_safe_float(row.get("Amount (€)", 0)),
                        cameras_number=_safe_int(row.get("Cameras", 0)) or None,
                        payment_date=None,
                        paid=_safe_str(row.get("Paid", "No")),
                        year=_safe_int(row.get("Year")) or None,
                        description=_safe_str(row.get("Description", "")).strip() or None,
                    )
                    invoices.append(inv)
                    if inv_no:
                        inv_map[(inv_no, project.lower())] = inv
                    else:
                        no_inv_map[(project.lower(), maint_year, inv_year)] = inv
                    new_count += 1
                else:
                    inv.invoice_number = inv_no
                    inv.project_name   = project
                    inv.maintenance_year = _safe_str(row.get("Maint. Year", ""))
                    inv.paid           = _safe_str(row.get("Paid", ""))
                    inv.payment_amount = _safe_float(row.get("Amount (€)", 0))
                    inv.cameras_number = _safe_int(row.get("Cameras", 0)) or None
                    inv.year           = _safe_int(row.get("Year")) or None
                inv.description = _safe_str(row.get("Description", "")).strip() or None
                inv.payment_date = _parse_invoice_date(row.get("Payment Date"))
            try:
                _save_invoices(invoices, _data_path)
                load_data.clear()
                st.session_state.pop("add_inv_row", None)
                st.session_state.pop("inv_editor", None)
                msg = f"Saved! {new_count} new invoice(s) added." if new_count else "Invoices saved successfully!"
                st.session_state["_flash_success"] = msg
                st.session_state["_flash_success_page"] = "🧾 Invoice Details"
                st.rerun()
            except Exception as e:
                st.error(f"Save failed: {e}")
    else:
        df_inv_display = (
            df_inv.style
            .map(color_paid, subset=["Paid"])
            .set_properties(subset=["Invoice #", "Amount (€)", "Cameras", "Year"], **{"text-align": "left"})
        )
        st.dataframe(
            df_inv_display,
            use_container_width=True,
            height=550,
        )
        _render_invoice_bottom_summary(df_inv)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: DEBT REPORT
# ══════════════════════════════════════════════════════════════════════════════
elif page == "💸 Debt Report":
    st.title("💸 Debt Report")

    debt_type_options = ["All", "New Installation (Y1)", "Paid Trials", "Maintenance (Y2+)"]
    pending_debt_type = st.session_state.pop("_pending_dr_debt_type", None)
    if pending_debt_type in debt_type_options:
        st.session_state["dr_debt_type"] = pending_debt_type
        st.session_state["dr_debt_type_widget"] = pending_debt_type
    active_debt_type = st.session_state.get("dr_debt_type", "All")
    if active_debt_type not in debt_type_options:
        active_debt_type = "All"
        st.session_state["dr_debt_type"] = active_debt_type
    widget_debt_type = st.session_state.get("dr_debt_type_widget", active_debt_type)
    if widget_debt_type not in debt_type_options:
        widget_debt_type = active_debt_type
        st.session_state["dr_debt_type_widget"] = widget_debt_type

    # ── Filters ───────────────────────────────────────────────────────────────
    with st.expander("🔍 Filters", expanded=False):
        fc1, fc2, fc3, fc4 = st.columns(4)
        debt_years    = sorted({inv.year for inv in invoices if inv.year}, reverse=True)
        dsel_year     = fc1.selectbox("Year", ["All"] + [str(y) for y in debt_years], key="dr_year")
        debt_countries = sorted({
            _safe_str(country).strip()
            for country in ([p.country for p in projects] + [ds.country for ds in debt_summaries])
            if _safe_str(country).strip()
        })
        dsel_country  = fc2.selectbox("Country", ["All"] + debt_countries, key="dr_country")
        dsel_search   = fc3.text_input("Search project", key="dr_search")
        dsel_debt_type = fc4.selectbox(
            "Debt Type",
            debt_type_options,
            index=debt_type_options.index(widget_debt_type),
            key="dr_debt_type_widget",
        )
        st.session_state["dr_debt_type"] = dsel_debt_type

    # Only unpaid invoices
    debt_inv = [i for i in invoices if i.is_unpaid()]
    if dsel_year != "All":
        debt_inv = [i for i in debt_inv if i.year == int(dsel_year)]

    # Country lookup — normalize names aggressively, fallback to debt_summaries
    import re as _re
    import unicodedata as _ud
    def _norm(s):
        s = str(s or "").strip()
        s = _re.sub(r'\s*\([^)]*\)', '', s).strip()          # strip "(coplementary)" etc.
        s = _ud.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')  # strip accents
        return _re.sub(r'\s+', ' ', s).strip().lower()
    proj_country_map = {}
    for p in projects:
        key = _norm(canonical_project_name(p.project_name))
        if not key:
            continue
        country = _normalize_country(_safe_str(p.country).strip())
        # Keep first non-empty country when duplicate normalized project names exist.
        if country and not _safe_str(proj_country_map.get(key, "")).strip():
            proj_country_map[key] = country
        elif key not in proj_country_map:
            proj_country_map[key] = country

    ds_country_map = {}
    for ds in debt_summaries:
        key = _norm(canonical_project_name(ds.project_name))
        if not key:
            continue
        country = _normalize_country(_safe_str(ds.country).strip())
        if country and not _safe_str(ds_country_map.get(key, "")).strip():
            ds_country_map[key] = country
        elif key not in ds_country_map:
            ds_country_map[key] = country

    orders_country_map = {}
    try:
        orders_rows, _orders_source_name = load_orders_data(_data_path)
    except Exception:
        orders_rows = []
    for order in orders_rows:
        key = _norm(canonical_project_name(_safe_str(order.get("project_name")).strip()))
        if not key:
            continue
        country = _normalize_country(_safe_str(order.get("country")).strip())
        if country and not _safe_str(orders_country_map.get(key, "")).strip():
            orders_country_map[key] = country
        elif key not in orders_country_map:
            orders_country_map[key] = country

    # Deterministic overrides for known project-name variants that may not match
    # project/order master data exactly.
    manual_country_overrides = {
        _norm("intermarche heusy"): "Belgium",
        _norm("intermarché heusy"): "Belgium",
    }

    merged_country_map = {}
    for source_map in (proj_country_map, ds_country_map, orders_country_map):
        for key, country in source_map.items():
            existing = _safe_str(merged_country_map.get(key, "")).strip()
            candidate = _safe_str(country).strip()
            if candidate and not existing:
                merged_country_map[key] = candidate
            elif key not in merged_country_map:
                merged_country_map[key] = candidate

    def _infer_country_from_project_name(name: str) -> str:
        key = _norm(canonical_project_name(name))
        if not key:
            return ""
        if "luxemb" in key:
            return "Luxembourg"
        if any(token in key for token in ("edeka", "rewe", "germany", "deutschland")):
            return "Germany"
        if any(token in key for token in ("belg", "belgium")):
            return "Belgium"
        return ""

    def _get_country(name):
        k = _norm(canonical_project_name(name))
        if not k:
            return ""
        manual_country = _safe_str(manual_country_overrides.get(k, "")).strip()
        if manual_country:
            return manual_country
        result = proj_country_map.get(k) or ds_country_map.get(k) or orders_country_map.get(k)
        if result:
            return result

        # Relaxed partial match handles suffixes/prefixes/noise in imported names.
        for source_map in (proj_country_map, ds_country_map, orders_country_map):
            for proj_k, country in source_map.items():
                if not proj_k or not _safe_str(country).strip():
                    continue
                if (
                    k.startswith(proj_k)
                    or proj_k.startswith(k)
                    or (k in proj_k)
                    or (proj_k in k)
                ):
                    return country

        # Token overlap fallback for long names with small textual differences.
        k_tokens = {token for token in k.split(" ") if token}
        best_country = ""
        best_overlap = 0
        if k_tokens:
            for source_map in (proj_country_map, ds_country_map, orders_country_map):
                for proj_k, country in source_map.items():
                    if not _safe_str(country).strip():
                        continue
                    proj_tokens = {token for token in proj_k.split(" ") if token}
                    overlap = len(k_tokens.intersection(proj_tokens))
                    if overlap > best_overlap and overlap >= 2:
                        best_overlap = overlap
                        best_country = country
        if best_country:
            return best_country

        # Final fuzzy fallback for renamed/edited invoice project labels.
        from difflib import SequenceMatcher

        best_ratio = 0.0
        best_fuzzy_country = ""
        compact_k = k.replace(" ", "")
        for proj_k, country in merged_country_map.items():
            country_text = _safe_str(country).strip()
            if not proj_k or not country_text:
                continue
            ratio = SequenceMatcher(None, k, proj_k).ratio()
            compact_ratio = SequenceMatcher(None, compact_k, proj_k.replace(" ", "")).ratio()
            candidate_ratio = max(ratio, compact_ratio)
            if candidate_ratio > best_ratio:
                best_ratio = candidate_ratio
                best_fuzzy_country = country_text
        if best_ratio >= 0.86:
            return best_fuzzy_country
        return _infer_country_from_project_name(name)

    if dsel_country != "All":
        debt_inv = [i for i in debt_inv if _get_country(i.project_name) == dsel_country]
    if dsel_search.strip():
        debt_inv = [i for i in debt_inv if dsel_search.lower() in i.project_name.lower()]
    if dsel_debt_type == "New Installation (Y1)":
        debt_inv = [i for i in debt_inv if _is_new_installation_category(i)]
    elif dsel_debt_type == "Paid Trials":
        debt_inv = [i for i in debt_inv if _is_paid_trial_category(i)]
    elif dsel_debt_type == "Maintenance (Y2+)":
        debt_inv = [i for i in debt_inv if _is_maintenance_category(i)]

    # Build invoice-level country hints from known rows.
    invoice_country_votes: dict[int, dict[str, int]] = {}
    for inv in debt_inv:
        inv_no = _safe_int(inv.invoice_number, default=0)
        if not inv_no:
            continue
        country = _normalize_country(_safe_str(_get_country(inv.project_name)).strip())
        if not country:
            continue
        bucket = invoice_country_votes.setdefault(inv_no, {})
        bucket[country] = bucket.get(country, 0) + 1

    invoice_country_hint: dict[int, str] = {}
    for inv_no, votes in invoice_country_votes.items():
        if votes:
            invoice_country_hint[inv_no] = max(votes.items(), key=lambda item: item[1])[0]

    # ── Summary metrics ───────────────────────────────────────────────────────
    total_debt_amt  = sum(i.payment_amount for i in debt_inv)
    proj_with_debt  = len({i.project_name for i in debt_inv})
    all_unpaid = [i for i in invoices if i.is_unpaid()]
    if dsel_year != "All":
        all_unpaid = [i for i in all_unpaid if i.year == int(dsel_year)]
    if dsel_country != "All":
        all_unpaid = [i for i in all_unpaid if _get_country(i.project_name) == dsel_country]
    if dsel_search.strip():
        all_unpaid = [i for i in all_unpaid if dsel_search.lower() in i.project_name.lower()]

    y1_total_amt = sum(i.payment_amount for i in all_unpaid if _is_new_installation_category(i))
    paid_trial_total_amt = sum(i.payment_amount for i in all_unpaid if _is_paid_trial_category(i))
    y2_total_amt = sum(i.payment_amount for i in all_unpaid if _is_maintenance_category(i))

    sent_invoice_rows = load_sent_invoices_log()
    sent_invoice_map = {}
    for row in sent_invoice_rows:
        invoice_number = _safe_int(row.get("invoice_number"), default=0)
        if not invoice_number:
            continue
        sent_at_text = _safe_str(row.get("sent_at", "")).replace("T", " ")[:19]
        current_row = sent_invoice_map.get(invoice_number)
        if current_row is None or sent_at_text >= _safe_str(current_row.get("Sent At", "")):
            sent_invoice_map[invoice_number] = {
                "Type": "Monthly Invoice",
                "For Month": f"{_safe_str(row.get('month', '')).strip()} {_safe_int(row.get('year'), default=0)}".strip(),
                "Sent At": sent_at_text,
            }

    from collections import defaultdict

    grouped_unpaid: dict[tuple, list] = defaultdict(list)
    for row_index, invoice_row in enumerate(debt_inv):
        raw_invoice_number = _safe_str(invoice_row.invoice_number).strip()
        if raw_invoice_number:
            try:
                invoice_key = ("invoice", int(float(raw_invoice_number)))
            except Exception:
                invoice_key = ("row", row_index)
        else:
            invoice_key = ("row", row_index)
        grouped_unpaid[invoice_key].append(invoice_row)

    grouped_unpaid_rows = []
    unmapped_project_names = set()
    for invoice_key, rows in grouped_unpaid.items():
        invoice_numbers = []
        project_names = []
        countries = []
        maint_years = []
        years = []
        total_amount = 0.0
        for invoice_row in rows:
            raw_invoice_number = _safe_str(invoice_row.invoice_number).strip()
            if raw_invoice_number:
                invoice_numbers.append(str(int(float(raw_invoice_number))))
            project_name = _safe_str(invoice_row.project_name).strip()
            if project_name:
                project_names.append(canonical_project_name(project_name))
            country = _normalize_country(_safe_str(_get_country(invoice_row.project_name)).strip())
            if not country:
                invoice_number_hint = _safe_int(invoice_row.invoice_number, default=0)
                if invoice_number_hint:
                    country = _safe_str(invoice_country_hint.get(invoice_number_hint, "")).strip()
            if country:
                countries.append(country)
            elif project_name:
                unmapped_project_names.add(project_name)
            maint_year = _safe_str(invoice_row.maintenance_year).strip()
            if maint_year:
                maint_years.append(maint_year)
            if invoice_row.year is not None:
                years.append(int(invoice_row.year))
            elif invoice_row.payment_date is not None:
                years.append(invoice_row.payment_date.year)
            total_amount += _safe_float(invoice_row.payment_amount)

        unique_project_names = sorted({name for name in project_names if name})
        unique_countries = sorted({country for country in countries if country})
        unique_maint_years = sorted({label for label in maint_years if label})
        grouped_unpaid_rows.append({
            "Invoice #": invoice_numbers[0] if invoice_numbers else "—",
            "Project Name": unique_project_names[0] if len(unique_project_names) <= 1 else f"{unique_project_names[0]} (+{len(unique_project_names) - 1} more)",
            "Projects": len(unique_project_names),
            "Country": ", ".join(unique_countries),
            "Maint. Year": ", ".join(unique_maint_years),
            "Amount (€)": total_amount,
            "Year": str(max(years)) if years else "",
            **sent_invoice_map.get(_safe_int(invoice_numbers[0], default=0), {"Type": "", "For Month": "", "Sent At": ""}),
            "Description": _safe_str(getattr(rows[0], "description", "")).strip() if rows else "",
        })

    grouped_unpaid_rows.sort(
        key=lambda row: (
            int(row["Invoice #"]) if str(row["Invoice #"]).isdigit() else 0,
            row["Project Name"].lower(),
        ),
        reverse=True,
    )

    mc1, mc2, mc3, mc4, mc5, mc6 = st.columns(6)
    mc1.metric("Unpaid Invoices",  len(grouped_unpaid_rows))
    mc2.metric("Projects with Debt", proj_with_debt)
    mc3.metric("Total Debt",       f"€{total_debt_amt:,.0f}")
    mc4.metric("Y1 Debt",          f"€{y1_total_amt:,.0f}")
    mc5.metric("Paid Trials",      f"€{paid_trial_total_amt:,.0f}")
    mc6.metric("Y2+ Debt",         f"€{y2_total_amt:,.0f}")

    if mc4.button("Show Y1 Invoice List", key="dr_show_y1", use_container_width=True):
        st.session_state["_pending_dr_debt_type"] = "New Installation (Y1)"
        st.rerun()
    if mc5.button("Show Paid Trials", key="dr_show_trials", use_container_width=True):
        st.session_state["_pending_dr_debt_type"] = "Paid Trials"
        st.rerun()
    if mc6.button("Show Y2+ Invoice List", key="dr_show_y2", use_container_width=True):
        st.session_state["_pending_dr_debt_type"] = "Maintenance (Y2+)"
        st.rerun()

    if dsel_debt_type != "All":
        if st.button("Clear Debt Type Filter", key="dr_clear_debt_type"):
            st.session_state["_pending_dr_debt_type"] = "All"
            st.rerun()

    if dsel_debt_type == "New Installation (Y1)":
        st.caption("Showing only first-year debt (Y1).")
    elif dsel_debt_type == "Paid Trials":
        st.caption("Showing only paid-trial debt.")
    elif dsel_debt_type == "Maintenance (Y2+)":
        st.caption("Showing only maintenance debt (Y2+).")
    else:
        st.caption("Use the Debt Type filter to switch between all debt, Y1, paid trials, and Y2+.")

    st.markdown("---")

    # ── Detailed table: one row per unpaid invoice ────────────────────────────
    st.subheader("Unpaid Invoices Detail")
    if grouped_unpaid_rows:
        detail_df = pd.DataFrame(grouped_unpaid_rows)
        st.dataframe(detail_df, use_container_width=True, hide_index=True, height=350)

        # Download detail as CSV
        csv_detail = detail_df.to_csv(index=False).encode("utf-8")
        st.download_button("⬇️ Download Detail CSV", csv_detail,
                           file_name="debt_detail.csv", mime="text/csv")

        if unmapped_project_names:
            with st.expander("Country Mapping Warnings", expanded=False):
                st.caption("These unpaid invoice project names could not be mapped to a country.")
                st.dataframe(
                    pd.DataFrame({"Unmapped Project Name": sorted(unmapped_project_names)}),
                    use_container_width=True,
                    hide_index=True,
                    height=min(280, 40 + 28 * len(unmapped_project_names)),
                )
    else:
        st.success("No unpaid invoices match the current filters.")

    st.markdown("---")

    # ── Summary table: grouped by project ────────────────────────────────────
    st.subheader("Debt Summary by Project")
    from collections import defaultdict
    proj_debt: dict = defaultdict(lambda: {"inv_nos": set(), "total": 0.0, "country": ""})
    for i in debt_inv:
        key = i.project_name
        if i.invoice_number:
            proj_debt[key]["inv_nos"].add(str(int(i.invoice_number)))
        proj_debt[key]["total"]   += i.payment_amount
        resolved_country = _normalize_country(_safe_str(_get_country(i.project_name)).strip())
        if not resolved_country:
            invoice_number_hint = _safe_int(i.invoice_number, default=0)
            if invoice_number_hint:
                resolved_country = _safe_str(invoice_country_hint.get(invoice_number_hint, "")).strip()
        if not resolved_country:
            resolved_country = _infer_country_from_project_name(i.project_name)
        if resolved_country and not _safe_str(proj_debt[key]["country"]).strip():
            proj_debt[key]["country"] = resolved_country

    summary_rows = [{
        "Project Name":    name,
        "Country":         d["country"],
        "Invoice Numbers": ", ".join(sorted(d["inv_nos"])),
        "Total Debt (€)":  int(round(d["total"])),
    } for name, d in sorted(proj_debt.items(), key=lambda x: -x[1]["total"])]

    if summary_rows:
        summary_df = pd.DataFrame(summary_rows)

        def color_debt_amt(val):
            try:
                return "color: #E74C3C; font-weight: bold" if float(val) > 0 else ""
            except Exception:
                return ""

        st.dataframe(
            summary_df.style.map(color_debt_amt, subset=["Total Debt (€)"]),
            use_container_width=True, hide_index=True, height=400,
        )

        # Download summary as CSV
        csv_summary = summary_df.to_csv(index=False).encode("utf-8")
        st.download_button("⬇️ Download Summary CSV", csv_summary,
                           file_name="debt_summary.csv", mime="text/csv")

        # Download summary as PDF (split Y1 vs Y2+)
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors as rl_colors
            from io import BytesIO

            # Split unpaid invoices into Y1, paid trials, and Y2+
            y1_inv  = [i for i in debt_inv if _is_new_installation_category(i)]
            trial_inv = [i for i in debt_inv if _is_paid_trial_category(i)]
            y2_inv  = [i for i in debt_inv if _is_maintenance_category(i)]

            def _proj_debt_rows(inv_list):
                pd_: dict = defaultdict(lambda: {"inv_nos": set(), "total": 0.0, "country": ""})
                for i in inv_list:
                    if i.invoice_number:
                        pd_[i.project_name]["inv_nos"].add(str(int(i.invoice_number)))
                    pd_[i.project_name]["total"]   += i.payment_amount
                    resolved_country = _normalize_country(_safe_str(_get_country(i.project_name)).strip())
                    if not resolved_country:
                        invoice_number_hint = _safe_int(i.invoice_number, default=0)
                        if invoice_number_hint:
                            resolved_country = _safe_str(invoice_country_hint.get(invoice_number_hint, "")).strip()
                    if not resolved_country:
                        resolved_country = _infer_country_from_project_name(i.project_name)
                    if resolved_country and not _safe_str(pd_[i.project_name]["country"]).strip():
                        pd_[i.project_name]["country"] = resolved_country
                return sorted(pd_.items(), key=lambda x: -x[1]["total"])

            def _make_section_table(rows, rl_colors):
                hdr_color = rl_colors.HexColor("#1B3A6B")
                tbl_data = [["Project Name", "Country", "Invoice #s", "Debt (€)"]]
                for name, d in rows:
                    tbl_data.append([
                        name, d["country"],
                        ", ".join(sorted(d["inv_nos"])),
                        f"€{d['total']:,.0f}",
                    ])
                t = Table(tbl_data, repeatRows=1,
                          colWidths=[7.5*cm, 2*cm, 5*cm, 3*cm])
                t.setStyle(TableStyle([
                    ("BACKGROUND",     (0, 0), (-1, 0), hdr_color),
                    ("TEXTCOLOR",      (0, 0), (-1, 0), rl_colors.white),
                    ("FONTNAME",       (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE",       (0, 0), (-1, -1), 8),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                     [rl_colors.white, rl_colors.HexColor("#EBF5FB")]),
                    ("GRID",           (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#BDC3C7")),
                    ("BOTTOMPADDING",  (0, 0), (-1, -1), 4),
                    ("TOPPADDING",     (0, 0), (-1, -1), 4),
                    ("ALIGN",          (3, 0), (3, -1), "RIGHT"),
                ]))
                return t

            pdf_buf = BytesIO()
            doc = SimpleDocTemplate(pdf_buf, pagesize=A4,
                                    leftMargin=1.5*cm, rightMargin=1.5*cm,
                                    topMargin=1.5*cm, bottomMargin=1.5*cm)
            styles = getSampleStyleSheet()
            sec_style = ParagraphStyle("sec", parent=styles["Heading2"],
                                       textColor=rl_colors.HexColor("#1B3A6B"))
            elems = []
            elems.append(Paragraph("<b>CaddyCheck CRM — Debt Report</b>", styles["Title"]))
            elems.append(Paragraph(
                f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d')}  |  "
                f"Total Debt: €{total_debt_amt:,.0f}  |  Projects: {proj_with_debt}",
                styles["Normal"]))
            elems.append(Spacer(1, 0.4*cm))

            # Section A: Y1
            y1_total = sum(i.payment_amount for i in y1_inv)
            elems.append(Paragraph(
                f"A. New Installation Debt (Y1)  —  €{y1_total:,.0f}", sec_style))
            elems.append(Spacer(1, 0.2*cm))
            y1_rows = _proj_debt_rows(y1_inv)
            if y1_rows:
                elems.append(_make_section_table(y1_rows, rl_colors))
            else:
                elems.append(Paragraph("No Y1 unpaid invoices.", styles["Normal"]))
            elems.append(Spacer(1, 0.6*cm))

            # Section B: Paid Trials
            trial_total = sum(i.payment_amount for i in trial_inv)
            elems.append(Paragraph(
                f"B. Paid Trials  —  €{trial_total:,.0f}", sec_style))
            elems.append(Spacer(1, 0.2*cm))
            trial_rows = _proj_debt_rows(trial_inv)
            if trial_rows:
                elems.append(_make_section_table(trial_rows, rl_colors))
            else:
                elems.append(Paragraph("No unpaid paid-trial invoices.", styles["Normal"]))
            elems.append(Spacer(1, 0.6*cm))

            # Section C: Y2+
            y2_total = sum(i.payment_amount for i in y2_inv)
            elems.append(Paragraph(
                f"C. Maintenance Debt (Y2+)  —  €{y2_total:,.0f}", sec_style))
            elems.append(Spacer(1, 0.2*cm))
            y2_rows = _proj_debt_rows(y2_inv)
            if y2_rows:
                elems.append(_make_section_table(y2_rows, rl_colors))
            else:
                elems.append(Paragraph("No Y2+ unpaid invoices.", styles["Normal"]))

            doc.build(elems)
            debt_pdf_bytes = pdf_buf.getvalue()
            st.download_button("⬇️ Download Debt PDF", debt_pdf_bytes,
                               file_name="debt_report.pdf", mime="application/pdf")

            st.markdown("### Email Debt Report")
            email_cfg = get_email_config()
            default_subject = f"Debt Report - {datetime.datetime.now().strftime('%Y-%m-%d')}"
            default_body = (
                "Dear Team,\n\n"
                f"Please find attached the current debt report.\n\n"
                f"Total debt: €{total_debt_amt:,.0f}\n"
                f"Y1 debt: €{y1_total:,.0f}\n"
                f"Y2+ debt: €{y2_total:,.0f}\n"
                f"Paid Trials debt: €{trial_total:,.0f}\n"
                f"Projects with debt: {proj_with_debt}\n\n"
                "Best regards,\n"
                "CaddyCheck CRM"
            )

            with st.form("debt_report_email_form"):
                debt_to_addrs = st.text_input(
                    "To (comma-separated)",
                    ", ".join(email_cfg.get("default_recipients", [])),
                    key="debt_report_email_to",
                )
                debt_cc_addrs = st.text_input(
                    "CC (comma-separated)",
                    ", ".join(email_cfg.get("default_cc", [])),
                    key="debt_report_email_cc",
                )
                debt_subject = st.text_input(
                    "Subject",
                    default_subject,
                    key="debt_report_email_subject",
                )
                debt_body = st.text_area(
                    "Body",
                    default_body,
                    height=140,
                    key="debt_report_email_body",
                )
                debt_send_btn = st.form_submit_button("Send Debt Report")

            if debt_send_btn:
                if not email_cfg.get("smtp_username"):
                    st.error("SMTP not configured. Go to ⚙️ Settings to set up email.")
                else:
                    from services.email_service import send_invoice_email
                    import tempfile

                    debt_recipients = [r.strip() for r in debt_to_addrs.split(",") if r.strip()]
                    debt_cc_list = [c.strip() for c in debt_cc_addrs.split(",") if c.strip()]
                    if not debt_recipients:
                        st.error("Enter at least one recipient email address.")
                    else:
                        with tempfile.TemporaryDirectory() as tmp_dir:
                            debt_pdf_path = Path(tmp_dir) / "debt_report.pdf"
                            debt_pdf_path.write_bytes(debt_pdf_bytes)
                            try:
                                send_invoice_email(
                                    attachment_path=debt_pdf_path,
                                    recipients=debt_recipients,
                                    cc=debt_cc_list,
                                    subject=debt_subject,
                                    body=debt_body,
                                    config=email_cfg,
                                )
                                st.success("Debt report email sent.")
                            except Exception as exc:
                                st.error(f"Debt report email failed: {exc}")
        except Exception as e:
            st.warning(f"PDF export unavailable: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: MONTHLY INVOICE
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📅 Monthly Invoice":
    st.title("📅 Monthly Invoice")

    col1, col2, col3 = st.columns([2, 1, 2])
    with col1:
        sel_month = st.selectbox("Month", MONTH_ORDER, index=datetime.date.today().month - 1)
    with col2:
        sel_year = st.number_input("Year", min_value=2015, max_value=2035,
                                   value=datetime.date.today().year, step=1)
    with col3:
        month_projects = get_monthly_invoice_projects(projects, sel_month, int(sel_year))
        suggested_inv_no = _suggest_month_invoice_number(month_projects, invoices, int(sel_year), _data_path)
        invoice_number = st.number_input("Invoice Number",
                                         min_value=1, value=suggested_inv_no, step=1)
    inv_no = int(invoice_number)

    st.markdown(f"**{len(month_projects)} project(s)** billed in **{sel_month}**")

    if not month_projects:
        st.warning("No projects found for the selected month.")
    else:
        preview_rows = get_invoice_preview_data(month_projects, sel_month, int(sel_year))
        preview_total_amount = sum(
            float(r["line_total"])
            for r in preview_rows
            if isinstance(r.get("line_total"), (int, float)) and r.get("rate") != "TOTAL"
        )
        preview_df = pd.DataFrame([{
            "Project": _safe_str(r["project_name"]),
            "# Cams": _safe_str(r["num_cams"]),
            "Maint. Year": _safe_str(r["maintenance_year"]),
            "Rate (€)": f"€{r['rate']:,.0f}" if isinstance(r["rate"], (int, float)) else _safe_str(r["rate"]),
            "Line Total (€)": f"€{r['line_total']:,.0f}" if isinstance(r["line_total"], (int, float)) else "",
        } for r in preview_rows])
        st.subheader("Invoice Preview")
        st.dataframe(preview_df, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.subheader("Monthly Invoice Status")
        st.caption(
            "Combined monthly invoices are inferred when the same invoice number appears on multiple projects, which matches the shared invoices introduced from Aug 2025 onward."
        )

        monthly_invoice_summaries = group_monthly_invoices(invoices)
        if monthly_invoice_summaries:
            selected_monthly_summary = next(
                (summary for summary in monthly_invoice_summaries if summary.invoice_number == inv_no),
                None,
            )
            if selected_monthly_summary and (
                round(selected_monthly_summary.total_amount, 2) != round(preview_total_amount, 2)
                or selected_monthly_summary.project_count != len(month_projects)
            ):
                st.warning(
                    f"Saved ledger rows for invoice #{inv_no} still total €{selected_monthly_summary.total_amount:,.0f} "
                    f"across {selected_monthly_summary.project_count} project(s), while the current preview totals "
                    f"€{preview_total_amount:,.0f} across {len(month_projects)} project(s). "
                    "Use Save to Invoice Ledger to replace the saved batch."
                )

            monthly_years = sorted({summary.year for summary in monthly_invoice_summaries if summary.year}, reverse=True)
            monthly_statuses = sorted({summary.status for summary in monthly_invoice_summaries})

            mi1, mi2 = st.columns(2)
            monthly_year_filter = mi1.selectbox(
                "Monthly Invoice Year",
                ["All"] + [str(year) for year in monthly_years],
                key="monthly_invoice_status_year",
            )
            monthly_status_filter = mi2.selectbox(
                "Monthly Invoice Status",
                ["All"] + monthly_statuses,
                key="monthly_invoice_status_filter",
            )

            filtered_monthly_summaries = [
                summary for summary in monthly_invoice_summaries
                if (monthly_year_filter == "All" or str(summary.year or "") == monthly_year_filter)
                and (monthly_status_filter == "All" or summary.status == monthly_status_filter)
            ]

            mm1, mm2, mm3, mm4 = st.columns(4)
            mm1.metric("Monthly invoices", len(filtered_monthly_summaries))
            mm2.metric(
                "Grouped amount",
                f"€{sum(summary.total_amount for summary in filtered_monthly_summaries):,.0f}",
            )
            mm3.metric(
                "Paid monthly invoices",
                sum(1 for summary in filtered_monthly_summaries if summary.status == "Paid"),
            )
            mm4.metric(
                "Open / mixed",
                sum(1 for summary in filtered_monthly_summaries if summary.status != "Paid"),
            )

            monthly_status_df = pd.DataFrame([
                {
                    "Invoice #": str(summary.invoice_number),
                    "Year": _safe_str(summary.year or ""),
                    "Projects": summary.project_count,
                    "Total (€)": f"{summary.total_amount:,.0f}",
                    "Status": summary.status,
                    "Last Payment": summary.last_payment_date.strftime("%Y-%m-%d") if summary.last_payment_date else "",
                    "Included Projects": ", ".join(summary.project_names),
                }
                for summary in filtered_monthly_summaries
            ])

            def color_monthly_status(val):
                value = str(val).strip().lower()
                if value == "paid":
                    return "color: #27AE60; font-weight: bold"
                if value in {"partial", "mixed", "paid / cancelled"}:
                    return "color: #F39C12; font-weight: bold"
                if value in {"unpaid", "unpaid / cancelled"}:
                    return "color: #E74C3C; font-weight: bold"
                if value == "cancelled":
                    return "color: #7F8C8D; font-weight: bold"
                return ""

            st.dataframe(
                monthly_status_df.style.map(color_monthly_status, subset=["Status"]),
                use_container_width=True,
                hide_index=True,
                height=260,
                column_config={
                    "Included Projects": st.column_config.TextColumn(width="large"),
                },
            )

            with st.expander("Show Full Included Projects", expanded=False):
                for summary in filtered_monthly_summaries:
                    st.markdown(
                        f"**Invoice #{summary.invoice_number}**  |  {summary.project_count} project(s)  |  {summary.status}"
                    )
                    st.write(", ".join(summary.project_names))
        else:
            st.info("No combined monthly invoices were found in the ledger yet.")

        if CAN_EDIT:
            st.markdown("---")
            st.subheader("Generate Invoice")

            from config.settings import get_data_paths
            from services.pdf_service import generate_invoice_pdf
            paths = get_data_paths()
            month_abbr = sel_month[:3]
            pdf_filename = f"CC_M-inv_{inv_no}_{month_abbr}_{int(sel_year)}.pdf"
            xlsx_filename = f"CC_M-inv_{inv_no}_{month_abbr}_{int(sel_year)}.xlsx"

            col_pdf, col_xlsx = st.columns(2)

            with col_pdf:
                try:
                    pdf_bytes = generate_invoice_pdf(
                        projects=month_projects,
                        month_name=sel_month,
                        year=int(sel_year),
                        invoice_number=inv_no,
                    )
                    st.download_button(
                        label="Download PDF Invoice",
                        data=pdf_bytes,
                        file_name=pdf_filename,
                        mime="application/pdf",
                        type="primary",
                    )
                except Exception as e:
                    st.error(f"PDF generation failed: {e}")

            with col_xlsx:
                template_ok = paths["invoice_template"].exists()
                if not template_ok:
                    st.warning("Excel template not found.")
                else:
                    if st.button("Generate Excel Invoice"):
                        import tempfile
                        with tempfile.TemporaryDirectory() as tmp_dir:
                            try:
                                out_path = generate_monthly_invoice(
                                    projects=month_projects,
                                    month_name=sel_month,
                                    year=int(sel_year),
                                    invoice_number=inv_no,
                                    output_dir=Path(tmp_dir),
                                    template_path=paths["invoice_template"],
                                )
                                with open(out_path, "rb") as f:
                                    file_bytes = f.read()
                                st.download_button(
                                    label=f"Download {out_path.name}",
                                    data=file_bytes,
                                    file_name=out_path.name,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                )
                            except Exception as e:
                                st.error(f"Excel generation failed: {e}")

            # ── Save to Ledger ─────────────────────────────────────────────
            st.markdown("---")
            st.subheader("Save to Invoice Ledger")
            st.caption(f"Saves invoice #{inv_no} rows for all {len(month_projects)} project(s) to Invoice Details. If invoice #{inv_no} already exists, its rows are replaced.")
            if st.button("💾 Save Invoice to Ledger", type="primary"):
                try:
                    n = _append_invoice_rows(
                        invoice_number=inv_no,
                        projects=month_projects,
                        year=int(sel_year),
                        source_name=_data_path,
                        description=f"Maintenance Monthly-{sel_month}-{sel_year}",
                    )
                    st.cache_data.clear()
                    st.session_state["_flash_success"] = (
                        f"Saved invoice #{inv_no} with {n} row(s) to Invoice Details."
                    )
                    st.session_state["_flash_success_page"] = "📅 Monthly Invoice"
                    st.rerun()
                except Exception as e:
                    st.error(f"Failed to save: {e}")

            # ── Email section ──────────────────────────────────────────────
            st.markdown("---")
            st.subheader("Send Invoice by Email")
            st.caption("Sending the PDF does not add invoice rows to the ledger unless you enable the option below.")
            email_cfg = get_email_config()

            with st.form("email_form"):
                to_addrs = st.text_input("To (comma-separated)",
                                         ", ".join(email_cfg.get("default_recipients", [])))
                cc_addrs = st.text_input("CC (comma-separated)",
                                         ", ".join(email_cfg.get("default_cc", [])))
                subject_tpl = email_cfg.get("default_subject_template", "Monthly Invoice - {month} {year}")
                subject = st.text_input("Subject", subject_tpl.format(month=sel_month, year=sel_year))
                body_tpl = email_cfg.get("default_body_template", "")
                body = st.text_area("Body", body_tpl.format(month=sel_month, year=sel_year), height=150)
                save_to_ledger_on_send = st.checkbox("Also save invoice to ledger before sending", value=True)
                send_btn = st.form_submit_button("Send Email")

            if send_btn:
                if not email_cfg.get("smtp_username"):
                    st.error("SMTP not configured. Go to ⚙️ Settings to set up email.")
                else:
                    from services.email_service import send_invoice_email
                    import tempfile
                    with tempfile.TemporaryDirectory() as tmp_dir:
                        try:
                            ledger_rows_added = None
                            if save_to_ledger_on_send:
                                ledger_rows_added = _append_invoice_rows(
                                    invoice_number=inv_no,
                                    projects=month_projects,
                                    year=int(sel_year),
                                    source_name=_data_path,
                                    description=f"Maintenance Monthly-{sel_month}-{sel_year}",
                                )
                            out_path = generate_monthly_invoice_pdf(
                                projects=month_projects,
                                month_name=sel_month,
                                year=int(sel_year),
                                invoice_number=inv_no,
                                output_dir=Path(tmp_dir),
                            )
                            archived_pdf_path = archive_sent_invoice_pdf(out_path)
                            recipients = [r.strip() for r in to_addrs.split(",") if r.strip()]
                            cc_list = [c.strip() for c in cc_addrs.split(",") if c.strip()]
                            send_invoice_email(
                                attachment_path=out_path,
                                recipients=recipients,
                                cc=cc_list,
                                subject=subject,
                                body=body,
                                config=email_cfg,
                            )
                            append_sent_invoice_log({
                                "sent_at": datetime.datetime.utcnow().isoformat(),
                                "invoice_number": inv_no,
                                "month": sel_month,
                                "year": int(sel_year),
                                "pdf_filename": out_path.name,
                                "pdf_archive_path": str(archived_pdf_path),
                                "recipients": recipients,
                                "cc": cc_list,
                                "subject": subject,
                                "project_count": len(month_projects),
                                "total_amount": preview_total_amount,
                                "saved_to_ledger": bool(save_to_ledger_on_send),
                                "ledger_rows_added": ledger_rows_added,
                            })
                            if save_to_ledger_on_send:
                                st.cache_data.clear()
                            if save_to_ledger_on_send and ledger_rows_added is not None:
                                st.session_state["_flash_success"] = (
                                    f"Email sent with PDF attachment: {out_path.name}. Ledger rows saved: {ledger_rows_added}."
                                )
                            else:
                                st.session_state["_flash_success"] = (
                                    f"Email sent with PDF attachment: {out_path.name}."
                                )
                            st.session_state["_flash_success_page"] = "📅 Monthly Invoice"
                            st.rerun()
                        except Exception as e:
                            st.error(f"Email failed: {e}")

            st.markdown("---")
            st.subheader("Sent PDF Invoices")
            sent_invoice_rows = load_sent_invoices_log()
            if sent_invoice_rows:
                from config.settings import get_data_paths
                import tempfile

                matching_sent_entry = next(
                    (
                        row for row in reversed(sent_invoice_rows)
                        if _safe_int(row.get("invoice_number"), default=0) == inv_no
                        and _safe_str(row.get("month", "")).strip().lower() == sel_month.strip().lower()
                        and _safe_int(row.get("year"), default=0) == int(sel_year)
                    ),
                    None,
                )
                logged_total_amount = _safe_float(
                    matching_sent_entry.get("total_amount", 0.0) if matching_sent_entry else 0.0,
                    0.0,
                )
                if matching_sent_entry and abs(logged_total_amount - preview_total_amount) >= 0.5:
                    st.warning(
                        f"Logged sent amount for invoice #{inv_no} is €{logged_total_amount:,.0f}, "
                        f"but the current preview total is €{preview_total_amount:,.0f}."
                    )
                    if st.button("Repair Logged Sent Amount", key="repair_sent_invoice_total"):
                        repaired = False
                        for row in reversed(sent_invoice_rows):
                            if (
                                _safe_int(row.get("invoice_number"), default=0) == inv_no
                                and _safe_str(row.get("month", "")).strip().lower() == sel_month.strip().lower()
                                and _safe_int(row.get("year"), default=0) == int(sel_year)
                            ):
                                row["total_amount"] = preview_total_amount
                                repaired = True
                                break
                        if repaired:
                            save_sent_invoices_log(sent_invoice_rows)
                            st.success(
                                f"Repaired sent-log total for invoice #{inv_no} to €{preview_total_amount:,.0f}."
                            )
                            st.rerun()

                sent_history_rows = []
                for row in reversed(sent_invoice_rows):
                    bucket = _safe_str(row.get("pdf_storage_bucket", "")).strip()
                    storage_path = _safe_str(row.get("pdf_storage_path", "")).strip()
                    signed_url = ""
                    if bucket and storage_path:
                        try:
                            signed_url = create_sent_invoice_pdf_signed_url_supabase(
                                bucket, storage_path
                            ) or ""
                        except Exception:
                            signed_url = ""
                    sent_at_text = _safe_str(row.get("sent_at", "")).replace("T", " ")[:19]
                    sent_month = _safe_str(row.get("month", "")).strip()
                    sent_year = _safe_int(row.get("year"), default=0)
                    sent_history_rows.append({
                        "Invoice #": _safe_int(row.get("invoice_number"), default=0),
                        "Type": "Monthly Invoice",
                        "For Month": f"{sent_month} {sent_year}".strip(),
                        "Sent At": sent_at_text,
                        "Total (€)": _safe_float(row.get("total_amount", 0.0), 0.0),
                        "Filename": _safe_str(row.get("pdf_filename", "")),
                        "Source PDF": signed_url,
                    })
                if sent_history_rows:
                    st.dataframe(
                        sent_history_rows,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Source PDF": st.column_config.LinkColumn(
                                "Source PDF", display_text="Download"
                            ),
                        },
                    )

                with st.expander("📎 Backfill Sent Invoice PDFs", expanded=False):
                    st.caption(
                        "Upload PDF(s) for monthly invoices that were sent before "
                        "PDFs were stored in the cloud. The invoice number is read "
                        "from the filename (e.g. `CC_M-inv_8669_May_2025.pdf`). "
                        "If a matching sent entry exists without a stored PDF, the "
                        "file is uploaded and the entry is updated."
                    )
                    backfill_files = st.file_uploader(
                        "Sent invoice PDF(s)",
                        type=["pdf"],
                        accept_multiple_files=True,
                        key="backfill_sent_invoice_pdfs",
                    )
                    if backfill_files and st.button(
                        "Attach to Sent Invoice History",
                        key="backfill_sent_invoice_btn",
                        type="primary",
                    ):
                        import re as _bf_re
                        import tempfile as _bf_tempfile

                        attached = 0
                        skipped = []
                        errors = []
                        updated_rows = list(sent_invoice_rows)
                        for up in backfill_files:
                            fname = up.name
                            m = _bf_re.search(r"(\d{4,})", fname)
                            if not m:
                                skipped.append(f"{fname}: no invoice number in filename")
                                continue
                            target_inv = int(m.group(1))
                            # find latest matching sent row without storage
                            target_row = None
                            for row in reversed(updated_rows):
                                if _safe_int(row.get("invoice_number"), default=0) != target_inv:
                                    continue
                                if _safe_str(row.get("pdf_storage_path", "")).strip():
                                    continue
                                target_row = row
                                break
                            if target_row is None:
                                # fallback: any matching invoice number even if already has storage
                                for row in reversed(updated_rows):
                                    if _safe_int(row.get("invoice_number"), default=0) == target_inv:
                                        target_row = row
                                        break
                            if target_row is None:
                                skipped.append(f"{fname}: no sent entry for invoice #{target_inv}")
                                continue
                            try:
                                file_bytes = up.read()
                                with _bf_tempfile.NamedTemporaryFile(
                                    suffix=".pdf", delete=False
                                ) as tmp:
                                    tmp.write(file_bytes)
                                    tmp_path = Path(tmp.name)
                                try:
                                    meta = upload_sent_invoice_pdf_supabase(tmp_path) or {}
                                finally:
                                    try:
                                        tmp_path.unlink()
                                    except Exception:
                                        pass
                                if meta.get("pdf_storage_bucket") and meta.get("pdf_storage_path"):
                                    target_row["pdf_storage_bucket"] = meta["pdf_storage_bucket"]
                                    target_row["pdf_storage_path"] = meta["pdf_storage_path"]
                                    if not _safe_str(target_row.get("pdf_filename", "")).strip():
                                        target_row["pdf_filename"] = fname
                                    attached += 1
                                else:
                                    errors.append(f"{fname}: upload returned no storage path")
                            except Exception as exc:
                                errors.append(f"{fname}: {exc}")

                        if attached:
                            try:
                                save_sent_invoices_log(updated_rows)
                                st.cache_data.clear()
                                st.success(f"Attached {attached} PDF(s) to sent invoice history.")
                            except Exception as exc:
                                st.error(f"Saved storage but could not update sent invoice log: {exc}")
                        if skipped:
                            st.warning("Skipped:\n- " + "\n- ".join(skipped))
                        if errors:
                            st.error("Errors:\n- " + "\n- ".join(errors))
                        if attached:
                            st.rerun()

                sent_download_options = {
                    f"#{_safe_int(row.get('invoice_number'), default=0)} | {_safe_str(row.get('month', ''))} {_safe_int(row.get('year'), default=0)} | {_safe_str(row.get('pdf_filename', ''))}": row
                    for row in reversed(sent_invoice_rows)
                }
                if sent_download_options:
                    selected_sent_label = st.selectbox(
                        "Sent PDF entry",
                        list(sent_download_options.keys()),
                        key="sent_pdf_entry",
                    )
                    selected_sent_row = sent_download_options[selected_sent_label]
                    archive_path_text = _safe_str(selected_sent_row.get("pdf_archive_path", "")).strip()
                    archive_path = Path(archive_path_text) if archive_path_text else None
                    download_bytes = None
                    download_name = _safe_str(selected_sent_row.get("pdf_filename", "")).strip() or "invoice.pdf"
                    if archive_path is None or not archive_path.exists():
                        pdf_filename = _safe_str(selected_sent_row.get("pdf_filename", "")).strip()
                        if pdf_filename:
                            fallback_archive_path = get_data_paths()["output_dir"] / "sent_invoices" / pdf_filename
                            if fallback_archive_path.exists():
                                archive_path = fallback_archive_path

                    if archive_path and archive_path.exists():
                        with open(archive_path, "rb") as pdf_file:
                            download_bytes = pdf_file.read()
                        download_name = archive_path.name
                    else:
                        storage_bucket = _safe_str(selected_sent_row.get("pdf_storage_bucket", "")).strip()
                        storage_path = _safe_str(selected_sent_row.get("pdf_storage_path", "")).strip()
                        if storage_bucket and storage_path:
                            try:
                                download_bytes = download_sent_invoice_pdf_supabase(storage_bucket, storage_path)
                                download_name = Path(storage_path).name or download_name
                            except Exception as exc:
                                st.caption(f"Supabase PDF archive is not available yet: {exc}")

                    if download_bytes is not None:
                        st.download_button(
                            "Download Sent PDF",
                            data=download_bytes,
                            file_name=download_name,
                            mime="application/pdf",
                            key="download_sent_pdf",
                        )
                    else:
                        st.info("No archived PDF is stored for this sent entry yet.")
                        sent_month = _safe_str(selected_sent_row.get("month", "")).strip()
                        sent_year = _safe_int(selected_sent_row.get("year"), default=0)
                        sent_inv_no = _safe_int(selected_sent_row.get("invoice_number"), default=0)
                        if sent_month and sent_year and sent_inv_no:
                            if st.button("Create Downloadable PDF For This Entry", key="rebuild_sent_pdf"):
                                sent_month_projects = get_monthly_invoice_projects(projects, sent_month, sent_year)
                                with tempfile.TemporaryDirectory() as tmp_dir:
                                    rebuilt_pdf_path = generate_monthly_invoice_pdf(
                                        projects=sent_month_projects,
                                        month_name=sent_month,
                                        year=sent_year,
                                        invoice_number=sent_inv_no,
                                        output_dir=Path(tmp_dir),
                                    )
                                    archived_pdf_path = archive_sent_invoice_pdf(rebuilt_pdf_path)
                                for row in reversed(sent_invoice_rows):
                                    if row is selected_sent_row:
                                        row["pdf_archive_path"] = str(archived_pdf_path)
                                        break
                                save_sent_invoices_log(sent_invoice_rows)
                                st.success(f"Archived PDF created for invoice #{sent_inv_no}.")
                                st.rerun()

                sent_invoice_df = pd.DataFrame([
                    {
                        "Sent At": _safe_str(row.get("sent_at", "")).replace("T", " ")[:19],
                        "Invoice #": _safe_int(row.get("invoice_number"), default=0),
                        "Month": _safe_str(row.get("month", "")),
                        "Year": _safe_int(row.get("year"), default=0),
                        "PDF": _safe_str(row.get("pdf_filename", "")),
                        "Projects": _safe_int(row.get("project_count"), default=0),
                        "Total (€)": f"€{_safe_float(row.get('total_amount', 0.0)):,.0f}",
                        "To": ", ".join(row.get("recipients", [])),
                        "Saved To Ledger": "Yes" if row.get("saved_to_ledger") else "No",
                    }
                    for row in reversed(sent_invoice_rows)
                ])
                st.dataframe(sent_invoice_df, use_container_width=True, hide_index=True, height=260)
            else:
                st.info("No invoice PDFs have been logged as sent yet.")
        else:
            st.info("Invoice generation requires Admin access.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: SETTINGS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "⚙️ Settings":
    st.title("⚙️ Settings")

    if not CAN_EDIT:
        st.info("Settings can only be changed by Admin users.")
    else:
        email_cfg = get_email_config()

        st.subheader("Email / SMTP Configuration")
        with st.form("smtp_form"):
            col1, col2 = st.columns(2)
            with col1:
                smtp_host = st.text_input("SMTP Host", email_cfg.get("smtp_host", "smtp.gmail.com"))
                smtp_user = st.text_input("Username / Email", email_cfg.get("smtp_username", ""))
                sender_name = st.text_input("Sender Name", email_cfg.get("sender_name", "CaddyCheck CRM"))
            with col2:
                smtp_port = st.number_input("SMTP Port", min_value=1, max_value=65535,
                                            value=int(email_cfg.get("smtp_port", 587)))
                smtp_pass = st.text_input("Password", email_cfg.get("smtp_password", ""),
                                          type="password")
                sender_email = st.text_input("Sender Email", email_cfg.get("sender_email", ""))
            use_tls = st.checkbox("Use STARTTLS", value=bool(email_cfg.get("smtp_use_tls", True)))

            st.markdown("---")
            st.subheader("Default Recipients")
            recipients = st.text_input("To (comma-separated)",
                                       ", ".join(email_cfg.get("default_recipients", [])))
            cc = st.text_input("CC (comma-separated)",
                                ", ".join(email_cfg.get("default_cc", [])))
            subject_tpl = st.text_input("Subject Template",
                                         email_cfg.get("default_subject_template",
                                                        "Monthly Invoice - {month} {year}"))
            body_tpl = st.text_area("Body Template",
                                     email_cfg.get("default_body_template", ""), height=120)

            save_btn = st.form_submit_button("Save Settings", type="primary")

        if save_btn:
            new_cfg = {
                "smtp_host": smtp_host.strip(),
                "smtp_port": int(smtp_port),
                "smtp_use_tls": use_tls,
                "smtp_username": smtp_user.strip(),
                "smtp_password": smtp_pass,
                "sender_name": sender_name.strip(),
                "sender_email": sender_email.strip(),
                "default_recipients": [r.strip() for r in recipients.split(",") if r.strip()],
                "default_cc": [c.strip() for c in cc.split(",") if c.strip()],
                "default_subject_template": subject_tpl.strip(),
                "default_body_template": body_tpl,
            }
            try:
                st.session_state["_smtp_password_override"] = smtp_pass
                save_email_config(new_cfg)
                st.success("Settings saved. SMTP password is kept only for this session unless configured in Streamlit secrets.")
            except Exception as e:
                st.error(f"Save failed: {e}")

        st.markdown("---")
        st.subheader("Test SMTP Connection")
        if st.button("Test Connection"):
            from services.email_service import test_smtp_connection
            cfg_to_test = {
                "smtp_host": email_cfg.get("smtp_host", ""),
                "smtp_port": email_cfg.get("smtp_port", 587),
                "smtp_use_tls": email_cfg.get("smtp_use_tls", True),
                "smtp_username": email_cfg.get("smtp_username", ""),
                "smtp_password": email_cfg.get("smtp_password", ""),
            }
            try:
                success, msg = test_smtp_connection(cfg_to_test)
                if success:
                    st.success(msg)
                else:
                    st.error(msg)
            except Exception as e:
                st.error(str(e))


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: TICKETS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🎫 Tickets":
    from services.supabase_service import (
        get_tickets, create_ticket, update_ticket, delete_ticket,
    )

    st.title("🎫 Support Tickets")

    def _append_camera_to_title(base_title: str, camera_name: str) -> str:
        cleaned_title = _safe_str(base_title).strip()
        cleaned_camera = _safe_str(camera_name).strip().upper()
        if not cleaned_title:
            return ""
        if not cleaned_camera:
            return cleaned_title
        if cleaned_camera in cleaned_title.upper():
            return cleaned_title
        return f"{cleaned_title} - {cleaned_camera}"

    def _extract_camera_from_title(title: str) -> str:
        text = _safe_str(title).strip().upper()
        match = re.search(r"\bK(10|[1-9])(B|TD)\b", text)
        return match.group(0) if match else ""

    # ── Summary metrics ───────────────────────────────────────────────────────
    all_tickets = get_tickets()
    open_count     = sum(1 for t in all_tickets if t["status"] == "Open")
    inprog_count   = sum(1 for t in all_tickets if t["status"] == "In Progress")
    resolved_count = sum(1 for t in all_tickets if t["status"] in ("Resolved", "Closed"))

    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("Total Tickets",   len(all_tickets))
    mc2.metric("Open",            open_count)
    mc3.metric("In Progress",     inprog_count)
    mc4.metric("Resolved/Closed", resolved_count)

    st.markdown("---")

    # ── New ticket form ───────────────────────────────────────────────────────
    if CAN_EDIT:
        with st.expander("➕ New Ticket", expanded=False):
            with st.form("new_ticket_form"):
                project_names = sorted({p.project_name for p in projects})
                t_project  = st.selectbox("Project", project_names, key="nt_proj")
                t_title_choice = st.selectbox("Title", TICKET_TITLE_OPTIONS, index=0, key="nt_title_choice")
                t_title_custom = st.text_input("Custom title (optional)", key="nt_title_custom")
                t_camera_name = st.selectbox("Camera Name", TICKET_CAMERA_OPTIONS, index=0, key="nt_camera_name")
                t_desc     = st.text_area("Description", height=100, key="nt_desc")
                t_subcategory = st.selectbox("Sub-category", TICKET_SUBCATEGORY_OPTIONS, index=0, key="nt_subcat")
                t_priority = st.selectbox("Priority", ["Low", "Medium", "High", "Critical"], index=1, key="nt_prio")
                submitted  = st.form_submit_button("Create Ticket", type="primary")
            if submitted:
                selected_title = _safe_str(t_title_custom).strip() or _safe_str(t_title_choice).strip()
                ticket_title = _append_camera_to_title(selected_title, t_camera_name)
                if not ticket_title:
                    st.error("Title is required.")
                else:
                    try:
                        ticket = create_ticket(
                            t_project,
                            ticket_title,
                            t_desc.strip(),
                            t_priority,
                            t_subcategory,
                        )
                        st.success(f"Ticket {ticket.get('ticket_number', '')} created!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Failed to create ticket: {e}")

    # ── Filters ───────────────────────────────────────────────────────────────
    with st.expander("🔍 Filters", expanded=True):
        fc1, fc2, fc3, fc4, fc5 = st.columns(5)
        project_names_all = ["All"] + sorted({t["project_name"] for t in all_tickets})
        tf_proj   = fc1.selectbox("Project",  project_names_all, key="tf_proj")
        tf_status = fc2.selectbox("Status",   ["All", "Open", "In Progress", "Resolved", "Closed"], key="tf_status")
        tf_prio   = fc3.selectbox("Priority", ["All", "Critical", "High", "Medium", "Low"], key="tf_prio")
        tf_subcat = fc4.selectbox("Sub-category", ["All"] + TICKET_SUBCATEGORY_OPTIONS, key="tf_subcat")
        tf_search = fc5.text_input("Search title", key="tf_search")

    filtered_tickets = all_tickets
    if tf_proj != "All":
        filtered_tickets = [t for t in filtered_tickets if t["project_name"] == tf_proj]
    if tf_status != "All":
        filtered_tickets = [t for t in filtered_tickets if t["status"] == tf_status]
    if tf_prio != "All":
        filtered_tickets = [t for t in filtered_tickets if t["priority"] == tf_prio]
    if tf_subcat != "All":
        filtered_tickets = [t for t in filtered_tickets if _safe_str(t.get("subcategory")).strip() == tf_subcat]
    if tf_search.strip():
        filtered_tickets = [t for t in filtered_tickets if tf_search.lower() in t["title"].lower()]

    st.caption(f"Showing {len(filtered_tickets)} of {len(all_tickets)} tickets")

    # ── Ticket table ──────────────────────────────────────────────────────────
    PRIORITY_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    STATUS_ORDER   = {"Open": 0, "In Progress": 1, "Resolved": 2, "Closed": 3}
    filtered_tickets = sorted(
        filtered_tickets,
        key=lambda t: (STATUS_ORDER.get(t["status"], 9), PRIORITY_ORDER.get(t["priority"], 9)),
    )

    def _prio_color(val):
        return {"Critical": "color:#C0392B;font-weight:bold",
                "High":     "color:#E67E22;font-weight:bold",
                "Medium":   "color:#2980B9",
                "Low":      "color:#27AE60"}.get(val, "")

    def _status_color(val):
        return {"Open":        "color:#E74C3C;font-weight:bold",
                "In Progress": "color:#F39C12;font-weight:bold",
                "Resolved":    "color:#27AE60",
                "Closed":      "color:#95A5A6"}.get(val, "")

    if filtered_tickets:
        ticket_df = pd.DataFrame([{
            "Ticket #":    t["ticket_number"],
            "Project":     t["project_name"],
            "Title":       t["title"],
            "Sub-category": _safe_str(t.get("subcategory")),
            "Priority":    t["priority"],
            "Status":      t["status"],
            "Created":     t["created_at"][:10] if t.get("created_at") else "",
            "Updated":     t["updated_at"][:10] if t.get("updated_at") else "",
        } for t in filtered_tickets])

        st.dataframe(
            ticket_df.style
                .map(_prio_color,    subset=["Priority"])
                .map(_status_color,  subset=["Status"]),
            use_container_width=True,
            hide_index=True,
            height=350,
        )
    else:
        st.info("No tickets match the current filters.")

    # ── Edit / update a ticket ────────────────────────────────────────────────
    if CAN_EDIT and filtered_tickets:
        st.markdown("---")
        st.subheader("Update Ticket")
        ticket_options = {f"{t['ticket_number']} — {t['title']}": t for t in filtered_tickets}
        selected_label = st.selectbox("Select ticket to update", list(ticket_options.keys()), key="sel_ticket")
        sel_ticket = ticket_options[selected_label]

        with st.form("update_ticket_form"):
            uc1, uc2 = st.columns(2)
            new_status   = uc1.selectbox("Status",   ["Open", "In Progress", "Resolved", "Closed"],
                                          index=["Open", "In Progress", "Resolved", "Closed"].index(sel_ticket["status"])
                                          if sel_ticket["status"] in ["Open", "In Progress", "Resolved", "Closed"] else 0,
                                          key="upd_status")
            new_priority = uc2.selectbox("Priority", ["Low", "Medium", "High", "Critical"],
                                          index=["Low", "Medium", "High", "Critical"].index(sel_ticket["priority"])
                                          if sel_ticket["priority"] in ["Low", "Medium", "High", "Critical"] else 1,
                                          key="upd_prio")
            current_ticket_title = _safe_str(sel_ticket.get("title")).strip()
            current_camera = _extract_camera_from_title(current_ticket_title)
            clean_title_for_selection = re.sub(r"\s*-\s*K(10|[1-9])(B|TD)\s*$", "", current_ticket_title, flags=re.IGNORECASE).strip()
            if clean_title_for_selection in TICKET_TITLE_OPTIONS:
                upd_title_choice = st.selectbox(
                    "Title",
                    TICKET_TITLE_OPTIONS,
                    index=TICKET_TITLE_OPTIONS.index(clean_title_for_selection),
                    key="upd_title_choice",
                )
                upd_title_custom = st.text_input("Custom title (optional)", value="", key="upd_title_custom")
            else:
                upd_title_choice = st.selectbox("Title", TICKET_TITLE_OPTIONS, index=0, key="upd_title_choice")
                upd_title_custom = st.text_input(
                    "Custom title (optional)",
                    value=clean_title_for_selection,
                    key="upd_title_custom",
                )
            upd_camera_index = TICKET_CAMERA_OPTIONS.index(current_camera) if current_camera in TICKET_CAMERA_OPTIONS else 0
            upd_camera_name = st.selectbox(
                "Camera Name",
                TICKET_CAMERA_OPTIONS,
                index=upd_camera_index,
                key="upd_camera_name",
            )
            new_subcategory = st.selectbox(
                "Sub-category",
                TICKET_SUBCATEGORY_OPTIONS,
                index=TICKET_SUBCATEGORY_OPTIONS.index(_safe_str(sel_ticket.get("subcategory")).strip())
                if _safe_str(sel_ticket.get("subcategory")).strip() in TICKET_SUBCATEGORY_OPTIONS else 0,
                key="upd_subcat",
            )
            new_notes = st.text_area("Notes / update comment", value=sel_ticket.get("notes") or "", height=80, key="upd_notes")
            col_upd, col_del = st.columns([3, 1])
            upd_btn = col_upd.form_submit_button("💾 Update Ticket", type="primary")
            del_btn = col_del.form_submit_button("🗑️ Delete", type="secondary")

        if upd_btn:
            try:
                resolved_title = _safe_str(upd_title_custom).strip() or _safe_str(upd_title_choice).strip()
                resolved_title = _append_camera_to_title(resolved_title, upd_camera_name)
                update_ticket(
                    sel_ticket["id"],
                    title=resolved_title,
                    status=new_status,
                    priority=new_priority,
                    subcategory=new_subcategory,
                    notes=new_notes,
                )
                st.success("Ticket updated!")
                st.rerun()
            except Exception as e:
                st.error(f"Update failed: {e}")

        if del_btn:
            try:
                delete_ticket(sel_ticket["id"])
                st.success(f"Ticket {sel_ticket['ticket_number']} deleted.")
                st.rerun()
            except Exception as e:
                st.error(f"Delete failed: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: BANK PAYMENT
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🏦 Bank Payment":
    import hashlib

    from services.bank_pdf_service import parse_swift_pdf
    from services.supabase_service import (
        get_invoices_by_number, mark_invoice_row_paid,
        get_subscription, upsert_subscription, create_renewal_link,
        append_bank_payment_with_allocations, load_bank_payments, load_bank_payment_allocations,
        upload_bank_payment_pdf as upload_bank_payment_pdf_supabase,
        create_bank_payment_pdf_signed_url as create_bank_payment_pdf_signed_url_supabase,
    )
    from config.settings import append_bank_payment_log, load_bank_payments_log

    st.title("🏦 Bank Payment Import")

    if not CAN_EDIT:
        st.warning("Read-only mode — contact an admin to record payments.")
        st.stop()

    # ── Show renewal result from previous confirmation ────────────────────────
    if "_renewal_result" in st.session_state:
        _rr = st.session_state.pop("_renewal_result")
        st.success(
            f"✅ Invoice **#{_rr['inv_no']}** — marked **{_rr['count']}** project(s) as paid."
        )
        _base_url = st.secrets.get("app", {}).get("base_url", "").rstrip("/")
        if _rr["links"]:
            st.markdown("### 🔑 Renewal Links")
            st.caption(
                "Send each link to the customer. "
                "When opened, the subscription is activated automatically."
            )
            for _rl in _rr["links"]:
                _url = f"{_base_url}/?token={_rl['token']}" if _base_url else f"?token={_rl['token']}"
                with st.container(border=True):
                    _lc1, _lc2 = st.columns([2, 3])
                    _lc1.markdown(
                        f"**{_rl['project']}**  \n"
                        f"📅 Valid until: **{_rl['valid_until'].strftime('%d %b %Y')}**  \n"
                        f"📷 Cameras: **{_rl['cameras']}**"
                    )
                    _lc2.code(_url, language=None)
        st.markdown("---")

    def _render_invoice_lookup(inv_no: int, pay_date: datetime.date, key_prefix: str, payment_context: Optional[dict] = None):
        """Look up rows for inv_no and render selection + confirm UI."""
        if inv_no <= 0:
            st.info("Enter a valid invoice number to look up matching rows.")
            return

        with st.spinner("Looking up invoice in database…"):
            try:
                rows = get_invoices_by_number(inv_no)
            except Exception as exc:
                st.error(f"Lookup failed: {exc}")
                return

        if not rows:
            st.warning(f"No invoice rows found for invoice **#{inv_no}**.")
            return

        df = pd.DataFrame([{
            "id":           r["id"],
            "Project":      r["project_name"],
            "Maint. Year":  r["maintenance_year"],
            "Amount (€)":   _safe_float(r.get("payment_amount")),
            "Cameras":      _safe_int(r.get("cameras_number")),
            "Paid":         r.get("paid", "No"),
            "Payment Date": (r["payment_date"] or "")[:10] if r.get("payment_date") else "",
        } for r in rows])

        st.dataframe(
            df[["Project", "Maint. Year", "Amount (€)", "Cameras", "Paid", "Payment Date"]],
            use_container_width=True,
            hide_index=True,
        )

        unpaid_projects = list(df[df["Paid"] != "Yes"]["Project"])
        if not unpaid_projects:
            st.success(f"All rows for invoice **#{inv_no}** are already marked as paid.")
            return

        selected = st.multiselect(
            "Select project row(s) to mark as paid",
            options=list(df["Project"]),
            default=unpaid_projects,
            key=f"{key_prefix}_sel",
        )

        if not selected:
            return

        confirm_date = st.date_input(
            "Payment date to record",
            value=pay_date,
            key=f"{key_prefix}_confirm_date",
        )

        sel_df = df[df["Project"].isin(selected)][["Project", "Amount (€)", "Cameras"]].copy()
        st.caption("The following rows will be marked paid. If subscription tables exist, renewal links will also be generated.")
        st.dataframe(sel_df, use_container_width=True, hide_index=True)

        if st.button("✅ Confirm — Mark as Paid & Generate Renewal Links", type="primary", key=f"{key_prefix}_confirm"):
            errors = []
            renewal_links = []
            renewal_warnings = []
            allocation_rows = []
            total_applied = 0.0

            for _, row in df[df["Project"].isin(selected)].iterrows():
                proj = row["Project"]
                try:
                    # 1. Mark invoice row as paid
                    mark_invoice_row_paid(
                        db_id=int(row["id"]),
                        payment_date=confirm_date,
                        payment_amount=row["Amount (€)"],
                    )

                    amount_applied = _safe_float(row["Amount (€)"])
                    total_applied += amount_applied
                    allocation_rows.append({
                        "invoice_row_id": int(row["id"]),
                        "invoice_number": inv_no,
                        "project_name": proj,
                        "maintenance_year": _safe_str(row["Maint. Year"]),
                        "year": None,
                        "amount_applied": amount_applied,
                    })

                    try:
                        # 2. Compute new valid_until (extend from current expiry or today)
                        sub = get_subscription(proj)
                        if sub and sub.get("valid_until"):
                            current_until = datetime.date.fromisoformat(sub["valid_until"][:10])
                            base = max(current_until, confirm_date)
                        else:
                            base = confirm_date
                        try:
                            target_until = base.replace(year=base.year + 1)
                        except ValueError:           # Feb 29 edge case
                            target_until = base.replace(year=base.year + 1, day=28)

                        cameras = _safe_int(row["Cameras"])

                        # 3. Upsert subscription record
                        upsert_subscription(
                            project_name=proj,
                            valid_until=target_until,
                            cameras_allowed=cameras,
                            valid_from=confirm_date,
                        )

                        # 4. Generate renewal token
                        token = create_renewal_link(
                            project_name=proj,
                            target_valid_until=target_until,
                            cameras_allowed=cameras,
                            invoice_number=str(inv_no),
                            payment_amount=row["Amount (€)"],
                        )

                        renewal_links.append({
                            "project":     proj,
                            "valid_until": target_until,
                            "cameras":     cameras,
                            "token":       token,
                        })
                    except Exception as renewal_exc:
                        if (
                            _is_missing_supabase_table_error(renewal_exc, "subscriptions")
                            or _is_missing_supabase_table_error(renewal_exc, "renewal_links")
                        ):
                            renewal_warnings.append(
                                f"{proj}: payment recorded, but renewal tables are not set up in Supabase yet."
                            )
                        else:
                            raise renewal_exc

                except Exception as exc:
                    errors.append(f"{proj}: {exc}")

            if errors:
                st.error("Some updates failed:\n" + "\n".join(errors))
            else:
                if renewal_warnings:
                    st.warning("\n".join(renewal_warnings))

                payment_entry = {
                    "payment_date": confirm_date.isoformat(),
                    "invoice_number": inv_no,
                    "source_name": payment_context.get("source_name") if payment_context else f"manual-invoice-{inv_no}",
                    "source_kind": payment_context.get("source_kind") if payment_context else "manual",
                    "payment_fingerprint": payment_context.get("payment_fingerprint") if payment_context else hashlib.sha256(
                        f"manual|{inv_no}|{confirm_date.isoformat()}|{','.join(str(r['invoice_row_id']) for r in allocation_rows)}|{total_applied:.2f}".encode("utf-8")
                    ).hexdigest(),
                    "instructed_amount": payment_context.get("instructed_amount") if payment_context else None,
                    "received_amount": payment_context.get("received_amount") if payment_context else None,
                    "applied_amount": total_applied,
                    "fee_amount": payment_context.get("fee_amount") if payment_context else None,
                    "currency": payment_context.get("currency") if payment_context else "EUR",
                    "raw_text": payment_context.get("raw_text") if payment_context else None,
                    "parsed_payload": payment_context.get("parsed_payload") if payment_context else {},
                    "notes": payment_context.get("notes") if payment_context else None,
                }

                try:
                    append_bank_payment_with_allocations(payment_entry, allocation_rows)
                except Exception as payment_exc:
                    append_bank_payment_log({**payment_entry, "allocations": allocation_rows})
                    st.warning(f"Payment was saved locally because the bank payment table is not ready yet: {payment_exc}")

                st.cache_data.clear()
                st.session_state["_renewal_result"] = {
                    "inv_no": inv_no,
                    "count":  len(selected),
                    "links":  renewal_links,
                }
                st.rerun()

    # ── PDF upload section ────────────────────────────────────────────────────
    st.markdown("Upload SWIFT MT103 bank transfer PDF(s) to auto-extract payment details.")
    uploaded_files = st.file_uploader("Upload bank transfer PDF(s)", type=["pdf"], accept_multiple_files=True)

    parsed_bank_files = []
    if uploaded_files:
        for index, uploaded in enumerate(uploaded_files, start=1):
            file_bytes = uploaded.read()
            uploaded_hash = hashlib.sha256(file_bytes).hexdigest()
            with st.expander(f"📄 {index}. {uploaded.name}", expanded=index == 1):
                with st.spinner("Parsing PDF…"):
                    try:
                        parsed = parse_swift_pdf(file_bytes)
                    except Exception as exc:
                        st.error(f"Failed to parse PDF: {exc}")
                        continue

                st.caption("All fields are editable — correct any parsing errors before proceeding.")

                ec1, ec2, ec3 = st.columns(3)
                inv_no = ec1.number_input(
                    "Invoice #",
                    value=int(parsed["invoice_number"]) if parsed["invoice_number"] else 0,
                    min_value=0,
                    step=1,
                    key=f"pdf_inv_no_{index}",
                )
                pay_date = ec2.date_input(
                    "Payment Date",
                    value=parsed["payment_date"] if parsed["payment_date"] else datetime.date.today(),
                    key=f"pdf_pay_date_{index}",
                )
                ec3.metric(
                    "Instructed Amount",
                    f"€{parsed['instructed_amount']:,.2f}" if parsed["instructed_amount"] else "—",
                )

                if parsed["received_amount"] and parsed["instructed_amount"]:
                    fee = parsed["instructed_amount"] - parsed["received_amount"]
                    if fee > 0:
                        st.info(
                            f"Bank fee deducted: €{fee:,.2f}  "
                            f"(instructed €{parsed['instructed_amount']:,.2f} → "
                            f"received €{parsed['received_amount']:,.2f}). "
                            "The invoiced amount will be kept as-is."
                        )

                if not parsed["invoice_number"] and not parsed["payment_date"]:
                    st.warning(
                        "Could not extract payment data from this PDF. "
                        "Use the manual lookup below instead."
                    )

                st.markdown("---")
                st.markdown("### Matching Invoice Rows")
                _render_invoice_lookup(
                    inv_no,
                    pay_date,
                    key_prefix=f"pdf_{index}",
                    payment_context={
                        "source_name": uploaded.name,
                        "source_kind": "pdf",
                        "payment_fingerprint": uploaded_hash,
                        "instructed_amount": parsed.get("instructed_amount"),
                        "received_amount": parsed.get("received_amount"),
                        "fee_amount": (
                            parsed.get("instructed_amount") - parsed.get("received_amount")
                            if parsed.get("instructed_amount") is not None and parsed.get("received_amount") is not None
                            else None
                        ),
                        "currency": "EUR",
                        "raw_text": parsed.get("raw_text"),
                        "parsed_payload": parsed,
                    },
                )

            parsed_bank_files.append({
                "index": index,
                "name": uploaded.name,
                "hash": uploaded_hash,
                "file_bytes": file_bytes,
                "parsed": parsed,
                "inv_no_key": f"pdf_inv_no_{index}",
                "pay_date_key": f"pdf_pay_date_{index}",
            })

        st.markdown("---")
        st.markdown("### 📋 Parsed Files Summary")
        summary_rows = []
        for item in parsed_bank_files:
            parsed = item["parsed"]
            current_inv = _safe_int(st.session_state.get(item["inv_no_key"], parsed.get("invoice_number") or 0), default=0)
            current_date = st.session_state.get(item["pay_date_key"], parsed.get("payment_date"))
            instructed = parsed.get("instructed_amount") or 0.0
            received = parsed.get("received_amount") or 0.0
            fee = (instructed - received) if (instructed and received) else 0.0
            summary_rows.append({
                "#": item["index"],
                "File": item["name"],
                "Invoice #": current_inv if current_inv else "—",
                "Payment Date": str(current_date) if current_date else "—",
                "Instructed (€)": float(instructed) if instructed else 0.0,
                "Received (€)": float(received) if received else 0.0,
                "Fee (€)": float(fee) if fee else 0.0,
            })
        st.dataframe(
            summary_rows,
            use_container_width=True,
            hide_index=True,
        )

        if st.button("✅ Save All Parsed Payments", type="primary", key="save_all_parsed_payments"):
            batch_saved = 0
            batch_skipped = []
            batch_errors = []
            batch_saved_records = []

            for item in parsed_bank_files:
                try:
                    item_inv_no = _safe_int(st.session_state.get(item["inv_no_key"], 0), default=0)
                    item_pay_date = st.session_state.get(item["pay_date_key"], item["parsed"].get("payment_date") or datetime.date.today())
                    item_storage_meta: dict = {}
                    try:
                        item_storage_meta = upload_bank_payment_pdf_supabase(item["file_bytes"], item["name"]) or {}
                    except Exception as exc:
                        logger.warning("Could not upload bank payment PDF to Supabase Storage: %s", exc)
                    item_payment_context = {
                        "source_name": item["name"],
                        "source_kind": "pdf-batch",
                        "payment_fingerprint": item["hash"],
                        "instructed_amount": item["parsed"].get("instructed_amount"),
                        "received_amount": item["parsed"].get("received_amount"),
                        "fee_amount": (
                            item["parsed"].get("instructed_amount") - item["parsed"].get("received_amount")
                            if item["parsed"].get("instructed_amount") is not None and item["parsed"].get("received_amount") is not None
                            else None
                        ),
                        "currency": "EUR",
                        "raw_text": item["parsed"].get("raw_text"),
                        "parsed_payload": item["parsed"],
                    }

                    if item_inv_no <= 0:
                        # Keep the payment record even when no invoice number is available.
                        payment_entry = {
                            "payment_date": item_pay_date.isoformat() if hasattr(item_pay_date, "isoformat") else str(item_pay_date),
                            "invoice_number": None,
                            "source_name": item["name"],
                            "source_kind": "pdf-batch",
                            "payment_fingerprint": item["hash"],
                            "instructed_amount": item_payment_context["instructed_amount"],
                            "received_amount": item_payment_context["received_amount"],
                            "applied_amount": None,
                            "fee_amount": item_payment_context["fee_amount"],
                            "currency": "EUR",
                            "raw_text": item_payment_context["raw_text"],
                            "parsed_payload": item_payment_context["parsed_payload"],
                            "notes": "Auto-saved from batch upload without invoice number.",
                        }
                        if item_storage_meta:
                            payment_entry["pdf_storage_bucket"] = item_storage_meta.get("pdf_storage_bucket")
                            payment_entry["pdf_storage_path"] = item_storage_meta.get("pdf_storage_path")
                        try:
                            append_bank_payment_with_allocations(payment_entry, [])
                        except Exception:
                            append_bank_payment_log({**payment_entry, "allocations": []})
                        batch_saved += 1
                        batch_saved_records.append(payment_entry)
                        batch_skipped.append(f"{item['name']}: saved without invoice match")
                        continue

                    rows = get_invoices_by_number(item_inv_no)
                    if not rows:
                        payment_entry = {
                            "payment_date": item_pay_date.isoformat() if hasattr(item_pay_date, "isoformat") else str(item_pay_date),
                            "invoice_number": item_inv_no,
                            "source_name": item["name"],
                            "source_kind": "pdf-batch",
                            "payment_fingerprint": item["hash"],
                            "instructed_amount": item_payment_context["instructed_amount"],
                            "received_amount": item_payment_context["received_amount"],
                            "applied_amount": None,
                            "fee_amount": item_payment_context["fee_amount"],
                            "currency": "EUR",
                            "raw_text": item_payment_context["raw_text"],
                            "parsed_payload": item_payment_context["parsed_payload"],
                            "notes": "Auto-saved from batch upload but no invoice rows were found.",
                        }
                        if item_storage_meta:
                            payment_entry["pdf_storage_bucket"] = item_storage_meta.get("pdf_storage_bucket")
                            payment_entry["pdf_storage_path"] = item_storage_meta.get("pdf_storage_path")
                        try:
                            append_bank_payment_with_allocations(payment_entry, [])
                        except Exception:
                            append_bank_payment_log({**payment_entry, "allocations": []})
                        batch_saved += 1
                        batch_saved_records.append(payment_entry)
                        batch_skipped.append(f"{item['name']}: no invoice rows found")
                        continue

                    unpaid_rows = [row for row in rows if str(row.get("paid", "No")).strip().lower() != "yes"]
                    if not unpaid_rows:
                        payment_entry = {
                            "payment_date": item_pay_date.isoformat() if hasattr(item_pay_date, "isoformat") else str(item_pay_date),
                            "invoice_number": item_inv_no,
                            "source_name": item["name"],
                            "source_kind": "pdf-batch",
                            "payment_fingerprint": item["hash"],
                            "instructed_amount": item_payment_context["instructed_amount"],
                            "received_amount": item_payment_context["received_amount"],
                            "applied_amount": None,
                            "fee_amount": item_payment_context["fee_amount"],
                            "currency": "EUR",
                            "raw_text": item_payment_context["raw_text"],
                            "parsed_payload": item_payment_context["parsed_payload"],
                            "notes": "Auto-saved from batch upload; invoice rows were already paid.",
                        }
                        if item_storage_meta:
                            payment_entry["pdf_storage_bucket"] = item_storage_meta.get("pdf_storage_bucket")
                            payment_entry["pdf_storage_path"] = item_storage_meta.get("pdf_storage_path")
                        try:
                            append_bank_payment_with_allocations(payment_entry, [])
                        except Exception:
                            append_bank_payment_log({**payment_entry, "allocations": []})
                        batch_saved += 1
                        batch_saved_records.append(payment_entry)
                        batch_skipped.append(f"{item['name']}: rows already paid")
                        continue

                    # Auto-mark every unpaid row for this invoice as paid.
                    allocation_rows = []
                    total_applied = 0.0
                    renewal_links = []
                    renewal_warnings = []
                    errors = []

                    for row in unpaid_rows:
                        proj = row["project_name"]
                        try:
                            mark_invoice_row_paid(
                                db_id=int(row["id"]),
                                payment_date=item_pay_date,
                                payment_amount=row.get("payment_amount"),
                            )

                            amount_applied = _safe_float(row.get("payment_amount"))
                            total_applied += amount_applied
                            allocation_rows.append({
                                "invoice_row_id": int(row["id"]),
                                "invoice_number": item_inv_no,
                                "project_name": proj,
                                "maintenance_year": _safe_str(row.get("maintenance_year")),
                                "year": _safe_int(row.get("year"), default=0) or None,
                                "amount_applied": amount_applied,
                            })

                            try:
                                sub = get_subscription(proj)
                                if sub and sub.get("valid_until"):
                                    current_until = datetime.date.fromisoformat(sub["valid_until"][:10])
                                    base = max(current_until, item_pay_date)
                                else:
                                    base = item_pay_date
                                try:
                                    target_until = base.replace(year=base.year + 1)
                                except ValueError:
                                    target_until = base.replace(year=base.year + 1, day=28)

                                cameras = _safe_int(row.get("cameras_number"))
                                upsert_subscription(
                                    project_name=proj,
                                    valid_until=target_until,
                                    cameras_allowed=cameras,
                                    valid_from=item_pay_date,
                                )
                                token = create_renewal_link(
                                    project_name=proj,
                                    target_valid_until=target_until,
                                    cameras_allowed=cameras,
                                    invoice_number=str(item_inv_no),
                                    payment_amount=row.get("payment_amount"),
                                )
                                renewal_links.append({
                                    "project": proj,
                                    "valid_until": target_until,
                                    "cameras": cameras,
                                    "token": token,
                                })
                            except Exception as renewal_exc:
                                if (
                                    _is_missing_supabase_table_error(renewal_exc, "subscriptions")
                                    or _is_missing_supabase_table_error(renewal_exc, "renewal_links")
                                ):
                                    renewal_warnings.append(f"{proj}: renewal tables are not set up yet.")
                                else:
                                    raise renewal_exc
                        except Exception as exc:
                            errors.append(f"{proj}: {exc}")

                    payment_entry = {
                        "payment_date": item_pay_date.isoformat() if hasattr(item_pay_date, "isoformat") else str(item_pay_date),
                        "invoice_number": item_inv_no,
                        "source_name": item["name"],
                        "source_kind": "pdf-batch",
                        "payment_fingerprint": item["hash"],
                        "instructed_amount": item_payment_context["instructed_amount"],
                        "received_amount": item_payment_context["received_amount"],
                        "applied_amount": total_applied,
                        "fee_amount": item_payment_context["fee_amount"],
                        "currency": "EUR",
                        "raw_text": item_payment_context["raw_text"],
                        "parsed_payload": item_payment_context["parsed_payload"],
                        "notes": item_payment_context.get("notes"),
                    }

                    if item_storage_meta:
                        payment_entry["pdf_storage_bucket"] = item_storage_meta.get("pdf_storage_bucket")
                        payment_entry["pdf_storage_path"] = item_storage_meta.get("pdf_storage_path")

                    if errors:
                        batch_skipped.append(f"{item['name']}: " + "; ".join(errors))
                    if renewal_warnings:
                        batch_skipped.append(f"{item['name']}: " + "; ".join(renewal_warnings))

                    try:
                        append_bank_payment_with_allocations(payment_entry, allocation_rows)
                    except Exception as payment_exc:
                        append_bank_payment_log({**payment_entry, "allocations": allocation_rows})
                        batch_skipped.append(f"{item['name']}: saved locally because the bank payment table is not ready yet ({payment_exc})")

                    batch_saved += 1
                    batch_saved_records.append(payment_entry)
                except Exception as exc:
                    batch_skipped.append(f"{item['name']}: {exc}")

            if batch_saved_records:
                st.cache_data.clear()
            if batch_saved:
                st.success(f"Saved {batch_saved} bank payment(s).")
            if batch_skipped:
                st.warning("\n".join(batch_skipped))
            st.rerun()

    # ── Manual lookup (no PDF) ────────────────────────────────────────────────
    st.markdown("---")
    with st.expander("🔍 Manual Lookup (no PDF)", expanded=not bool(uploaded_files)):
        mc1, mc2 = st.columns(2)
        manual_inv_no = mc1.number_input(
            "Invoice #", min_value=0, step=1, key="manual_inv_no",
        )
        manual_date = mc2.date_input(
            "Payment Date", value=datetime.date.today(), key="manual_date",
        )
        _render_invoice_lookup(manual_inv_no, manual_date, key_prefix="manual")

    st.markdown("---")
    st.subheader("Saved Bank Payments")
    bank_payment_rows = load_bank_payments_log()
    if bank_payment_rows:
        bank_pdf_links: dict[str, str | None] = {}
        for row in bank_payment_rows:
            storage_bucket = _safe_str(row.get("pdf_storage_bucket")).strip()
            storage_path = _safe_str(row.get("pdf_storage_path")).strip()
            fp_key = _safe_str(row.get("payment_fingerprint"))
            link: str | None = None
            if storage_bucket and storage_path:
                link = create_bank_payment_pdf_signed_url_supabase(storage_bucket, storage_path)
            bank_pdf_links[fp_key] = link

        # Build invoice# -> sent monthly invoice PDF signed URL map
        sent_invoice_rows = load_sent_invoices_log() or []
        invoice_pdf_links: dict[int, str] = {}
        for sent_row in sent_invoice_rows:
            inv_num = _safe_int(sent_row.get("invoice_number"), default=0)
            if not inv_num or inv_num in invoice_pdf_links:
                continue
            sent_bucket = _safe_str(sent_row.get("pdf_storage_bucket")).strip()
            sent_path = _safe_str(sent_row.get("pdf_storage_path")).strip()
            if not (sent_bucket and sent_path):
                continue
            try:
                signed = create_sent_invoice_pdf_signed_url_supabase(sent_bucket, sent_path)
            except Exception:
                signed = None
            if signed:
                invoice_pdf_links[inv_num] = signed

        # Build payment_id -> invoice_number map from allocations (covers
        # rows where the bank_payments.invoice_number is 0 because the
        # SWIFT parser didn't extract it but the user matched rows manually).
        payment_invoice_map: dict[int, int] = {}
        try:
            all_allocations = load_bank_payment_allocations()
        except Exception:
            all_allocations = []
        for alloc in all_allocations or []:
            pid = alloc.get("payment_id")
            inv = _safe_int(alloc.get("invoice_number"), default=0)
            if pid is None or not inv:
                continue
            try:
                pid_int = int(pid)
            except Exception:
                continue
            payment_invoice_map.setdefault(pid_int, inv)

        def _row_invoice_number(row: dict) -> int:
            inv = _safe_int(row.get("invoice_number"), default=0)
            if inv:
                return inv
            try:
                pid = int(row.get("id")) if row.get("id") is not None else None
            except Exception:
                pid = None
            if pid is not None:
                inv = payment_invoice_map.get(pid, 0)
                if inv:
                    return inv
            for alloc in row.get("allocations", []) or []:
                inv = _safe_int(alloc.get("invoice_number"), default=0)
                if inv:
                    return inv
            return 0

        def _row_invoice_pdf_link(row: dict) -> str:
            inv = _row_invoice_number(row)
            return invoice_pdf_links.get(inv, "") if inv else ""

        payment_df = pd.DataFrame([
            {
                "Saved At": _safe_str(row.get("created_at") or row.get("updated_at") or "")[:19].replace("T", " "),
                "Payment Date": _safe_str(row.get("payment_date")),
                "Invoice #": _row_invoice_number(row),
                "Source": _safe_str(row.get("source_name")),
                "Kind": _safe_str(row.get("source_kind")),
                "Applied (€)": f"€{_safe_float(row.get('applied_amount', 0.0)):,.0f}",
                "Fee (€)": f"€{_safe_float(row.get('fee_amount', 0.0)):,.0f}" if row.get("fee_amount") not in (None, "") else "",
                "Fingerprint": _safe_str(row.get("payment_fingerprint"))[:12],
                "Source PDF": bank_pdf_links.get(_safe_str(row.get("payment_fingerprint"))) or "",
                "Invoice PDF": _row_invoice_pdf_link(row),
            }
            for row in bank_payment_rows
        ])
        st.dataframe(
            payment_df,
            use_container_width=True,
            hide_index=True,
            height=260,
            column_config={
                "Source PDF": st.column_config.LinkColumn(
                    "Source PDF",
                    help="Click to open the originally uploaded bank transfer PDF.",
                    display_text="Download",
                ),
                "Invoice PDF": st.column_config.LinkColumn(
                    "Invoice PDF",
                    help="Click to open the monthly invoice PDF that was paid by this bank transfer.",
                    display_text="Download",
                ),
            },
        )

        payment_labels = [
            f"#{_safe_int(row.get('invoice_number'), default=0)} | {_safe_str(row.get('payment_date'))} | {_safe_str(row.get('source_name'))}"
            for row in bank_payment_rows
        ]
        selected_payment_label = st.selectbox("Payment details", payment_labels, key="saved_bank_payment")
        selected_payment_row = bank_payment_rows[payment_labels.index(selected_payment_label)]
        payment_id = selected_payment_row.get("id")
        allocations = []
        if payment_id is not None:
            allocations = load_bank_payment_allocations(int(payment_id))
        if not allocations:
            allocations = selected_payment_row.get("allocations", []) or []
        if allocations:
            alloc_df = pd.DataFrame([
                {
                    "Project": _safe_str(a.get("project_name")),
                    "Invoice #": _safe_int(a.get("invoice_number"), default=0),
                    "Maint. Year": _safe_str(a.get("maintenance_year")),
                    "Amount (€)": f"€{_safe_float(a.get('amount_applied', 0.0)):,.0f}",
                }
                for a in allocations
            ])
            st.dataframe(alloc_df, use_container_width=True, hide_index=True)
    else:
        st.info("No bank payments have been saved yet.")
