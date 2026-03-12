"""One-time script to remove duplicate invoice rows added by the save bug.

Rows WITHOUT an invoice number are deduplicated by (Project name, Maintenance Year, Year).
Only the FIRST occurrence is kept.

Run: python dedup_invoices.py
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

import openpyxl

DATA_FILE = Path(__file__).parent / "data" / "CaddyCheckProjectsInfo.xlsx"

wb = openpyxl.load_workbook(DATA_FILE)
ws = wb["Invoice details"]

rows = list(ws.iter_rows(values_only=True))
if not rows:
    print("Sheet is empty.")
    sys.exit(0)

header = rows[0]
print(f"Header: {header}")
print(f"Total rows (incl. header): {len(rows)}")

# Find column indices (0-based within the row tuple)
col_map = {str(h).strip().lower() if h else "": i for i, h in enumerate(header)}
print(f"Columns: {col_map}")

# Identify key columns (adjust names to match actual headers)
# Expected: Invoice Number, Project name, Maintenance Year, Year
inv_col   = next((i for k, i in col_map.items() if "invoice" in k and "number" in k), None)
proj_col  = next((i for k, i in col_map.items() if "project" in k), None)
my_col    = next((i for k, i in col_map.items() if "maintenance" in k), None)
year_col  = next((i for k, i in col_map.items() if k == "year"), None)

print(f"Invoice# col: {inv_col}, Project col: {proj_col}, Maint. Year col: {my_col}, Year col: {year_col}")

seen_with_inv = set()       # invoice numbers seen
seen_no_inv = set()         # (project, maint_year, year) tuples for no-invoice rows
kept = [header]
dups = 0

for row in rows[1:]:
    inv_no = row[inv_col] if inv_col is not None else None
    project = str(row[proj_col]).strip().lower() if proj_col is not None and row[proj_col] else ""
    maint_yr = str(row[my_col]).strip() if my_col is not None and row[my_col] else ""
    yr = str(row[year_col]).strip() if year_col is not None and row[year_col] else ""

    # Skip formula strings like "=A501+1" — treat as no invoice number
    if inv_no and not str(inv_no).startswith("="):
        try:
            key = (float(inv_no), project)  # monthly invoices share invoice# across projects
        except (ValueError, TypeError):
            key = None
    else:
        key = None

    if key is not None:
        if key in seen_with_inv:
            dups += 1
            continue
        seen_with_inv.add(key)
    else:
        key = None  # already set above, just for clarity
        key = (project, maint_yr, yr)
        if key in seen_no_inv:
            dups += 1
            continue
        seen_no_inv.add(key)

    kept.append(row)

print(f"Duplicates removed: {dups}")
print(f"Rows kept: {len(kept) - 1} data rows")

if dups == 0:
    print("No duplicates found — file unchanged.")
    sys.exit(0)

# Rewrite the sheet
for row_idx in range(ws.max_row, 0, -1):
    ws.delete_rows(row_idx)

for row_data in kept:
    ws.append(list(row_data))

wb.save(DATA_FILE)
print(f"Saved: {DATA_FILE}")
